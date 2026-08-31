"""
VCF Compound Layer — Step 10: reliability anchors for the new VCF corpus.

Placement note: the spec says "add to the existing harness in
src/lib/reliability/" — that directory is 100% live-API/live-LLM checks
against the flavor_network/RecipeNLG/Tradition-DB frontend lenses (see
checks.js: every function takes a live `api` object or calls a live LLM).
None of these 9 VCF anchors need or have that — the VCF pipeline isn't
wired into the frontend or artifact API yet (spec's own status line:
"Steps 1-8 are self-contained and testable without touching the
frontend"), and every one of these checks reads the pipeline's own JSONL
artifacts directly. This repo already has exactly that kind of test for
the pre-existing flavor_network dataset — pipeline/tests/test_reliability.py
— so this file is its sibling, not a forced fit into the JS harness.
Flagged to James rather than silently reinterpreted; see the Step 10 report.

No live API or LLM required — every check reads the pipeline's own frozen
JSONL artifacts directly, the same pattern test_reliability.py already
uses for the pre-existing flavor_network dataset (corpus_tables/
compound_tables fixtures skip loudly if artifacts are missing; this file
does the same for the VCF artifacts).

Covers the spec's 9 Step 10 bullets:
  1. Coffee's top matches are roasted products, not beverages.
  2. Roasted peanut and cocoa liquor share pyrazines.
  3. Mushroom's profile contains 1-octen-3-one.
  4. No product occupies more than 40% of top-10 slots.
  5. Every multi-state base ingredient produces a non-empty gained/lost
     diff at df >= 3.
  6. Compound counts per product match the frozen artifact version.
  7. No compound appears under two compound_id values after Step 3b
     canonicalisation.
  8. Every product profile carries a single profile_source.
  9. Descriptor coverage on top-IDF explanatory compounds is reported per
     source, not assumed.

Plus Step 11's 5 anchors (added 2026-08-29, same run that built Step 11
itself — see build_vcf_phase.py / build_vcf_phase_frames.py /
build_vcf_competition.py):
  10. Hydrocarbons' median XLogP is the highest of the 18 groups.
  11. Bases and Furans fall below Hydrocarbons, Esters, and Phenols.
  12. Every compound with a CAS in the crosswalk either has an xlogp
      value or is explicitly marked null — no silent zeros.
  13. No phase-behaviour sentence appears in output that is not present
      in phase_frames.jsonl.
  14. Volatility claims are suppressed where boiling-point coverage is
      below threshold.
These use a SEPARATE fixture (vcf_phase_artifacts) that loud-skips only
itself if phase_frames.jsonl/competition.jsonl are missing, rather than
folding into the original vcf_artifacts fixture — that keeps tests 1-9
running (and loud-skipping only on their own 6 original files) even in a
checkout that predates Step 11, instead of newly breaking on files that
didn't exist when they were written.

REVISION 4 (final revision on the smoke/phenol frame — see
build_vcf_phase_frames.py and build_vcf_competition.py's docstrings for
the full account): the flagship frame is renamed
fat_phase_smoke_terpene_competition -> fat_phase_phenol_terpene_carrier
and its trigger no longer references smoke_marker/role_counts/
role_shares at all — it's a pure group_pair + percentile behaviour claim.
Tests 15 and 26 are rewritten accordingly (they previously reconstructed
a_smoke_marker_count/share fields that no longer exist on competition
rows). Test 22's smoke-marker-literal-threshold check is retired for the
same reason (replaced by test 27, which checks the NEW conflict_subtype
mechanism's role-percentile gate instead). New anchors added for
Revision 4's other two changes: no frame sentence asserts a
preparation/provenance claim unless its trigger actually reads the
preparation/cure_state field (test 28), and the new smoked_product_fat_phase
stub stays dormant — pending_authoring, empty sentence, never attached to
any row — regardless of its trigger firing (test 29).

BEEF INGESTION BUILD SPEC (2026-08-29): the pipeline's first non-VCF
source. Tests 30-34 cover its six required anchors (one of the six —
"no profile mixes sources" — is folded into test 8's rewrite, since it's
the same concern test 8 already owned). See ingest_protein_beef.py's
module docstring for the two spec premises that turned out to be false
against the real workbook, and the section header above test 30 for the
detail on each anchor.
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import pytest

VCF_DIR = Path(__file__).resolve().parents[1] / "artifacts" / "vcf"
SCRIPTS_DIR = Path(__file__).resolve().parents[1] / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from vcf_trigger_lib import evaluate_trigger, generate_trigger_description  # noqa: E402
from ingest_protein_beef import (  # noqa: E402
    normalize_name,
    loose_key,
    ROUTING_XLSX,
    ROUTING_SHEET,
    VALID_MR17_ROUTING_STATES,
)


def _loud_skip(missing: list[Path]):
    names = ", ".join(p.name for p in missing)
    pytest.skip(
        f"VCF ARTIFACTS MISSING: {names} — run the pipeline/scripts/build_vcf_*.py "
        f"chain first. This is a LOUD skip: it must show up in the run summary as "
        f"'skipped', and a CI config for this suite must fail (or print an "
        f"unmissable summary) if any test here is skipped rather than run — "
        f"per Step 10, a green run must not be able to happen for the wrong reason."
    )


def _load_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text().splitlines() if line.strip()]


REQUIRED = ["pairs.jsonl", "profiles.jsonl", "compounds.jsonl", "form_diffs.jsonl",
            "spine.jsonl", "meta.json"]


@pytest.fixture(scope="module")
def vcf_artifacts():
    missing = [VCF_DIR / name for name in REQUIRED if not (VCF_DIR / name).exists()]
    if missing:
        _loud_skip(missing)
    return {
        "pairs": _load_jsonl(VCF_DIR / "pairs.jsonl"),
        "profiles": _load_jsonl(VCF_DIR / "profiles.jsonl"),
        "compounds": _load_jsonl(VCF_DIR / "compounds.jsonl"),
        "form_diffs": _load_jsonl(VCF_DIR / "form_diffs.jsonl"),
        "spine": _load_jsonl(VCF_DIR / "spine.jsonl"),
        "meta": json.loads((VCF_DIR / "meta.json").read_text()),
    }


# --- 1. Coffee's top matches are roasted products, not beverages -----------

# Substring match, not literal equality: raw_name spellings can carry
# incidental qualifiers (binomial, "generic", cultivar) that a brittle
# literal check would break on for no reason connected to what this test
# is actually protecting against.
COFFEE_EXPECTED_IN_TOP10 = [
    "hazelnut",   # FILBERT, HAZELNUT (roasted)
    "peanut",     # PEANUT (roasted)
    "cocoa liquor",
    "roasted cocoa beans",
    "popcorn",
]
# The specific failure mode Step 5 named by name: unweighted (raw
# shared-count) ranking returns big, generic beverage profiles because
# they're large, not because they're culinarily related to coffee.
BEVERAGE_CONTAMINATION_TOKENS = ["beer", "black tea", "white wine"]


def test_coffee_top_matches_are_roasted_not_beverages(vcf_artifacts):
    pairs = vcf_artifacts["pairs"]
    coffee_rows = sorted(
        (p for p in pairs if p["anchor_raw_name"] == "COFFEE"),
        key=lambda p: p["rank"],
    )
    assert coffee_rows, "no pairs anchored on COFFEE — has the anchor raw_name changed?"
    top10_names = [r["match_raw_name"].lower() for r in coffee_rows[:10]]

    missing = [
        expected for expected in COFFEE_EXPECTED_IN_TOP10
        if not any(expected in name for name in top10_names)
    ]
    assert not missing, (
        f"coffee's top-10 is missing expected roasted/aroma-related matches: "
        f"{missing} — top10 was {top10_names}"
    )

    contaminated = [
        name for name in top10_names
        if any(tok in name for tok in BEVERAGE_CONTAMINATION_TOKENS)
    ]
    assert not contaminated, (
        f"coffee's top-10 contains beverage contamination {contaminated} — "
        f"this is the exact unweighted raw-shared-count failure signature "
        f"Step 5 named (beer/black tea/white wine dominating instead of "
        f"IDF-weighted culinary matches)"
    )


# --- 2. Roasted peanut and cocoa liquor share pyrazines ---------------------

def test_roasted_peanut_cocoa_liquor_share_pyrazines(vcf_artifacts):
    compounds_by_id = {c["compound_id"]: c for c in vcf_artifacts["compounds"]}
    profiles_by_name = {p["raw_name"]: p for p in vcf_artifacts["profiles"]}

    peanut = profiles_by_name.get("PEANUT (roasted)")
    cocoa = profiles_by_name.get("COCOA LIQUOR")
    assert peanut and cocoa, "PEANUT (roasted) or COCOA LIQUOR raw_name not found in profiles"

    shared = set(peanut["compound_ids"]) & set(cocoa["compound_ids"])
    shared_pyrazines = [
        compounds_by_id[c]["raw_compound"] for c in shared
        if "pyrazine" in compounds_by_id[c]["raw_compound"].lower()
    ]
    # Checked directly against compound names, not VCF's own "Bases" group
    # label — VCF files pyrazines under "Bases" (an artifact of its own
    # taxonomy), so asserting on the group label would silently pass if a
    # future re-pull relabeled the group but broke the actual pyrazine
    # overlap, or vice versa.
    assert len(shared_pyrazines) >= 5, (
        f"only {len(shared_pyrazines)} shared pyrazine compounds between "
        f"roasted peanut and cocoa liquor (of {len(shared)} total shared "
        f"compounds) — expected the well-known Maillard pyrazine signature"
    )


# --- 3. Mushroom's profile contains 1-octen-3-one ---------------------------

def test_mushroom_contains_1_octen_3_one(vcf_artifacts):
    compounds_by_id = {c["compound_id"]: c for c in vcf_artifacts["compounds"]}
    mushroom_profiles = [
        p for p in vcf_artifacts["profiles"] if p["raw_name"].upper().startswith("MUSHROOM")
    ]
    assert mushroom_profiles, "no product with raw_name starting 'MUSHROOM' found"
    for m in mushroom_profiles:
        names = [compounds_by_id[c]["raw_compound"].lower() for c in m["compound_ids"]]
        assert any("1-octen-3-one" in n for n in names), (
            f"{m['raw_name']!r} profile is missing 1-octen-3-one "
            f"(mushroom's signature volatile)"
        )


# --- 4. No product occupies more than 40% of top-10 slots -------------------

def test_no_product_exceeds_hub_cap(vcf_artifacts):
    pairs = vcf_artifacts["pairs"]
    n_anchors = len({p["anchor_vcf_product_id"] for p in pairs})
    assert n_anchors > 0
    top10 = [p for p in pairs if p["rank"] <= 10]
    occupancy = Counter(p["match_vcf_product_id"] for p in top10)
    worst_id, worst_count = occupancy.most_common(1)[0]
    worst_fraction = worst_count / n_anchors
    assert worst_fraction <= 0.40, (
        f"product id {worst_id} occupies {worst_fraction:.1%} of all top-10 "
        f"slots ({worst_count}/{n_anchors}) — exceeds the 40% hub cap, "
        f"meaning the IDF weighting is not controlling for profile size"
    )


# --- 5. Every multi-state base ingredient produces a non-empty diff --------

def test_form_diffs_nonempty_for_every_multistate_base(vcf_artifacts):
    spine = vcf_artifacts["spine"]
    form_diffs = vcf_artifacts["form_diffs"]

    # Build 1 (Sequenced Builds spec, spine near-duplicate clustering,
    # 2026-08-30) introduced a SECOND kind of multi-member spine entry:
    # apply_spine_clusters.py merges genuinely near-duplicate products
    # (>=85% compound overlap, genus- or containment-matched) onto one
    # spine_id so pairs.jsonl's same-spine exclusion catches them — a
    # different job from build_vcf_spine.py's original base_ingredient
    # grouping, which only ever grouped genuinely distinct preparation
    # states of one food (PEANUT raw/roasted) where a real gained/lost
    # difference is expected practically by construction. A clustered
    # near-duplicate pair can legitimately clear the clustering gate
    # (>=85% shared) and still produce ZERO signal at form_diffs' own
    # df_culinary>=3 floor (AGASTACHE SPECIES / ANISE HYSSOP, merged into
    # culin:anise_hyssop) — that isn't a floor-logic bug, it's what "these
    # two are near-duplicates" actually looks like from the Form lens's
    # side. apply_spine_clusters.py stamps `spine_cluster_id` on every
    # entry it merges, so those are excluded from this assertion; every
    # organically multi-member entry (this test's original, still-valid
    # target) is held to the same strict standard as before.
    clustered_spine_ids = {e["spine_id"] for e in spine if e.get("spine_cluster_id")}

    multi_state_spine_ids = {
        e["spine_id"] for e in spine
        if sum(1 for m in e["members"] if m["class"] == "culinary") >= 2
        and e["spine_id"] not in clustered_spine_ids
    }
    assert multi_state_spine_ids, "no multi-culinary-member spine entries found at all"

    by_spine = defaultdict(list)
    for row in form_diffs:
        by_spine[row["spine_id"]].append(row)

    empty_bases = []
    for spine_id in multi_state_spine_ids:
        rows = by_spine.get(spine_id, [])
        if not rows:
            empty_bases.append((spine_id, "no form_diffs rows at all"))
            continue
        # Beef Ingestion Build Spec Step 7 added a second diff shape
        # (diff_type == "aging_delta", value_a/value_b quantitative deltas,
        # no gained/lost keys at all) sharing form_diffs.jsonl with the
        # original gained/lost state_diff rows. Restrict this anchor to
        # state_diff rows explicitly rather than assume every row has
        # gained/lost — a bare KeyError here would be the wrong failure mode
        # for "a new diff shape was added on purpose."
        state_diff_rows = [r for r in rows if r.get("diff_type", "state_diff") == "state_diff"]
        has_signal = any(len(r["gained"]) > 0 or len(r["lost"]) > 0 for r in state_diff_rows)
        if not has_signal:
            empty_bases.append((spine_id, "every diff row is empty at the df floor"))

    assert not empty_bases, (
        f"{len(empty_bases)} multi-state base ingredient(s) produce no "
        f"gained/lost signal at all: {empty_bases}"
    )


# --- 6. Compound counts per product match the frozen artifact version -----

# Frozen 2026-08-29, updated same day for the Beef Ingestion Build Spec:
# beef adds 4 profiles (beef:muscle:raw/cooked/smoked, beef:fat:cooked) on
# top of the prior 573-VCF-profile baseline -> 577 profiles, N=525
# culinary. Hash is over the sorted (vcf_product_id, n_compounds) pairs
# from profiles.jsonl, sorted by str() since a beef product id is a
# descriptive string ("beef:muscle:raw") rather than VCF's int — this is a
# regression gate, not a correctness proof: a real, reviewed re-pull,
# parse fix, or (as here) a deliberate new-source ingestion SHOULD change
# this hash, and whoever changes it must update this constant deliberately
# (with a comment saying why), not silently.
# Updated 2026-08-29: MR-17/MR-18 (flavour-relevance classification) removed
# 5 excluded + 6 unresolved compounds from beef's raw/cooked/fat profiles
# (Furaneol, p-Cymene, and a pyrazine were reclassified INTO their profiles
# under real compound_groups instead — those 3 don't change n_compounds).
# The resolver-identity-collision fix also merged 2 previously-duplicated
# compounds ((E,E)-2,4-Heptadienal, Pentadecanal) into their pre-existing
# VCF identities. Deliberate, reviewed change — see meta.json's
# protein_beef_validation entry. Profile COUNT is unchanged (577); only
# per-product compound counts for the affected beef profiles moved.
#
# Updated 2026-08-30: egg ingestion adds 4 profiles
# (egg:whole_chicken:scrambled, egg:yolk_chicken:boiled,
# egg:yolk_duck_salted:salted, egg:yolk_duck_salted:roasted_salted) on top
# of the 577-profile beef baseline -> 581 profiles, N=529 culinary.
# Deliberate, reviewed change — see meta.json's protein_egg entry
# (MR-17 occurrence-gating fix + identity-fragmentation audit documented
# there). Recomputed from the live profiles.jsonl after the full
# downstream rebuild.
FROZEN_PROFILE_COMPOUND_COUNT_HASH = (
    "fd660bb47353e3f0f71e0b0a3ad95c0f5bffc98d4e7f3e55f24282c568cca88d"
)
FROZEN_PROFILE_COUNT = 581


def test_compound_counts_match_frozen_snapshot(vcf_artifacts):
    profiles = vcf_artifacts["profiles"]
    assert len(profiles) == FROZEN_PROFILE_COUNT, (
        f"{len(profiles)} profiles now vs {FROZEN_PROFILE_COUNT} frozen — "
        f"the product count itself changed; update FROZEN_PROFILE_COUNT and "
        f"FROZEN_PROFILE_COMPOUND_COUNT_HASH together, deliberately, with a "
        f"note on why (a re-pull, a parse fix — not silently)"
    )
    rows = sorted(((p["vcf_product_id"], p["n_compounds"]) for p in profiles), key=lambda t: str(t[0]))
    payload = json.dumps(rows, separators=(",", ":"))
    current_hash = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    assert current_hash == FROZEN_PROFILE_COMPOUND_COUNT_HASH, (
        "per-product compound counts drifted from the frozen artifact "
        "version — if this is an intentional re-pull or bugfix, recompute "
        "and update the frozen hash deliberately; if not, something "
        "upstream changed silently"
    )


# --- 7. No compound appears under two compound_id values -------------------

def test_no_compound_split_across_two_compound_ids(vcf_artifacts):
    compounds = vcf_artifacts["compounds"]

    cas_to_ids = defaultdict(set)
    id_to_cas = {}
    conflicts_cas_to_id = []
    conflicts_id_to_cas = []
    for c in compounds:
        cid, cas = c["compound_id"], c["cas"]
        if cas:
            cas_to_ids[cas].add(cid)
        if cid in id_to_cas and id_to_cas[cid] != cas:
            conflicts_id_to_cas.append((cid, id_to_cas[cid], cas))
        id_to_cas[cid] = cas

    conflicts_cas_to_id = {cas: ids for cas, ids in cas_to_ids.items() if len(ids) > 1}

    assert not conflicts_cas_to_id, (
        f"{len(conflicts_cas_to_id)} CAS number(s) map to more than one "
        f"compound_id after canonicalisation — the same real molecule is "
        f"being scored as two different compounds: {conflicts_cas_to_id}"
    )
    assert not conflicts_id_to_cas, (
        f"{len(conflicts_id_to_cas)} compound_id value(s) map to more than "
        f"one CAS number — two different molecules were merged under one "
        f"identity: {conflicts_id_to_cas}"
    )


# --- 7b. Standing anchor: no two compound_ids share a normalized name. -----
#
# James, 2026-08-29: test 7 above only cross-checks compounds that BOTH
# carry a CAS — it never caught (E,E)-2,4-Heptadienal / Pentadecanal
# because in each pair, only ONE side (the beef-minted row) has a CAS; the
# pre-existing VCF row is a name-only provisional id with cas=None, so
# the two never land in the same cas_to_ids bucket. This is the actual
# gap: BeefCompoundResolver's CAS-first priority order (ingest_protein_
# beef.py) minted a new identity whenever a source's CAS wasn't already
# known, without ever checking whether the NAME already resolved to an
# existing provisional id. The resolver itself is now fixed (see its
# step-2 branch); this test is the standing anchor so a future family
# supplying its own CAS values can't reintroduce the same bug.
#
# UPDATE (same day, MR-17/MR-18 rebuild): the 2 instances below WERE
# retroactively merged — the classification pass required a full rebuild
# of compounds.jsonl/profiles.jsonl anyway (ingest_protein_beef.py can't
# reclassify an already-"existing" compound the second time it runs), so
# the fixed resolver picked them up for free on that same rebuild. Both
# names now resolve to a single compound_id (confirmed: zero collisions
# in the current corpus). Left in this allowlist, inert, as a named
# record of the 2 real instances this bug produced — not because either
# is still live. A third instance would still be caught.
KNOWN_DEFERRED_NAME_COLLISIONS = {
    frozenset({"4313-03-5", "vcf:e_e_2_4_heptadienal"}),
    frozenset({"2765-11-9", "vcf:pentadecanal"}),
}


def test_no_two_compound_ids_share_a_normalized_name(vcf_artifacts):
    compounds = vcf_artifacts["compounds"]

    ids_by_norm_name = defaultdict(set)
    for c in compounds:
        ids_by_norm_name[normalize_name(c["raw_compound"])].add(c["compound_id"])

    unexpected_collisions = {
        name: ids
        for name, ids in ids_by_norm_name.items()
        if len(ids) > 1 and frozenset(ids) not in KNOWN_DEFERRED_NAME_COLLISIONS
    }

    assert not unexpected_collisions, (
        f"{len(unexpected_collisions)} normalized compound name(s) map to "
        f"more than one compound_id, beyond the 2 known-and-deferred "
        f"instances — a resolver is minting a duplicate identity for a "
        f"compound the corpus already has under a name-only provisional "
        f"id: {unexpected_collisions}"
    )


# --- 7c. Standing anchor: no two compound_ids share a LOOSE (space/hyphen- -
#         collapsed) normalized name. ----------------------------------------
#
# James, 2026-08-31: 7b's normalize_name() collapses repeated whitespace but
# never removes it, so it never caught beef:4_methyl_phenol duplicating the
# pre-existing p-cresol (106-44-5) — '4-methylphenol (=p-cresol)' normalizes
# to '4-methylphenol', '4-Methyl phenol' normalizes to '4-methyl phenol',
# one space apart, no collision detected. Auditing the full corpus
# vocabulary this way (not just the one reported case) found 5 more
# instances of the exact same shape, all beef-minted, all sitting in the
# corpus undetected until this audit: 2,5-dimethylpyrazine,
# 2-ethyl-5-methylpyrazine, 2-pentylfuran, 3-ethyl-2,5-dimethylpyrazine,
# trimethylpyrazine. The resolver now has a loose_key() fallback tier (see
# its docstring in ingest_protein_beef.py) so none of these can be minted
# again — this test is the standing anchor, same role as 7b, so a future
# family can't reintroduce the pattern a fourth time undetected.
#
# The 6 pairs below are NOT yet merged — that's a data decision pending
# James's row-by-row review (same gate as the CID-collision candidates and
# the Furaneol stereochemistry question), not a resolver bug anymore. This
# allowlist exists so the anchor can go live today without waiting on that
# review, and should shrink to empty once the merge lands.
KNOWN_DEFERRED_LOOSE_NAME_COLLISIONS = {
    frozenset({"123-32-0", "beef:2_5_dimethyl_pyrazine"}),
    frozenset({"13360-64-0", "beef:2_ethyl_5_methyl_pyrazine"}),
    frozenset({"3777-69-3", "beef:2_pentyl_furan"}),
    frozenset({"13360-65-1", "beef:3_ethyl_2_5_dimethyl_pyrazine"}),
    frozenset({"106-44-5", "beef:4_methyl_phenol"}),
    frozenset({"14667-55-1", "beef:trimethyl_pyrazine"}),
}


def test_no_two_compound_ids_share_a_loose_normalized_name(vcf_artifacts):
    compounds = vcf_artifacts["compounds"]

    ids_by_loose_name = defaultdict(set)
    for c in compounds:
        ids_by_loose_name[loose_key(c["raw_compound"])].add(c["compound_id"])

    unexpected_collisions = {
        name: ids
        for name, ids in ids_by_loose_name.items()
        if len(ids) > 1 and frozenset(ids) not in KNOWN_DEFERRED_LOOSE_NAME_COLLISIONS
    }

    assert not unexpected_collisions, (
        f"{len(unexpected_collisions)} loosely-normalized (space/hyphen-"
        f"collapsed) compound name(s) map to more than one compound_id, "
        f"beyond the 6 known-and-deferred instances — a resolver is "
        f"minting a duplicate identity for a compound the corpus already "
        f"has under a differently-spaced name: {unexpected_collisions}"
    )


# --- 8. Every product profile carries a single profile_source, and no ------
#        profile ever mixes sources (Beef Ingestion Build Spec Anchor 1) ----
#
# REVISION (Beef Ingestion): this test used to hard-fail the instant a
# second profile_source value appeared anywhere in the corpus — written
# when profile_source had, in fact, never carried a value other than
# "VCF" (see this file's own prior comment, and ingest_protein_beef.py's
# "rule zero" docstring). Beef ingestion makes that fail unconditionally
# and for the wrong reason: a second source is now the EXPECTED, correct
# state, not evidence of a bug. The real invariant — per spec, "no profile
# mixes sources" — was never actually about the corpus having only one
# source; it's about a SINGLE profile (or a single spine entry's members)
# never blending two sources' data together. That's what this rewrite
# checks instead.

# Updated 2026-08-30: egg ingestion adds a third known source,
# culinai_protein_v30_egg (see ingest_protein_egg.py's own "rule zero"
# docstring) — same reasoning as the beef revision above: a new source
# appearing is the expected, correct state, not a bug, as long as no
# single profile blends two sources together.
ALLOWED_PROFILE_SOURCES = {"VCF", "culinai_protein_v21", "culinai_protein_v30_egg"}


def test_every_profile_has_single_source(vcf_artifacts):
    profiles = vcf_artifacts["profiles"]
    missing_source = [p["vcf_product_id"] for p in profiles if not p.get("profile_source")]
    assert not missing_source, f"{len(missing_source)} profiles have no profile_source set"

    # Every profile's own source is a single, known string — never null,
    # never a list, never anything outside the sources this pipeline
    # actually knows how to build (a typo'd or invented source string
    # would otherwise slip through as if it were a real, handled one).
    unknown_source = {p["vcf_product_id"]: p["profile_source"] for p in profiles
                       if p["profile_source"] not in ALLOWED_PROFILE_SOURCES}
    assert not unknown_source, f"profile(s) with an unrecognized profile_source: {unknown_source}"

    # No SPINE ENTRY mixes sources across its own members — a base
    # ingredient's raw/cooked/smoked states (or VCF's raw/roasted
    # variants) must all come from the same extraction, never some states
    # from VCF and others from an external source under one spine_id.
    # This is the concrete form of "profiles never mix or fall back
    # across sources" (Step 5) that's actually checkable mechanically.
    spine = vcf_artifacts["spine"]
    source_by_product_id = {p["vcf_product_id"]: p["profile_source"] for p in profiles}
    mixed_spine = []
    for e in spine:
        sources = {source_by_product_id[m["vcf_product_id"]] for m in e["members"]
                   if m["vcf_product_id"] in source_by_product_id}
        if len(sources) > 1:
            mixed_spine.append((e["spine_id"], sorted(sources)))
    assert not mixed_spine, f"spine entries whose members mix profile_source: {mixed_spine}"

    # Reported, not failed: more than one source is now the correct state
    # of the world post-beef-ingestion. This is exactly the assumption
    # every downstream lens (Step 5's pairing especially) needs to know
    # changed — surfaced here so it's never silently invisible, without
    # treating its truth as a bug.
    sources_used = Counter(p["profile_source"] for p in profiles)
    print(f"profile_source breakdown: {dict(sources_used)}")


# --- 9. Descriptor coverage is reported per source, not assumed ------------

def test_descriptor_coverage_reported_per_source(vcf_artifacts):
    descriptors_meta = vcf_artifacts["meta"].get("descriptors")
    assert descriptors_meta, "meta.json has no 'descriptors' block at all"

    assert "source_counts" in descriptors_meta, (
        "descriptors meta doesn't break coverage down per source — the "
        "spec's requirement is 'reported per source, not assumed', a "
        "single blended number isn't enough to audit which source is "
        "actually contributing"
    )
    assert descriptors_meta["source_counts"], "source_counts is present but empty"

    for field in ("n_compounds_with_any_descriptor", "n_compounds_total", "coverage_fraction"):
        assert field in descriptors_meta, f"descriptors meta missing {field!r}"

    # This is a reporting check, not a threshold gate — the spec is explicit
    # that low coverage is EXPECTED (IDF selects rare compounds, published
    # descriptors cover well-studied ones, the two sets are anti-correlated
    # by construction). Failing this test on a low number would just
    # pressure someone into inflating coverage rather than reporting it
    # honestly, which is the opposite of what Step 6b asked for.


# --- Step 11 anchors ---------------------------------------------------

PHASE_REQUIRED = ["phase_frames.jsonl", "competition.jsonl", "meta.json"]


@pytest.fixture(scope="module")
def vcf_phase_artifacts():
    missing = [VCF_DIR / name for name in PHASE_REQUIRED if not (VCF_DIR / name).exists()]
    if missing:
        _loud_skip(missing)
    return {
        "compounds": _load_jsonl(VCF_DIR / "compounds.jsonl"),
        "phase_frames": _load_jsonl(VCF_DIR / "phase_frames.jsonl"),
        "competition": _load_jsonl(VCF_DIR / "competition.jsonl"),
        "meta": json.loads((VCF_DIR / "meta.json").read_text()),
    }


# --- Revision 3 fixture: compound_roles.jsonl + role-augmented profiles ---

ROLES_REQUIRED = ["compound_roles.jsonl", "profiles.jsonl", "compounds.jsonl",
                  "phase_frames.jsonl", "competition.jsonl", "meta.json"]


@pytest.fixture(scope="module")
def vcf_role_artifacts():
    missing = [VCF_DIR / name for name in ROLES_REQUIRED if not (VCF_DIR / name).exists()]
    if missing:
        _loud_skip(missing)
    profiles = _load_jsonl(VCF_DIR / "profiles.jsonl")
    if not any("role_counts" in p for p in profiles):
        pytest.skip(
            "VCF ARTIFACTS MISSING: profiles.jsonl has no role_counts field "
            "— run build_vcf_compound_roles.py then build_vcf_profile_roles.py "
            "before these Revision 3 tests can mean anything."
        )
    return {
        "compounds": _load_jsonl(VCF_DIR / "compounds.jsonl"),
        "compound_roles": _load_jsonl(VCF_DIR / "compound_roles.jsonl"),
        "profiles": profiles,
        "phase_frames": _load_jsonl(VCF_DIR / "phase_frames.jsonl"),
        "competition": _load_jsonl(VCF_DIR / "competition.jsonl"),
        "meta": json.loads((VCF_DIR / "meta.json").read_text()),
    }


# --- 10. Hydrocarbons' median XLogP is the highest of the 18 groups --------

def test_hydrocarbons_median_xlogp_is_highest(vcf_phase_artifacts):
    phase_meta = vcf_phase_artifacts["meta"].get("phase")
    assert phase_meta, "meta.json has no 'phase' block — run build_vcf_phase.py"
    group_medians = phase_meta["group_medians_xlogp"]
    assert "Hydrocarbons" in group_medians, "no Hydrocarbons group in group_medians_xlogp"

    top_group = max(group_medians, key=lambda g: group_medians[g]["median_xlogp"])
    assert top_group == "Hydrocarbons", (
        f"{top_group!r} has the highest median XLogP "
        f"({group_medians[top_group]['median_xlogp']}), not Hydrocarbons "
        f"({group_medians['Hydrocarbons']['median_xlogp']}) — 'fat carries "
        f"terpenes' is supposed to be measured, not assumed"
    )


# --- 11. Bases and Furans fall below Hydrocarbons, Esters, and Phenols -----

def test_bases_furans_below_hydrocarbons_esters_phenols(vcf_phase_artifacts):
    group_medians = vcf_phase_artifacts["meta"]["phase"]["group_medians_xlogp"]
    for required in ("Hydrocarbons", "Esters", "Phenols", "Bases", "Furans"):
        assert required in group_medians, f"{required!r} missing from group_medians_xlogp"

    ceiling = min(
        group_medians["Hydrocarbons"]["median_xlogp"],
        group_medians["Esters"]["median_xlogp"],
        group_medians["Phenols"]["median_xlogp"],
    )
    for low_group in ("Bases", "Furans"):
        m = group_medians[low_group]["median_xlogp"]
        assert m < ceiling, (
            f"{low_group}'s median XLogP ({m}) is not below the lowest of "
            f"Hydrocarbons/Esters/Phenols ({ceiling}) — the Maillard groups "
            f"are supposed to be among the least lipophilic in the dataset"
        )


# --- 12. Every compound with a CAS either has xlogp or is explicitly null --

def test_no_silent_zero_xlogp(vcf_phase_artifacts):
    compounds = vcf_phase_artifacts["compounds"]
    with_cas = [c for c in compounds if c.get("cas")]
    assert with_cas, "no compound rows carry a CAS at all — canonicalization ran?"

    missing_key = [c["compound_id"] for c in with_cas if "xlogp" not in c]
    assert not missing_key, (
        f"{len(missing_key)} compound(s) with a CAS have no 'xlogp' key at "
        f"all (should be present and either a float or explicit null): "
        f"{missing_key[:10]}"
    )

    # The actual "no silent zero" risk: a genuinely-missing value getting
    # coerced to 0.0, which would misfile it into water_phase (xlogp < 0
    # is NOT this — 0.0 is a real, valid, lipophilic-neutral reading). What
    # must never happen is a compound whose crosswalk row could not be
    # parsed as a number ('$null$' or otherwise) ending up with xlogp==0.0
    # AND phase_bucket=='both_phases' as if 0.0 were a real reading.
    suspicious = [
        c["compound_id"] for c in with_cas
        if c.get("xlogp") is None and c.get("phase_bucket") is not None
    ]
    assert not suspicious, (
        f"{len(suspicious)} compound(s) have xlogp=null but a non-null "
        f"phase_bucket — a bucket should never be assigned without a real "
        f"xlogp backing it: {suspicious[:10]}"
    )


# --- 13. No phase-behaviour sentence outside phase_frames.jsonl ------------

def test_no_phase_sentence_outside_phase_frames(vcf_phase_artifacts):
    frames = vcf_phase_artifacts["phase_frames"]
    frame_ids = {f["frame_id"] for f in frames}
    assert frame_ids, "phase_frames.jsonl is empty"

    for f in frames:
        if f.get("pending_authoring"):
            # A pending_authoring stub (Revision 4: smoked_product_fat_phase)
            # is intentionally unauthored — see test_smoked_stub_frame_is_dormant
            # for the anchor that no row may render against it while empty.
            continue
        assert isinstance(f.get("sentence"), str) and f["sentence"].strip(), (
            f"frame {f.get('frame_id')!r} has no authored sentence text"
        )
        assert f.get("source"), f"frame {f.get('frame_id')!r} has no source attribution"

    competition = vcf_phase_artifacts["competition"]
    dangling = [
        row for row in competition
        if row.get("frame_id") is not None and row["frame_id"] not in frame_ids
    ]
    assert not dangling, (
        f"{len(dangling)} competition row(s) reference a frame_id that "
        f"doesn't exist in phase_frames.jsonl — this would surface a "
        f"phase-behaviour claim with no authored sentence backing it: "
        f"{dangling[:5]}"
    )


# --- 14. Volatility claims are suppressed below coverage threshold --------

def test_volatility_suppressed_below_coverage_threshold(vcf_phase_artifacts):
    volatility_meta = vcf_phase_artifacts["meta"].get("volatility")
    if volatility_meta is None:
        pytest.skip(
            "VCF ARTIFACTS MISSING: meta.json has no 'volatility' block yet "
            "— Step 11c (boiling-point fetch) has not completed. This is a "
            "LOUD skip, same as a missing JSONL file: PubChem's PUG View "
            "endpoint was returning 503 for this build as of 2026-08-29 "
            "(confirmed not a pacing issue — the lighter /property/ "
            "endpoint used for XLogP worked throughout); re-run "
            "build_vcf_volatility.py once that clears, or against a "
            "different boiling-point source, before this test can mean "
            "anything."
        )
    coverage = volatility_meta["coverage_fraction"]
    claims_suppressed = volatility_meta.get("volatility_claims_suppressed")
    if coverage < 0.5:
        assert claims_suppressed is True, (
            f"boiling-point coverage is {coverage:.1%}, below the spec's "
            f"~50% threshold, but volatility_claims_suppressed is not True "
            f"— thin data must not be allowed to make volatility claims"
        )
    else:
        assert claims_suppressed is False, (
            f"boiling-point coverage is {coverage:.1%}, at/above threshold, "
            f"but claims are still marked suppressed — check the threshold "
            f"logic in build_vcf_volatility.py"
        )


# --- 15. Every flagship-frame row satisfies its own trigger, re-derived ----
# from stored fields only. REVISION 4: fat_phase_phenol_terpene_carrier
# (renamed from fat_phase_smoke_terpene_competition) is a pure behaviour
# claim — group_pair=={Hydrocarbons,Phenols} (order-independent) AND both
# sides' own group_percentile >= 75. No role/provenance fields at all.
# Reconstructs a PAIR context from each row's OWN stored fields
# (bucket/a_group/b_group/a_group_percentile/b_group_percentile) and
# re-evaluates the trigger via evaluate_trigger — independent of whatever
# code path in build_vcf_competition.py assigned frame_id.

FLAGSHIP_FRAME_ID = "fat_phase_phenol_terpene_carrier"
ALDEHYDE_FRAME_ID = "fat_phase_aldehyde_load_crowding"  # Post-Ingestion Fixes, Fix 2


def test_flagship_frame_rows_satisfy_trigger_independently(vcf_phase_artifacts):
    competition = vcf_phase_artifacts["competition"]
    frames = vcf_phase_artifacts["phase_frames"]
    flagship_trigger = next(f["trigger"] for f in frames if f["frame_id"] == FLAGSHIP_FRAME_ID)
    fat_affine_buckets = {"fat_phase", "fat_leaning"}

    flagged = [r for r in competition if r.get("frame_id") == FLAGSHIP_FRAME_ID]
    assert flagged, (
        "no competition row carries the flagship frame_id at all. As of "
        "Revision 4 roughly 37 rows should qualify (anise, ouzo, fennel, "
        "star anise, calamus, chervil, pimento leaf, Korean mint...) — an "
        "empty result here is a regression to investigate, not an "
        "expected outcome the way it briefly was under Revision 3's much "
        "narrower (and since-retired) smoke_marker-gated trigger."
    )

    bad = []
    for r in flagged:
        if r.get("bucket") not in fat_affine_buckets:
            bad.append(("frame_id set on a non-fat-affine bucket", r))
            continue
        reconstructed = {
            "bucket": r["bucket"],
            "group_pair": [r["a_group"], r["b_group"]],
            "a_group_percentile": r["a_group_percentile"],
            "b_group_percentile": r["b_group_percentile"],
        }
        if not evaluate_trigger(reconstructed, flagship_trigger):
            bad.append(("frame_id set but stored fields do not satisfy the structured trigger", r))
        if r.get("conflict_type") != "cross_group_carrier":
            bad.append(("frame_id set on a non-cross-group conflict_type", r))

    assert not bad, (
        f"{len(bad)} row(s) carry frame_id={FLAGSHIP_FRAME_ID!r} but fail "
        f"independent re-evaluation of the trigger stored in "
        f"phase_frames.jsonl: {bad[:5]}"
    )


# --- 16. No same-dominant-group pair ever carries the flagship frame -------
# The original bug's exact shape: a_group == b_group (e.g. both
# Hydrocarbons) but frame_id set anyway. REVISION 4: structurally
# impossible now (group_pair "equals" against a 2-DISTINCT-item set can
# never match a_group==b_group), but still checked directly against the
# stored field rather than relied upon, in case a future trigger edit
# reintroduces the possibility.

def test_no_same_group_row_carries_flagship_frame(vcf_phase_artifacts):
    competition = vcf_phase_artifacts["competition"]
    offenders = [
        r for r in competition
        if r.get("frame_id") == FLAGSHIP_FRAME_ID
        and r.get("a_group") == r.get("b_group")
    ]
    assert not offenders, (
        f"{len(offenders)} row(s) with a_group == b_group carry the "
        f"phenol/terpene carrier frame — group_pair equality against a "
        f"2-distinct-item set should make this impossible by construction: "
        f"{offenders[:5]}"
    )


# --- 17. Competition tightening is real and reported (both_phases excluded)

def test_competition_tightening_reported_and_both_phases_excluded(vcf_phase_artifacts):
    competition = vcf_phase_artifacts["competition"]
    comp_meta = vcf_phase_artifacts["meta"].get("competition")
    assert comp_meta, "meta.json has no 'competition' block"

    both_phases_rows = [r for r in competition if r.get("bucket") == "both_phases"]
    assert not both_phases_rows, (
        f"{len(both_phases_rows)} competition row(s) still have bucket == "
        f"'both_phases' — that bucket is supposed to be excluded entirely "
        f"(a compound there has an alternative route, not a shared limited "
        f"carrier)"
    )

    for key in (
        "n_candidates_same_bucket_excl_both_phases",
        "n_competition_rows",
        "n_competition_rows_previous_revision_flat_floor",
        "tightening",
        "top_group_pairs_missing_frame",
        "group_percentiles",
        "conflicts_per_dish",
    ):
        assert key in comp_meta, f"meta.json's competition block is missing {key!r}"

    assert comp_meta["n_competition_rows"] < comp_meta["n_candidates_same_bucket_excl_both_phases"], (
        "the percentile gate should strictly reduce the row count relative "
        "to the pre-gate candidate pool — if it didn't, the gate isn't "
        "being applied"
    )


# --- 18. Both sides of every surviving row clear their own group's p75 ----
# (Revision 2, Change 1) — re-derived independently from group_percentiles,
# not by trusting the code path that computed a_group_share/b_group_share.

def test_both_sides_clear_own_group_75th_percentile(vcf_phase_artifacts):
    competition = vcf_phase_artifacts["competition"]
    comp_meta = vcf_phase_artifacts["meta"]["competition"]
    percentiles = comp_meta["group_percentiles"]
    assert percentiles, "meta.json's competition.group_percentiles is empty"

    bad = []
    for r in competition:
        a_table = percentiles.get(r["a_group"])
        b_table = percentiles.get(r["b_group"])
        if a_table is None or b_table is None:
            bad.append(("no percentile table for a_group/b_group", r))
            continue
        if r["a_group_share"] < a_table["p75_share"] or r["b_group_share"] < b_table["p75_share"]:
            bad.append(("share below own group's p75", r))
    assert not bad, (
        f"{len(bad)} row(s) failed independent re-verification against "
        f"group_percentiles: {bad[:5]}"
    )


# --- 19. Per-group percentile table is present and internally consistent --
# (reproducibility from profiles.jsonl/compounds.jsonl is checked at build
# time by construction — compute_group_percentile_table takes no other
# input — this test checks the table's own shape rather than recomputing
# it a second time here.)

def test_group_percentile_table_shape(vcf_phase_artifacts):
    percentiles = vcf_phase_artifacts["meta"]["competition"]["group_percentiles"]
    assert percentiles, "group_percentiles table is empty"
    for group, t in percentiles.items():
        for key in ("n_products", "median_share", "p75_share", "p90_share"):
            assert key in t, f"group_percentiles[{group!r}] missing {key!r}"
        assert t["n_products"] > 0, f"group_percentiles[{group!r}] has n_products == 0"
        assert 0 <= t["median_share"] <= t["p75_share"] <= t["p90_share"] <= 1, (
            f"group_percentiles[{group!r}] percentiles are not monotonic/in "
            f"[0,1]: {t}"
        )


# --- 20. same_group_crowding rows never carry the FLAGSHIP's frame_id ----
# (Revision 2, Change 2) — the flagship frame requires exactly
# {Hydrocarbons, Phenols}, which by definition can't be a_group == b_group,
# but this checks the STORED conflict_type field directly rather than
# re-deriving it, so a future frame with a looser trigger can't silently
# violate this. Post-Ingestion Fixes, Fix 2: NARROWED to the flagship
# specifically — fat_phase_aldehyde_load_crowding is deliberately authored
# to fire on same_group_crowding rows too (both sides aldehyde-dominant is
# a real crowding claim, not a bug), so "no frame_id at all on a
# same-group row" is no longer the right invariant. What must still never
# happen: the flagship's *own* group_pair=={Hydrocarbons,Phenols} claim
# attaching to a same-group row, which is what this test now checks.

def test_same_group_crowding_never_carries_cross_group_frame(vcf_phase_artifacts):
    competition = vcf_phase_artifacts["competition"]
    offenders = [
        r for r in competition
        if r.get("conflict_type") == "same_group_crowding" and r.get("frame_id") == FLAGSHIP_FRAME_ID
    ]
    assert not offenders, (
        f"{len(offenders)} same_group_crowding row(s) carry the flagship's "
        f"frame_id — {FLAGSHIP_FRAME_ID} is a cross-group claim by "
        f"construction (group_pair equality against a 2-distinct-item set "
        f"can't match a_group==b_group), so none should ever attach to a "
        f"same-group row: {offenders[:5]}"
    )


# --- 21. render_mode / frame_id / sentence-field invariants (Change 3) ----

def test_render_mode_matches_frame_id_and_no_stray_sentences(vcf_phase_artifacts):
    competition = vcf_phase_artifacts["competition"]
    frames = vcf_phase_artifacts["phase_frames"]
    frame_ids = {f["frame_id"] for f in frames}

    for r in competition:
        has_frame = r.get("frame_id") is not None
        expected_mode = "framed" if has_frame else "data_only"
        assert r.get("render_mode") == expected_mode, (
            f"render_mode {r.get('render_mode')!r} doesn't match frame_id "
            f"presence ({has_frame}): {r}"
        )
        if has_frame:
            assert r["frame_id"] in frame_ids, (
                f"row's frame_id {r['frame_id']!r} isn't in phase_frames.jsonl: {r}"
            )
        assert "sentence" not in r, (
            f"competition row carries a 'sentence' field — sentence text "
            f"must live only in phase_frames.jsonl, joined by frame_id at "
            f"render time: {r}"
        )


# =====================================================================
# Revision 3 anchors (compound role layer + structured triggers)
# =====================================================================

# --- 22. Every framed row clears the literal group_pair/percentile floor,
# checked against the literal values (not via evaluate_trigger) — decoupled
# from the trigger evaluator itself so a bug in vcf_trigger_lib can't blind
# this anchor the way it could test 15. REVISION 4: replaces the retired
# smoke_marker-threshold check (that role no longer drives any frame — see
# build_vcf_competition.py's docstring).

def test_framed_rows_clear_literal_group_pair_and_percentile_floor(vcf_phase_artifacts):
    competition = vcf_phase_artifacts["competition"]
    percentile_floor = 75

    flagged = [r for r in competition if r.get("frame_id") == FLAGSHIP_FRAME_ID]
    assert flagged, "no row carries the flagship frame_id — see test 15 for context"

    bad = []
    for r in flagged:
        pair_ok = {r["a_group"], r["b_group"]} == {"Hydrocarbons", "Phenols"}
        pctl_ok = r["a_group_percentile"] >= percentile_floor and r["b_group_percentile"] >= percentile_floor
        if not (pair_ok and pctl_ok):
            bad.append(r)
    assert not bad, (
        f"{len(bad)} framed row(s) fail the literal group_pair=="
        f"{{'Hydrocarbons','Phenols'}} AND both percentiles >= "
        f"{percentile_floor} check: {bad[:5]}"
    )


# --- 23. No explicitly-excluded compound ever carries the smoke_marker role
# Imports the SAME exclusion sets build_vcf_compound_roles.py declares
# (not a re-typed copy) and checks the actual compound_roles.jsonl output
# against them — this catches a future edit that widens a curated set or a
# pattern match without re-running the build's own internal assertion.

def test_no_excluded_compound_carries_smoke_marker_role(vcf_role_artifacts):
    from build_vcf_compound_roles import (
        SMOKE_MARKER_EXCLUDED_CAS,
        SMOKE_MARKER_EXCLUDED_NAME_PATTERN_IDS,
    )

    compounds = vcf_role_artifacts["compounds"]
    cas_by_id = {c["compound_id"]: c.get("cas") for c in compounds}
    excluded_ids = set(SMOKE_MARKER_EXCLUDED_NAME_PATTERN_IDS) | {
        cid for cid, cas in cas_by_id.items() if cas in SMOKE_MARKER_EXCLUDED_CAS
    }
    assert excluded_ids, "exclusion set resolved to zero compound_ids — cas_by_id join is broken"

    smoke_tagged = {r["compound_id"] for r in vcf_role_artifacts["compound_roles"] if r["role"] == "smoke_marker"}
    bad = smoke_tagged & excluded_ids
    assert not bad, (
        f"{len(bad)} explicitly-excluded compound(s) carry the smoke_marker "
        f"role in compound_roles.jsonl: {bad} — this is the exact failure "
        f"mode build_vcf_compound_roles.py's own build-time assertion "
        f"exists to catch; re-checked here against the ARTIFACT, "
        f"independent of whether that assertion actually ran"
    )


# --- 24. trigger_description is mechanically regenerated, never stale ------
# Every frame's trigger_description must equal
# generate_trigger_description(trigger) freshly computed from the SAME
# module the build script uses — catches a hand-edit to either field that
# left the other one behind.

def test_trigger_descriptions_match_regenerated_form(vcf_role_artifacts):
    frames = vcf_role_artifacts["phase_frames"]
    assert frames, "phase_frames.jsonl is empty"
    bad = []
    for f in frames:
        assert "trigger" in f and isinstance(f["trigger"], dict), (
            f"frame {f.get('frame_id')!r} trigger is not a structured dict — "
            f"Revision 3 requires {{'all': [...]}} , not a hand-written string"
        )
        regenerated = generate_trigger_description(f["trigger"])
        if regenerated != f.get("trigger_description"):
            bad.append((f["frame_id"], f.get("trigger_description"), regenerated))
    assert not bad, (
        f"{len(bad)} frame(s) have a stored trigger_description that "
        f"disagrees with generate_trigger_description(trigger): {bad}"
    )


# --- 25. Role coverage is reported honestly — no false completeness claim --
# meta.json's compound_roles.coverage counts must match the actual
# compound_roles.jsonl artifact (no silent inflation between what was
# computed and what was written), and the smoke_marker role — the one role
# with an explicit target count from the review (36) — must admit any
# shortfall rather than claim the target was hit.

def test_role_coverage_reported_honestly(vcf_role_artifacts):
    roles_meta = vcf_role_artifacts["meta"].get("compound_roles")
    assert roles_meta and roles_meta.get("coverage"), "meta.json has no compound_roles.coverage block"
    coverage = roles_meta["coverage"]

    role_rows = vcf_role_artifacts["compound_roles"]
    actual_counts = Counter(r["role"] for r in role_rows)

    for role_name in ("smoke_marker", "maillard_marker", "lipid_oxidation_marker", "terpene_mono", "terpene_sesqui"):
        assert role_name in coverage, f"compound_roles.coverage is missing {role_name!r}"
        reported_n = coverage[role_name].get("n_tagged")
        assert reported_n == actual_counts.get(role_name, 0), (
            f"{role_name!r} coverage claims n_tagged={reported_n} but "
            f"compound_roles.jsonl actually contains "
            f"{actual_counts.get(role_name, 0)} rows for that role — "
            f"the report must match the artifact"
        )

    smoke_n = actual_counts.get("smoke_marker", 0)
    smoke_target = coverage["smoke_marker"].get("n_curated_target")
    assert smoke_target is not None, "smoke_marker coverage has no n_curated_target to compare against"
    if smoke_n < smoke_target:
        gap_note = coverage["smoke_marker"].get("gap_note", "")
        assert gap_note and str(smoke_n) in gap_note, (
            f"smoke_marker landed at {smoke_n} of a stated target of "
            f"{smoke_target}, but coverage's gap_note doesn't mention the "
            f"actual count {smoke_n} — a shortfall must be stated plainly, "
            f"not left implicit in two numbers a reader has to subtract "
            f"themselves"
        )


# --- 26. Re-evaluating EVERY row's trigger from stored fields reproduces --
# frame_id exactly — not just the rows that already carry a frame_id (test
# 15's scope). This also catches a FALSE NEGATIVE: a row that should have
# been framed under the stored trigger but wasn't, which test 15 alone
# cannot see because it only inspects already-flagged rows. REVISION 4:
# reconstructs from group_pair/percentile fields, not smoke_marker ones.
# Post-Ingestion Fixes, Fix 2: generalized from "check the flagship only"
# to "check every assignable frame" — fat_phase_aldehyde_load_crowding is
# a second non-pending PAIR-context frame now, with its OWN conflict_type
# rule (none — it fires on same_group_crowding and cross_group_carrier
# alike, unlike the flagship). Mirrors build_vcf_competition.py's own
# _assert_frame_assignments_correct, independently re-implemented here
# rather than imported, per this test file's existing convention of never
# trusting the build's own assertion to catch its own bugs.

ASSIGNABLE_FRAME_CONFLICT_TYPE_CONSTRAINT = {
    FLAGSHIP_FRAME_ID: "cross_group_carrier",
    ALDEHYDE_FRAME_ID: None,  # no restriction
}


def test_every_row_frame_id_matches_trigger_reevaluation(vcf_phase_artifacts):
    competition = vcf_phase_artifacts["competition"]
    frames = vcf_phase_artifacts["phase_frames"]
    frames_by_id = {f["frame_id"]: f for f in frames}

    bad = []
    for r in competition:
        reconstructed = {
            "bucket": r["bucket"],
            "group_pair": [r["a_group"], r["b_group"]],
            "groups_present": {r["a_group"], r["b_group"]},
            "a_group_percentile": r["a_group_percentile"],
            "b_group_percentile": r["b_group_percentile"],
        }
        fired = []
        for frame_id, required_conflict_type in ASSIGNABLE_FRAME_CONFLICT_TYPE_CONSTRAINT.items():
            frame = frames_by_id.get(frame_id)
            if not frame or frame.get("pending_authoring"):
                continue
            trigger_fires = evaluate_trigger(reconstructed, frame["trigger"])
            satisfies_constraint = required_conflict_type is None or r["conflict_type"] == required_conflict_type
            if trigger_fires and satisfies_constraint:
                fired.append(frame_id)
        expected_frame_id = fired[0] if len(fired) == 1 else (None if not fired else "AMBIGUOUS:" + str(fired))
        if r["frame_id"] != expected_frame_id:
            bad.append((r["frame_id"], expected_frame_id, r))
    assert not bad, (
        f"{len(bad)} row(s) have a stored frame_id that disagrees with "
        f"independent re-evaluation of every assignable frame's trigger "
        f"from their own stored fields (stored, expected, row): {bad[:5]}"
    )


# --- 27. Every row with conflict_subtype set clears the relevant terpene --
# role's 75th percentile on BOTH sides (Revision 4, Change 3) — re-derived
# from meta.json's own role_percentiles table, not by trusting the code
# path that computed conflict_subtype.

def test_conflict_subtype_rows_clear_role_percentile(vcf_phase_artifacts):
    competition = vcf_phase_artifacts["competition"]
    comp_meta = vcf_phase_artifacts["meta"]["competition"]
    role_percentiles = comp_meta["terpene_subtype_gate"]["role_percentiles"]
    assert role_percentiles, "meta.json's competition.terpene_subtype_gate.role_percentiles is empty"

    subtyped = [r for r in competition if r.get("conflict_subtype") is not None]
    assert subtyped, "no competition row carries a conflict_subtype at all — regression to investigate"

    bad = []
    for r in subtyped:
        subtype = r["conflict_subtype"]
        assert subtype in ("mono_crowding", "sesqui_crowding"), (
            f"unexpected conflict_subtype value {subtype!r} — per spec, "
            f"mono_vs_sesqui must never be stored as a conflict_subtype "
            f"(it's a drop reason, not a kept label): {r}"
        )
        role = "terpene_mono" if subtype == "mono_crowding" else "terpene_sesqui"
        gate = role_percentiles[role]["p75_share"]
        a_share = r[f"a_{role}_share"]
        b_share = r[f"b_{role}_share"]
        if a_share is None or b_share is None or a_share < gate or b_share < gate:
            bad.append((role, gate, r))
    assert not bad, (
        f"{len(bad)} conflict_subtype row(s) fail independent "
        f"re-verification against their own role's 75th percentile: {bad[:5]}"
    )

    # Every OTHER row (not Hydrocarbons-vs-Hydrocarbons) must carry
    # conflict_subtype: null — the gate is scoped to that one group pair.
    non_hh_with_subtype = [
        r for r in competition
        if r.get("conflict_subtype") is not None
        and not (r["a_group"] == "Hydrocarbons" and r["b_group"] == "Hydrocarbons")
    ]
    assert not non_hh_with_subtype, (
        f"{len(non_hh_with_subtype)} row(s) outside Hydrocarbons-vs-"
        f"Hydrocarbons carry a conflict_subtype — the terpene split is "
        f"scoped to that one same-group pair only: {non_hh_with_subtype[:5]}"
    )


# --- 28. No frame sentence asserts a preparation/provenance claim unless --
# its trigger actually reads the preparation/cure_state field (Revision 4's
# core rule — see build_vcf_phase_frames.py's docstring). A crude but
# effective proxy: if the authored sentence contains a provenance-shaped
# word, the trigger had better be reading a preparation-shaped field.

_PROVENANCE_WORDS = ("smoke", "smoked", "roast", "roasted", "cured", "curing", "kiln", "kilned", "barrel")


def _trigger_field_names(trigger):
    return [c["field"] for c in trigger.get("all", [])]


def test_no_frame_sentence_asserts_provenance_without_preparation_trigger(vcf_phase_artifacts):
    frames = vcf_phase_artifacts["phase_frames"]
    assert frames, "phase_frames.jsonl is empty"

    bad = []
    for f in frames:
        sentence = (f.get("sentence") or "").lower()
        if not any(word in sentence for word in _PROVENANCE_WORDS):
            continue
        fields = _trigger_field_names(f["trigger"])
        reads_preparation = any("preparation" in fld or "cure_state" in fld for fld in fields)
        if not reads_preparation:
            bad.append((f["frame_id"], sentence))
    assert not bad, (
        f"{len(bad)} frame(s) have a sentence asserting a preparation/"
        f"provenance claim (smoke/roast/cure/kiln/barrel) but a trigger "
        f"that never reads the preparation or cure_state field — this is "
        f"exactly the category error three prior revisions made (a "
        f"molecule licensing a claim about how a product was prepared): "
        f"{bad}"
    )


# --- 29. The smoked-product stub stays dormant regardless of its trigger --
# (Revision 4, Change 2): pending_authoring, empty sentence, and — the part
# that actually matters — no row anywhere renders frame_id set to it, even
# though its trigger is real and evaluated for n_firing_rows reporting.

def test_smoked_stub_frame_is_dormant(vcf_phase_artifacts):
    frames = vcf_phase_artifacts["phase_frames"]
    stub = next((f for f in frames if f["frame_id"] == "smoked_product_fat_phase"), None)
    assert stub is not None, "smoked_product_fat_phase frame is missing from phase_frames.jsonl"
    assert stub["pending_authoring"] is True, (
        "smoked_product_fat_phase must stay pending_authoring — authoring "
        "the sentence is James's task, not a build step"
    )
    assert not (stub.get("sentence") or "").strip(), (
        "smoked_product_fat_phase must have an empty sentence while "
        "pending_authoring is True"
    )

    competition = vcf_phase_artifacts["competition"]
    offenders = [r for r in competition if r.get("frame_id") == "smoked_product_fat_phase"]
    assert not offenders, (
        f"{len(offenders)} competition row(s) carry frame_id== "
        f"'smoked_product_fat_phase' despite it being pending_authoring "
        f"with an empty sentence — no row may ever render against an "
        f"unauthored frame: {offenders[:5]}"
    )

    comp_meta = vcf_phase_artifacts["meta"]["competition"]
    stub_report = comp_meta.get("smoked_product_fat_phase_frame")
    assert stub_report is not None, "meta.json is missing the smoked_product_fat_phase_frame report block"
    assert "n_firing_rows" in stub_report, "smoked_product_fat_phase_frame report has no n_firing_rows"


# =============================================================================
# BEEF INGESTION BUILD SPEC — six required anchors (2026-08-29). This is the
# first non-VCF source the pipeline has ingested; see
# ingest_protein_beef.py's module docstring for the two spec premises that
# turned out to be false against the real workbook ("all beef rows carry
# CAS"; the preparation_state->state_tier table's own dry-aged/wet-aged row
# contradicting the spec's own prose) and how each was resolved.
#
# 30. Every beef profile carries profile_source=culinai_protein_v21, and no
#     profile mixes sources (the corpus-wide half of this is test 8, above;
#     this test is the beef-specific half — CHECKING that beef products in
#     particular landed with the right, single source).
# 31. No stored/spoilage row appears in any profile reachable by a lens.
# 32. beef:fat exists as a distinct product; no muscle profile contains
#     fat-only compounds (concretely: the BP-052 fat-duplicate row never
#     contributes to beef:muscle).
# 33. No straight-chain alkane carries a terpene role (MR-15).
# 34. Rows with evidence_mode != measured are excluded from document-
#     frequency counts (mechanism check — this pass's real beef data is
#     100% measured, so there is no non-measured row to show a count
#     *change* against; what's checked is that the exclusion machinery
#     itself is wired in and structurally correct, per Step 4: "implement
#     the handling now" even though beef alone can't exercise it).
#     Plus: vocabulary_version differs from the pre-beef baseline.
# =============================================================================

BEEF_OBSERVATIONS_JSONL = Path(__file__).resolve().parents[1] / "artifacts" / "protein" / "beef_observations.jsonl"
BEEF_PROFILES_PREBUILT_JSONL = (
    Path(__file__).resolve().parents[1] / "artifacts" / "protein" / "beef_profiles_prebuilt.jsonl"
)


@pytest.fixture(scope="module")
def protein_beef_artifacts(vcf_artifacts):
    missing = [p for p in (BEEF_OBSERVATIONS_JSONL, BEEF_PROFILES_PREBUILT_JSONL) if not p.exists()]
    if missing:
        pytest.skip(
            f"PROTEIN BEEF ARTIFACTS MISSING: {[p.name for p in missing]} — run "
            f"pipeline/scripts/ingest_protein_beef.py first. LOUD skip, same "
            f"reporting requirement as _loud_skip above."
        )
    return {
        "observations": _load_jsonl(BEEF_OBSERVATIONS_JSONL),
        "beef_profiles_prebuilt": _load_jsonl(BEEF_PROFILES_PREBUILT_JSONL),
    }


# --- 30. Every beef profile carries the beef profile_source, none mixed ----

def test_beef_profiles_carry_single_correct_source(vcf_artifacts, protein_beef_artifacts):
    beef_product_ids = {p["vcf_product_id"] for p in protein_beef_artifacts["beef_profiles_prebuilt"]}
    profiles_by_id = {p["vcf_product_id"]: p for p in vcf_artifacts["profiles"]}
    missing = beef_product_ids - profiles_by_id.keys()
    assert not missing, f"beef product(s) built by ingestion but absent from the merged profiles.jsonl: {missing}"
    wrong_source = {
        pid: profiles_by_id[pid]["profile_source"] for pid in beef_product_ids
        if profiles_by_id[pid]["profile_source"] != "culinai_protein_v21"
    }
    assert not wrong_source, f"beef product(s) with the wrong profile_source after merge: {wrong_source}"


# --- 31. No stored/spoilage row appears in any profile reachable by a lens -

def test_spoilage_row_excluded_from_every_profile(protein_beef_artifacts):
    observations = protein_beef_artifacts["observations"]
    spoilage_obs = [o for o in observations if (o.get("exclusion_reason") or "").startswith("pat_beef_005")]
    assert spoilage_obs, "no spoilage-marker observation found at all — PAT-BEEF-005's row may not have landed"
    not_excluded = [o for o in spoilage_obs if not o.get("excluded")]
    assert not not_excluded, f"spoilage observation(s) not marked excluded: {not_excluded}"
    has_tier = [o for o in spoilage_obs if o.get("tiers")]
    assert not has_tier, f"spoilage observation(s) still assigned to a state tier: {has_tier}"

    # Cross-check against the actual built profiles: the spoilage
    # compound (3-octanone, CAS 106-68-3) must never enter a beef:muscle
    # profile VIA the spoilage row specifically — other legitimate
    # detections of the same compound elsewhere are a different claim and
    # not what PAT-BEEF-005 excludes.
    muscle_obs_same_compound = [
        o for o in observations
        if o.get("resolved_compound_id") == "106-68-3" and o.get("product_id") == "beef:muscle" and not o["excluded"]
    ]
    for o in muscle_obs_same_compound:
        assert o.get("source_tab") != "Verified Beef Profiles" or o.get("detected_record_id") != "BP-053", (
            f"the spoilage row (BP-053) contributed to a beef:muscle profile: {o}"
        )


# --- 32. beef:fat is distinct; no muscle profile contains fat-only compounds

def test_fat_product_split_from_muscle(vcf_artifacts, protein_beef_artifacts):
    spine_ids = {e["spine_id"] for e in vcf_artifacts["spine"]}
    assert "beef:muscle" in spine_ids, "beef:muscle spine entry missing"
    assert "beef:fat" in spine_ids, "beef:fat spine entry missing"

    observations = protein_beef_artifacts["observations"]
    superseded = [
        o for o in observations
        if (o.get("exclusion_reason") or "") == "superseded_duplicate_of_verified_beef_fat_profile_bfp_001"
    ]
    assert superseded, "the BP-052 fat-duplicate row was not found/excluded — Step 1's product split may not have run"
    for o in superseded:
        assert o.get("excluded") is True and not o.get("tiers"), (
            f"BP-052 (or an equivalent superseded fat row) still contributed to a profile: {o}"
        )
        assert o.get("product_id") != "beef:muscle", (
            f"a superseded fat row was attributed to beef:muscle instead of being dropped: {o}"
        )


# --- 33. No straight-chain alkane carries a terpene role (MR-15) -----------

STRAIGHT_CHAIN_ALKANE_RE_FOR_TEST = re.compile(
    r"^(meth|eth|prop|but|pent|hex|hept|oct|non|dec|undec|dodec|tridec|"
    r"tetradec|pentadec|hexadec|heptadec|octadec|nonadec|icos)ane$",
    re.IGNORECASE,
)


def test_no_straight_chain_alkane_carries_terpene_role(vcf_role_artifacts):
    compounds_by_id = {c["compound_id"]: c for c in vcf_role_artifacts["compounds"]}
    roles = vcf_role_artifacts["compound_roles"]
    offenders = []
    for r in roles:
        if r["role"] not in ("terpene_mono", "terpene_sesqui"):
            continue
        c = compounds_by_id.get(r["compound_id"])
        if not c:
            continue
        raw = (c.get("raw_compound") or "").strip()
        raw_core = re.sub(r"\s*\(=.*\)\s*$", "", raw)
        if STRAIGHT_CHAIN_ALKANE_RE_FOR_TEST.match(raw_core):
            offenders.append((r["compound_id"], raw, r["role"]))
    assert not offenders, f"straight-chain alkane(s) carrying a terpene role (MR-15 violation): {offenders}"

    # The guard's own role must actually exist and be populated with beef's
    # real alkanes — a passing test above for the wrong reason (the role
    # simply never fires) would be worse than no test.
    alkane_role_ids = {r["compound_id"] for r in roles if r["role"] == "lipid_degradation_fragment"}
    assert alkane_role_ids, "lipid_degradation_fragment role never fired at all — MR-15 guard may not be wired in"
    tagged_names = {compounds_by_id[cid]["raw_compound"] for cid in alkane_role_ids if cid in compounds_by_id}
    assert any(n.lower() == "hexane" for n in tagged_names) or any(n.lower() == "pentane" for n in tagged_names), (
        f"expected at least one of beef's own straight-chain alkanes (hexane, pentane, ...) tagged "
        f"lipid_degradation_fragment; got: {sorted(tagged_names)[:10]}"
    )


# --- 34. evidence_mode != measured excluded from df; vocabulary refrozen ---

def test_evidence_mode_df_exclusion_and_vocabulary_refrozen(vcf_artifacts, protein_beef_artifacts):
    for p in protein_beef_artifacts["beef_profiles_prebuilt"]:
        eligible = set(p["df_eligible_compound_ids"])
        full = set(p["compound_ids"])
        assert eligible <= full, (
            f"{p['vcf_product_id']}: df_eligible_compound_ids is not a subset of compound_ids — "
            f"the df-exclusion mechanism would inflate document frequency for a compound not even "
            f"in the profile"
        )
    # This pass's real beef data is 100% evidence_mode=measured (Step 4's
    # own premise, verified against the workbook), so eligible == full for
    # every beef profile right now — that is the CORRECT current state,
    # not a sign the mechanism is unused; it becomes observable the moment
    # a future family (avian's 18 founder_equivalence rows) has anything
    # to exclude.
    all_equal = all(
        set(p["df_eligible_compound_ids"]) == set(p["compound_ids"])
        for p in protein_beef_artifacts["beef_profiles_prebuilt"]
    )
    assert all_equal, "beef data is expected to be 100% measured this pass — an unequal set means something else changed"

    vocab_version = vcf_artifacts["meta"]["spine"]["vocabulary_version"]
    assert vocab_version != "vcf_spine_v1", (
        f"vocabulary_version ({vocab_version!r}) was not bumped from the pre-beef-ingestion baseline "
        f"'vcf_spine_v1' — Step 6 requires the refreeze to be visible in meta.json"
    )


# =============================================================================
# POST-INGESTION FIXES (2026-08-29) — five required anchors, following the
# smoked_product_fat_phase mechanism-gate fix and the new
# fat_phase_aldehyde_load_crowding frame prompted by beef's own firing rows.
#
# 35. No row carries frame_id=smoked_product_fat_phase unless its contending
#     group is Phenols — trivially true today since frame_id can never be
#     set to a pending_authoring frame at all (test 29 checks that), but
#     this test additionally checks the STORED TRIGGER itself requires the
#     mechanism gate, so a future removal of that condition (which would
#     make the "trivially true" argument stop holding once someone
#     un-pends the frame) is caught here first.
# 36. The aldehyde frame's sentence contains no reference to smoke, curing,
#     or any preparation word.
# 37. Every pairs.jsonl row carries a suppressed_reason field (present, may
#     be null) — none are deleted, matching the "auditable and reversible"
#     instruction.
# 38. phase_frames.jsonl has exactly 5 entries: 4 authored, 1 pending.
# =============================================================================

def test_smoked_stub_trigger_has_mechanism_gate(vcf_phase_artifacts):
    frames = vcf_phase_artifacts["phase_frames"]
    stub = next(f for f in frames if f["frame_id"] == "smoked_product_fat_phase")
    conditions = stub["trigger"]["all"]
    mechanism_conditions = [
        c for c in conditions
        if c["field"] == "either_side.group" and c.get("eq") == "Phenols"
    ]
    assert mechanism_conditions, (
        f"smoked_product_fat_phase's stored trigger has no "
        f"'either_side.group eq Phenols' condition — without a mechanism "
        f"gate alongside the provenance gate (either_side.preparation "
        f"contains 'smoked'), this frame would fire on ANY smoked product "
        f"regardless of what it's chemically contending on, which is "
        f"exactly the bug beef ingestion caught (4 firing rows, all "
        f"contending on Carbonyls,aldehydes, none on Phenols)."
    )
    provenance_conditions = [
        c for c in conditions
        if c["field"] == "either_side.preparation" and c.get("contains") == "smoked"
    ]
    assert provenance_conditions, "smoked_product_fat_phase lost its provenance gate entirely"

    # Both conditions must be either_side (same-side binding), never
    # other_side — otherwise the mechanism gate could be satisfied by the
    # WRONG side (the one that isn't actually smoked).
    assert not any(c["field"] == "other_side.group" for c in conditions), (
        "mechanism gate must be either_side.group (bound to the smoked "
        "side), not other_side.group"
    )


def test_aldehyde_frame_sentence_has_no_provenance_language(vcf_phase_artifacts):
    frames = vcf_phase_artifacts["phase_frames"]
    aldehyde_frame = next(f for f in frames if f["frame_id"] == "fat_phase_aldehyde_load_crowding")
    sentence = aldehyde_frame["sentence"].lower()
    banned_words = ("smoke", "smoked", "smoking", "cure", "cured", "curing", "roast", "grill", "braise", "prepar")
    hits = [w for w in banned_words if w in sentence]
    assert not hits, (
        f"fat_phase_aldehyde_load_crowding's sentence contains provenance/"
        f"preparation language ({hits}) — this frame has no preparation "
        f"gate in its trigger, so its sentence must not imply one either "
        f"(the exact category error four smoked_product_fat_phase "
        f"revisions made in the other direction): {aldehyde_frame['sentence']!r}"
    )
    assert not aldehyde_frame["pending_authoring"], (
        "fat_phase_aldehyde_load_crowding should be a fully authored, "
        "non-pending frame — it has real evidence and an authored sentence"
    )


def test_every_pair_row_carries_suppressed_reason_field(vcf_artifacts):
    pairs = vcf_artifacts["pairs"]
    assert pairs, "pairs.jsonl is empty"
    missing = [r for r in pairs if "suppressed_reason" not in r]
    assert not missing, (
        f"{len(missing)} pairs.jsonl row(s) have no suppressed_reason field at all — "
        f"every row must carry it (null when not suppressed), per the "
        f"'auditable and reversible, nothing deleted' instruction: {missing[:3]}"
    )
    suppressed = [r for r in pairs if r["suppressed_reason"]]
    for r in suppressed:
        assert r["suppressed_reason"] == "near_duplicate_same_spine", (
            f"unexpected suppressed_reason value {r['suppressed_reason']!r} on row {r}"
        )


def test_phase_frames_count_and_authored_split(vcf_phase_artifacts):
    frames = vcf_phase_artifacts["phase_frames"]
    assert len(frames) == 5, f"expected 5 phase frames (4 authored, 1 pending stub), got {len(frames)}"
    n_authored = sum(1 for f in frames if not f["pending_authoring"])
    n_pending = sum(1 for f in frames if f["pending_authoring"])
    assert n_authored == 4, f"expected 4 authored frames, got {n_authored}"
    assert n_pending == 1, f"expected 1 pending-authoring frame, got {n_pending}"


# --- Routing-tab-read gate (Build 1, backport, 2026-08-30) -----------------
#
# James: "the script reads the routing tab, and a test asserts every state
# is one of the three defined values with a non-empty authority field. A
# routing decision without a stated basis fails the build." This is the
# standing, corpus-wide anchor — ingest_protein_beef.load_routing_table()
# already gates its OWN family's rows at load time (and is the thing that
# actually stops a bad row from becoming ingested corpus behavior on a
# given run); this test re-checks the SAME source across every family
# present in the tab, not just whichever one last ran an ingestion, so a
# bad row added for a family that hasn't ingested yet is still caught.
#
# Skips (not fails) if the pinned workbook isn't present in this checkout
# — it's a large binary vendor file (see James, 2026-08-30: "commit it,
# it's decided by #5" — the workbook is code-path input once ingestion
# reads routing from it, so it belongs in the repo, but a shallow clone or
# CI mirror missing large-file support shouldn't hard-fail this suite over
# a fetch problem unrelated to the pipeline's own correctness).
def test_routing_table_states_and_authority_are_well_formed():
    if not ROUTING_XLSX.exists():
        pytest.skip(f"{ROUTING_XLSX} not present in this checkout")
    import openpyxl

    wb = openpyxl.load_workbook(ROUTING_XLSX, data_only=True, read_only=True)
    assert ROUTING_SHEET in wb.sheetnames, f"{ROUTING_XLSX} has no {ROUTING_SHEET!r} tab"
    ws = wb[ROUTING_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # title row
    header = next(rows_iter)
    rows = [dict(zip(header, r)) for r in rows_iter if any(r)]
    assert rows, f"{ROUTING_SHEET} has no data rows"

    bad_state = [
        (r.get("route_id"), r.get("family"), r.get("compound_name"), r.get("mr17_state"))
        for r in rows if r.get("mr17_state") not in VALID_MR17_ROUTING_STATES
    ]
    assert not bad_state, (
        f"{len(bad_state)} routing row(s) use an mr17_state outside "
        f"VALID_MR17_ROUTING_STATES {sorted(VALID_MR17_ROUTING_STATES)} — "
        f"(route_id, family, compound_name, state): {bad_state}"
    )

    bad_authority = [
        (r.get("route_id"), r.get("family"), r.get("compound_name"))
        for r in rows if not r.get("authority") or not str(r.get("authority")).strip()
    ]
    assert not bad_authority, (
        f"{len(bad_authority)} routing row(s) have no stated authority — a "
        f"routing decision without a basis fails the build, per spec "
        f"(route_id, family, compound_name): {bad_authority}"
    )

    # A compound can legitimately be routed more than once within a family
    # (egg: Dibutyl phthalate — chicken yolk and duck salted yolk each have
    # their own row) since load_routing_table() stores one decision per
    # compound identity, not per product. Two rows that AGREE are
    # redundant, not conflicting. Two that DISAGREE on state or authority
    # are a real ambiguity a human has to resolve, not something the
    # ingestion script should silently pick a winner for.
    seen: dict[tuple, dict] = {}
    conflicts = []
    for r in rows:
        key = (r.get("family"), r.get("compound_name"))
        entry = {"mr17_state": r.get("mr17_state"), "authority": str(r.get("authority") or "").strip()}
        if key in seen and seen[key] != entry:
            conflicts.append((key, seen[key], entry, r.get("route_id")))
        seen[key] = entry
    assert not conflicts, (
        f"routing rows disagree for the same (family, compound_name) — "
        f"(key, first_seen, conflicting_row, conflicting_route_id): {conflicts}. "
        f"This may be a genuine mistake, or it may mean this compound's "
        f"disposition legitimately varies by product and routing needs to "
        f"become per-product rather than per-compound for it — don't resolve "
        f"by editing one row to match the other without checking which."
    )
