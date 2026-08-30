"""
VCF Compound Layer — Build 1 (Sequenced Builds spec), "apply" step.

Run from the repo root, AFTER human review of spine_cluster_proposals.xlsx
and BEFORE build_vcf_resolution_policy.py:
    python pipeline/scripts/apply_spine_clusters.py

Reads:  pipeline/artifacts/vcf/spine.jsonl                    (Step 3 output)
        pipeline/artifacts/vcf/spine_cluster_proposals.xlsx    (human review)
Writes: pipeline/artifacts/vcf/spine.jsonl        (same file, in place —
        merges each approved cluster's member spine entries into one)
        pipeline/artifacts/vcf/meta.json          (adds a "spine_clustering"
        block)

This is the one-time "apply" half of spine_cluster_candidates.py, which
only ever proposed clusters for review and never wrote to spine.jsonl.
James approved all 41 proposed clusters in the first pass (2026-08-30) —
every row in that sheet, not a subset — so this script originally did not
read the approve/override columns at all; it merged every cluster the
sheet listed. The second pass (same day, post binomial-parser fix, 44
proposed clusters) was genuinely partial — one outright reject
(ARTICHOKE/JERUSALEM ARTICHOKE, later removed at the source by a
containment-rule fix) plus 7 more (STRAWBERRY/STRAWBERRY JAM, BARLEY/
MALTED BARLEY, RICE/RICE BRAN, SOYBEAN/DEFATTED SOYBEAN, MALT/PEATED MALT,
the three-species berry bucket, BROCCOLI/CAULIFLOWER) rejected as
different-products-sharing-a-name rather than near-duplicates, plus 15
policy overrides beyond wine/whisky — so this script now reads the
approve (y/n) and policy_override columns for real, per cluster (every
row of one cluster must agree, checked below).

--- What "merge" means here ---

Clustering assigns one shared spine_id to near-duplicate products (WINE,
RED WINE, PORT WINE, ...) so build_vcf_pairs.py's existing same-spine
exclusion at candidate generation — already unconditional there — finally
reaches them. It does NOT merge or average their compound profiles; each
product keeps its own profile row in profiles.jsonl, keyed by its own
vcf_product_id. Only the `spine_id` a profile resolves to changes.

Because this script runs on spine.jsonl AFTER build_vcf_spine.py, and
BEFORE build_vcf_resolution_policy.py, `policy`/`default_member`/
`resolution_confidence` on every entry (merged or not) are reset to None
here, matching build_vcf_spine.py's own contract that those fields belong
to Step 9. Do not run build_vcf_spine.py again after this script without
re-running this one too — build_vcf_spine.py rebuilds spine.jsonl from
vcf_product_parse.jsonl's base_ingredient grouping from scratch and would
silently erase every cluster merge.

--- Canonical spine_id / display_name pick ---

A cluster's constituent products usually come from DIFFERENT pre-cluster
spine entries (one per base_ingredient) — merging must settle on ONE
spine_id and display_name to survive. Rule: among the cluster's TOUCHED
SPINE ENTRIES (not raw member rows), the one with the SHORTEST
`base_ingredient` string wins (tie-break: lowest spine_id string); its
spine_id/display_name become the merged entry's.

base_ingredient, not raw_name, is the length comparison's input
deliberately: raw_name still carries whatever parenthetical qualifier
Step 2 left on it (a binomial suffix, "(cooked)", "and/or STALKS") and
that qualifier's presence is an accident of which member happens to have
one recorded, not a signal of genericness — CELERY (Apium graveolens L.)
is 28 characters, longer than CELERY LEAVES and/or STALKS at 27, even
though "celery" is obviously the base and "leaves and/or stalks" the
qualified part. base_ingredient is Step 2's own already-stripped name
(the same field build_vcf_spine.py itself groups on), so the comparison
sees "celery" (6) vs "celery leaves and/or stalks" (27) — the right
answer, and the right one generally: "wine" beats "red wine", "whisky"
beats "malt whisky", "rice" beats "rice bran", "soybean" beats "defatted
soybean", "coconut" beats "coconut meat". This isn't a semantic guarantee
(see build_vcf_resolution_policy.py's own default_member docstring for
the same caveat pattern) — a cluster with no bare/generic member at all
(e.g. THYMUS ZYGIS L. / THYMUS, OTHER TYPES: neither base_ingredient is
"thymus" alone) still picks one specific name over another on the same
length rule, which is the best available answer, not a claim that name is
uniquely correct. Flagged in the printed report, not asserted correct
without a look.

--- Policy overrides ---

James: Wine and Whisky both take `category`, not the `expand` the
resolution-policy rule's mechanical fields would produce (no form/
cure_state/binomial variation among the merged members — "a range across
port, madeira, dessert wine and white wine describes nothing a chef can
use"). This script does NOT itself set policy (that's Step 9's job) — it
records which merged spine_ids need an override so
build_vcf_resolution_policy.py's own POLICY_OVERRIDES dict (added there,
mirroring its existing DEFAULT_MEMBER_OVERRIDES pattern) can apply it and
report it audibly. Recorded here too, in meta.json, so the two scripts'
bookkeeping agrees.
"""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = REPO_ROOT / "pipeline" / "artifacts" / "vcf"
SPINE_JSONL = OUT_DIR / "spine.jsonl"
PROPOSALS_XLSX = OUT_DIR / "spine_cluster_proposals.xlsx"
PARSE_JSONL = OUT_DIR / "vcf_product_parse.jsonl"
META_JSON = OUT_DIR / "meta.json"


def compute_binomial_coverage() -> tuple[int, int]:
    """Live count, not a frozen constant — see the module-level note above
    on why a hardcoded number in an artifact writer is a lie waiting to
    happen. Returns (n_with_binomial, n_culinary_total)."""
    products = [json.loads(l) for l in PARSE_JSONL.read_text().splitlines() if l.strip()]
    culinary = [p for p in products if p["class"] == "culinary"]
    with_binomial = [p for p in culinary if p.get("binomial")]
    return len(with_binomial), len(culinary)


def count_coverage_gap_pairs(xlsx_path: Path) -> dict:
    """Live count of the coverage-gap sheets' data rows. Split 2026-08-30
    (James, after the 70-row asymmetric triage) into 'Needs Review' (a
    textual signal ties the untagged side to the tagged one — actually
    worth a look) and 'No Signal (Info)' (pure compound-overlap
    coincidence, or neither side genus-confirmable at all — informational,
    no review expected) — reporting one combined number as before would
    re-flatten exactly the distinction the split exists to preserve.
    Falls back to the old single-sheet name if a file predates the split,
    so this doesn't crash on an older artifact."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    def count_sheet(name):
        if name not in wb.sheetnames:
            return None
        ws = wb[name]
        return sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[0] is not None)

    if "Coverage Gap - Manual Call" in wb.sheetnames:
        n = count_sheet("Coverage Gap - Manual Call")
        return {"needs_review": None, "no_signal_info": None, "total": n, "pre_split_format": True}

    n_review = count_sheet("Coverage Gap - Needs Review") or 0
    n_info = count_sheet("Coverage Gap - No Signal (Info)") or 0
    return {"needs_review": n_review, "no_signal_info": n_info, "total": n_review + n_info, "pre_split_format": False}

# Superseded 2026-08-30 (second pass): policy overrides now come from the
# reviewed sheet's own "policy_override" column (wine/whisky are still in
# there, alongside 13 more — plant-part and processing-state clusters
# James decided should show members rather than average them, same
# reasoning as wine/whisky). Reading it from the review artifact instead
# of a hardcoded set here means a future family's overrides don't need a
# code change to take effect. See load_clusters() and main()'s
# policy_overrides dict. Binomial coverage and the coverage-gap count are
# likewise no longer hardcoded module constants — see the same lesson
# applied in build_vcf_pairs.py's near-duplicate report: a frozen number
# in an artifact writer silently re-asserts itself as true on every future
# rebuild. Both are now computed live in main() from the current corpus.

# James, 2026-08-30 (second review pass): the shortest-base_ingredient
# canonical-name rule (see module docstring) picks a real generic root for
# containment clusters (wine, whisky, rice, ...) but picks an ARBITRARY
# sibling name for a pure-genus cluster where no member's base_ingredient
# is actually a generic term the others contain — Allium's cluster has no
# member called "allium" in English at all, so shortest-string picked
# "leek" out of onion/garlic/chive/shallot/leek by coincidence of word
# length, not genericness. A chef who types "garlic" and gets told it
# resolves to "Leek" reads that as the system confusing two different
# vegetables — a trust-costing defect, not a cosmetic one, per James.
#
# The fix James specified is explicit: don't invent a curated botanical
# category name (that's the "allium family" move, rejected), and don't
# keep the arbitrary sibling name either — set display_name=null and let
# the consuming lens render whatever alias the chef actually queried
# (already present in this entry's merged `aliases` list). This is an
# explicit, named override list — the same "deterministic lookup, not a
# model decision" pattern as POLICY_OVERRIDE_SPINE_IDS above and
# DEFAULT_MEMBER_OVERRIDES in build_vcf_resolution_policy.py — not a
# general "no true generic" detector run automatically over all 41
# clusters. A general detector would also flag SAGE (its cluster includes
# "SALVIA SPECIES", whose base_ingredient is "salvia species", not
# literally containing "sage") even though "Sage" reads fine to a chef
# (sage/salvia are the recognizable common/Latin name pair) — that's a
# judgment call, and this list only encodes the specific clusters James
# reviewed and named, not an inferred rule that could silently relabel
# others without review.
DISPLAY_NAME_NULL_OVERRIDES = {
    "culin:leek": "allium",       # ALLIUM cluster: onion/garlic/chive/leek/shallot/nira/nobiru
    "culin:bilberry": "vaccinium",  # blueberry/cranberry/lingonberry/bilberry
    "culin:kuini": "mangifera",   # mango/bachang/bambangan/binjai/kuini
    "culin:chekur": "alpinia",    # galangal varieties
    "culin:summer_savory": "satureja",  # summer/winter savory
}
VALID_POLICY_OVERRIDE_VALUES = {"single", "expand", "category"}


def load_clusters(xlsx_path: Path):
    """Reads the reviewed proposals sheet for real: only clusters marked
    approve='y' are returned for merging; a policy_override value (must be
    one of VALID_POLICY_OVERRIDE_VALUES — a REJECT: note on a rejected row
    never reaches here) is collected per approved cluster. Every row of a
    given cluster_id must agree on both columns — a human editing the
    sheet by hand could otherwise leave one row approved and another not,
    which is exactly the kind of silent inconsistency this read exists to
    catch rather than average away.

    Returns (clusters, policy_overrides, rejected):
      clusters         {cluster_id: [(vcf_product_id, raw_name), ...]} — approved only
      policy_overrides {cluster_id: override_value} — only where set
      rejected         [{"cluster_id":..., "reason":...}] — for the report
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Cluster Proposals"]
    raw_rows: dict[str, list[tuple]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        cluster_id, vcf_product_id, raw_name = row[0], row[1], row[2]
        approve, override = row[9], row[10]
        raw_rows[cluster_id].append((vcf_product_id, raw_name, approve, override))

    clusters: dict[str, list[tuple[int, str]]] = {}
    policy_overrides: dict[str, str] = {}
    rejected = []
    for cluster_id, rows in raw_rows.items():
        approvals = {(r[2] or "").strip().lower() for r in rows}
        if len(approvals) != 1:
            raise SystemExit(
                f"{cluster_id}: rows disagree on approve (y/n): {approvals} — "
                f"every row of one cluster must carry the same decision."
            )
        approve = approvals.pop()
        if approve not in ("y", "n"):
            raise SystemExit(
                f"{cluster_id}: approve column is {approve!r}, not 'y' or 'n' — "
                f"every row needs an explicit decision before this can run."
            )
        overrides = {(r[3] or "").strip() for r in rows}
        if len(overrides) != 1:
            raise SystemExit(f"{cluster_id}: rows disagree on policy_override: {overrides}")
        override = overrides.pop()

        if approve == "n":
            rejected.append({"cluster_id": cluster_id, "reason": override or "(no reason recorded)"})
            continue

        clusters[cluster_id] = [(r[0], r[1]) for r in rows]
        if override:
            if override not in VALID_POLICY_OVERRIDE_VALUES:
                raise SystemExit(
                    f"{cluster_id}: policy_override {override!r} is not one of "
                    f"{sorted(VALID_POLICY_OVERRIDE_VALUES)} — an approved cluster's "
                    f"override must be a real policy value, not free text (REJECT: "
                    f"notes belong on approve='n' rows only)."
                )
            policy_overrides[cluster_id] = override

    return clusters, policy_overrides, rejected


def main():
    if not SPINE_JSONL.exists():
        raise SystemExit(f"{SPINE_JSONL} not found — run build_vcf_spine.py first.")
    if not PROPOSALS_XLSX.exists():
        raise SystemExit(f"{PROPOSALS_XLSX} not found — run spine_cluster_candidates.py first.")

    entries = [json.loads(l) for l in SPINE_JSONL.read_text().splitlines() if l.strip()]
    entry_by_spine_id = {e["spine_id"]: e for e in entries}
    spine_id_by_product_id: dict[int, str] = {
        m["vcf_product_id"]: e["spine_id"] for e in entries for m in e["members"]
    }

    clusters, policy_overrides, rejected = load_clusters(PROPOSALS_XLSX)

    merge_report = []
    policy_override_report = []
    already_merged = []
    consumed_spine_ids: set[str] = set()

    for cluster_id, members in clusters.items():
        product_ids = [pid for pid, _ in members]
        touched_spine_ids = sorted({spine_id_by_product_id[pid] for pid in product_ids})

        if len(touched_spine_ids) == 1:
            # base_ingredient parsing already put every member of this
            # cluster under one spine entry — nothing to merge structurally,
            # but the policy-override bookkeeping below still needs to see
            # it if it's wine/whisky (not the case in this pass).
            already_merged.append({"cluster_id": cluster_id, "spine_id": touched_spine_ids[0]})
            spine_id = touched_spine_ids[0]
        else:
            # Canonical pick: shortest base_ingredient among the cluster's
            # touched SPINE ENTRIES (tie-break lowest spine_id string) —
            # see module docstring for why base_ingredient, not raw_name.
            canonical_spine_id = min(
                touched_spine_ids,
                key=lambda sid: (len(entry_by_spine_id[sid]["base_ingredient"]), sid),
            )
            canonical_entry = entry_by_spine_id[canonical_spine_id]
            canonical_raw_name = next(
                raw_name for pid, raw_name in members
                if spine_id_by_product_id[pid] == canonical_spine_id
            )

            other_spine_ids = [sid for sid in touched_spine_ids if sid != canonical_spine_id]
            other_entries = [entry_by_spine_id[sid] for sid in other_spine_ids]

            product_groups = {canonical_entry["product_group"]} | {
                e["product_group"] for e in other_entries
            }
            if len(product_groups) != 1:
                raise SystemExit(
                    f"cluster {cluster_id} spans product_group values "
                    f"{sorted(product_groups)} across {touched_spine_ids} — "
                    f"needs a human call, not a silent merge."
                )

            merged_members = list(canonical_entry["members"])
            merged_member_ids = {m["vcf_product_id"] for m in merged_members}
            for e in other_entries:
                for m in e["members"]:
                    if m["vcf_product_id"] not in merged_member_ids:
                        merged_members.append(m)
                        merged_member_ids.add(m["vcf_product_id"])

            merged_aliases = sorted(
                set(canonical_entry["aliases"]) | {a for e in other_entries for a in e["aliases"]}
            )
            class_counts: dict[str, int] = defaultdict(int)
            for m in merged_members:
                class_counts[m["class"]] += 1

            canonical_entry["members"] = merged_members
            canonical_entry["aliases"] = merged_aliases
            canonical_entry["n_members"] = len(merged_members)
            canonical_entry["class_counts"] = dict(class_counts)
            # Step 9's job, not this script's — reset so a stale pre-merge
            # value (e.g. "single") can never survive under new membership.
            canonical_entry["policy"] = None
            canonical_entry["default_member"] = None
            canonical_entry["resolution_confidence"] = None
            canonical_entry["spine_cluster_id"] = cluster_id
            canonical_entry["spine_cluster_merged_from"] = other_spine_ids
            if canonical_spine_id in DISPLAY_NAME_NULL_OVERRIDES:
                canonical_entry["display_name"] = None
                canonical_entry["display_name_null_reason"] = (
                    f"No member of this {DISPLAY_NAME_NULL_OVERRIDES[canonical_spine_id]} "
                    "genus cluster is more generic than its siblings — the shortest-"
                    "base_ingredient tiebreak would otherwise surface one arbitrary "
                    "species name as if it represented the whole cluster (e.g. "
                    "'Leek' for onion/garlic/chive/shallot). Render whatever alias "
                    "the caller queried instead; see `aliases` on this entry."
                )

            consumed_spine_ids.update(other_spine_ids)
            spine_id = canonical_spine_id

            merge_report.append(
                {
                    "cluster_id": cluster_id,
                    "canonical_spine_id": canonical_spine_id,
                    "canonical_raw_name": canonical_raw_name,
                    "merged_spine_ids": other_spine_ids,
                    "n_members_after_merge": len(merged_members),
                }
            )

        if cluster_id in policy_overrides:
            policy_override_report.append(
                {"cluster_id": cluster_id, "spine_id": spine_id, "policy": policy_overrides[cluster_id]}
            )

    remaining_entries = [e for e in entries if e["spine_id"] not in consumed_spine_ids]

    with open(SPINE_JSONL, "w") as f:
        for e in remaining_entries:
            f.write(json.dumps(e, ensure_ascii=False) + "\n")

    n_with_binomial, n_culinary = compute_binomial_coverage()
    n_coverage_gap_pairs = count_coverage_gap_pairs(PROPOSALS_XLSX)
    policy_overrides_by_spine_id = {
        r["spine_id"]: r["policy"] for r in policy_override_report
    }

    meta = json.loads(META_JSON.read_text()) if META_JSON.exists() else {}
    meta["spine_clustering"] = {
        "source": "spine_cluster_candidates.py + human review (spine_cluster_proposals.xlsx)",
        "n_clusters_approved": len(clusters),
        "n_clusters_rejected": len(rejected),
        "clusters_rejected": rejected,
        "n_clusters_merged_new_spine_entries": len(merge_report),
        "n_clusters_already_single_spine_entry": len(already_merged),
        "n_spine_entries_before": len(entries),
        "n_spine_entries_after": len(remaining_entries),
        "policy_overrides_applied_to": policy_overrides_by_spine_id,
        "policy_override_reason": (
            "Second review pass (2026-08-30): 15 clusters given an explicit "
            "policy_override in spine_cluster_proposals.xlsx rather than "
            "left to build_vcf_resolution_policy.py's mechanical rule — "
            "wine and whisky (no form/cure_state/binomial variation among "
            "merged members; the real distinction, style/region, lives "
            "only in free text in raw_name), plant-part clusters (lovage, "
            "caraway, myrtle, pimento, blackcurrant, parsley — different "
            "tissues, different profiles), and processing-state clusters "
            "(tea, olive, mate — green/black/roasted are not the same "
            "product to average over). James: these should show members "
            "or ask which, not be collapsed into a range. Recorded here "
            "as {spine_id: policy}; build_vcf_resolution_policy.py should "
            "read this block directly rather than keep its own second copy "
            "of the same mapping."
        ),
        "binomial_coverage": f"{n_with_binomial}/{n_culinary}",
        "binomial_coverage_note": (
            f"{n_with_binomial}/{n_culinary} ({n_with_binomial/n_culinary:.1%}) of "
            "VCF's culinary products carry a binomial field, computed live "
            "from vcf_product_parse.jsonl at apply time (not a frozen "
            "constant — see the module note on why that broke before). "
            "The clustering rule's genus half only has data to work with "
            "for that slice — the rest can only qualify via name "
            "containment. This is a floor set by data coverage, not a "
            "ceiling on how many real near-duplicates exist in the corpus."
        ),
        "clustering_pass": "partial",
        "clustering_pass_note": (
            f"{len(clusters)} clusters approved ({len(rejected)} rejected) "
            "this pass; the coverage-gap counts below are read live from "
            "spine_cluster_proposals.xlsx, not a frozen constant. Split "
            "2026-08-30 (James) into 'Needs Review' (a textual signal ties "
            "the untagged side to the tagged one) and 'No Signal (Info)' "
            "(pure compound-overlap coincidence, or neither side "
            "genus-confirmable at all) — see "
            "n_coverage_gap_pairs_pending_review for both counts "
            "separately; conflating them into one number is exactly the "
            "thing the split was meant to stop. Duplicate suppression "
            "downstream (pairs.jsonl) is incomplete by that known amount; "
            "anyone reading the rebaselined median top-1 score after this "
            "pass needs to know that, not just the after-number. Note also "
            "that this pass's baseline is VCF-only (521 products) — beef "
            "is temporarily excluded from compounds/profiles/pairs, so "
            "this number is not comparable to any prior baseline that "
            "included beef's 4 products."
        ),
        "n_coverage_gap_pairs_pending_review": n_coverage_gap_pairs,
        "clusters_merged": merge_report,
        "clusters_already_single_spine_entry": already_merged,
    }
    META_JSON.write_text(json.dumps(meta, indent=2))

    print(f"Spine entries: {len(entries)} -> {len(remaining_entries)}")
    print(f"Clusters approved: {len(clusters)}  |  rejected: {len(rejected)}")
    for r in rejected:
        print(f"  REJECTED {r['cluster_id']:<14} {r['reason']}")
    print(f"Clusters requiring an actual merge: {len(merge_report)}")
    print(f"Clusters already single-spine-entry pre-merge: {len(already_merged)}")
    print(f"Policy overrides applied this pass: {policy_overrides_by_spine_id}")
    print(f"Binomial coverage (live): {n_with_binomial}/{n_culinary}")
    if n_coverage_gap_pairs.get("pre_split_format"):
        print(f"Coverage-gap pairs pending review (live, pre-split sheet format): "
              f"{n_coverage_gap_pairs['total']}")
    else:
        print(f"Coverage-gap pairs (live): {n_coverage_gap_pairs['needs_review']} needs review, "
              f"{n_coverage_gap_pairs['no_signal_info']} no-signal/informational "
              f"({n_coverage_gap_pairs['total']} total)")
    for m in merge_report:
        print(f"  {m['cluster_id']:<14} -> {m['canonical_spine_id']:<24} "
              f"({m['n_members_after_merge']} members, merged {m['merged_spine_ids']})")


if __name__ == "__main__":
    main()
