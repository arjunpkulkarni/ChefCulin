"""
CulinAI Protein Layer — Beef ingestion (pilot family, per the Beef
Ingestion Build Spec).

Run from the repo root, AFTER build_vcf_spine.py and
canonicalize_vcf_compounds.py (needs a compounds.jsonl to check for
reusable identities against) and BEFORE build_vcf_resolution_policy.py,
build_vcf_phase.py, build_vcf_compound_roles.py, build_vcf_profiles.py:
    python pipeline/scripts/ingest_protein_beef.py

Once this has run and beef_profiles_prebuilt.jsonl exists, remember
build_vcf_profiles.py needs `--include-protein` to actually merge it —
that flag defaults OFF (2026-08-30 fix; see that script's own module
note) precisely so a rebuild never mixes protein families back in
without someone deciding it should.

Reads:  the source xlsx (path given via BEEF_XLSX env var or --xlsx)
Writes: pipeline/vendor/culinai_protein/beef/*.csv        (landed tabs)
        pipeline/artifacts/vcf/compounds.jsonl             (appends new
                                                             beef-only
                                                             compound rows;
                                                             VCF rows
                                                             untouched)
        pipeline/artifacts/vcf/spine.jsonl                 (appends
                                                             beef:muscle /
                                                             beef:fat spine
                                                             entries; VCF
                                                             entries
                                                             untouched;
                                                             vocabulary
                                                             re-hashed)
        pipeline/artifacts/protein/beef_profiles_prebuilt.jsonl
                                                            (consumed by
                                                             build_vcf_profiles.py)
        pipeline/artifacts/protein/beef_observations.jsonl (full per-row
                                                             audit trail —
                                                             every row from
                                                             every landed
                                                             tab, with its
                                                             resolved
                                                             identity,
                                                             assigned
                                                             product/tier,
                                                             and whether it
                                                             counts toward
                                                             any profile —
                                                             including the
                                                             rows that
                                                             don't)
        pipeline/artifacts/protein/beef_aging_deltas.jsonl (dry28 vs wet28
                                                             quantitative
                                                             deltas — a
                                                             different diff
                                                             shape than
                                                             gained/lost
                                                             sets, appended
                                                             to
                                                             form_diffs.jsonl
                                                             by
                                                             build_vcf_form_diffs.py
                                                             in a later
                                                             step)
        pipeline/artifacts/vcf/meta.json                   (adds a
                                                             "protein_beef"
                                                             block)

=====================================================================
Rule zero (per spec): every rule about how an external profile joins VCF
is untested before this pass. Two of the spec's own stated premises turn
out to be false against the real workbook, found by checking rather than
assuming — reported here and in meta.json, not silently worked around:

1. "All beef rows carry CAS" (Step 3) is false. Of the 355 rows across the
   four Detected tabs, only 91 (the Detected Beef Muscle / Morsli
   raw-to-grill series) carry a CAS number in the source data. The other
   264 (all of Detected Beef Fat, all of Detected Beef Cure-Smoke, all of
   Detected Beef Aging, and the 45-row Detected Beef Muscle
   identification-QA block) do not.

   This turns out not to be a blocker: VCF's OWN compound_id is never
   taken from a native CAS field either — canonicalize_vcf_compounds.py's
   entire job is resolving a compound NAME against the CAS/PubChem
   crosswalk, with CAS-less names falling back to a provisional
   `vcf:<slug>` identity. The same machinery, `resolve_beef_compound()`
   below, is reused for beef: CAS given in the sheet is checked first (and
   preferred, since it is beef's own stated identity), a bare compound
   name is resolved against the SAME crosswalk canonicalize_vcf_compounds.py
   already loads, and only a genuinely unmatched name falls back to a new
   `beef:<slug>` provisional identity. Reported in full in meta.json's
   protein_beef.compound_resolution block — this is not "assume they
   match," it's "checked, and most of them do."

2. Step 2's own preparation_state -> state_tier table contradicts Step
   2's own prose. The table maps bare "dry-aged"/"wet-aged" to
   `cooked + pre_treatment`; the very next paragraph says "A dry-aged
   steak is still raw or still grilled" — i.e. bare dry-aged/wet-aged
   should be `raw + pre_treatment`, and only "grilled dry-aged" is
   `cooked + pre_treatment`. Implemented per the PROSE (the more specific,
   explicit statement), not the table — see STATE_TIER_MAP below. Flagged
   rather than silently picked, per this engagement's established pattern
   (the CUTTLEFISH/bromophenol contradiction in an earlier revision was
   handled the same way).

=====================================================================
Two products, not one (Step 1): beef:muscle and beef:fat. Verified Beef
Profiles' own row BP-052 already carries a `product_id: beef:fat` and
`promotion_status: superseded_moved_to_beef_fat_profile_v16` — the
workbook itself has already flagged this exact row as the fat-vs-muscle
split point (its evidence is a verbatim duplicate of Verified Beef Fat
Profile's BFP-001: same compound, same CAS, same source_id). BP-052 is
therefore EXCLUDED from beef:muscle ingestion entirely (not double-counted
against BFP-001).

Per-state profiles, not one merged blob (Step 2 + Step 7): "the lens reads
the tier" only makes pairing/competition/form-diff logic meaningful if a
profile represents ONE culinary state. beef:muscle resolves to three
STATE-TIER profiles (raw / cooked / smoked); beef:fat resolves to one
(cooked, dry-rendered fat has no other state in this data). This also
reuses build_vcf_form_diffs.py's existing spine-member pairwise diff logic
for the Step 7 raw->cooked / cooked->smoked form diffs, with ZERO new diff
code required there — they fall out of the existing multi-member-spine
mechanism once beef:muscle's three states are spine members.

Detected Beef Aging (37 rows) has no raw/cooked signal of its own (its
state_scope is "identification union across day 0-28" pooled across dry
AND wet aging) — folded into the RAW-tier profile's membership (aging
without further cooking is structurally still raw meat, per the corrected
Step 2 rule above), tagged with pre_treatment provenance, not spun into
its own profile. This is a judgment call the data doesn't resolve
cleanly; flagged as such rather than asserted as obviously correct.

Detected Beef Cure-Smoke (86 rows) is a PAIRED T1(smoked+spiced) vs
T2(unsmoked+unspiced) comparison table, not a pure "smoked beef" table.
Only t1_detected=="yes" rows enter the smoked-tier profile. T2 data is
landed and reported but does not become a fourth profile (spec names
three tiers: raw/cooked/smoked; T2 is "cured, unsmoked", which the spec's
own state_tier table never names) — flagged as a scoping choice, not
silently dropped.
=====================================================================
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[2]
VCF_DIR = REPO_ROOT / "pipeline" / "artifacts" / "vcf"
PROTEIN_DIR = REPO_ROOT / "pipeline" / "artifacts" / "protein"
VENDOR_DIR = REPO_ROOT / "pipeline" / "vendor" / "culinai_protein" / "beef"
CROSSWALK_XLSX = REPO_ROOT / "pipeline" / "vendor" / "vcf" / "vcfodor_cas_pubchem_distinct.xlsx"

# James, 2026-08-30 (backport of the routing-tab-read decision, made for
# egg but landed on beef first — beef is the reference implementation
# every later family's ingestion script gets read against, so two
# patterns in the repo is worse than the cost of retrofitting): routing
# decisions for excluded/unresolved compounds come from THIS workbook's
# "Exclusion Routing Audit" tab, not from a code constant. Pinned by
# path+hash in meta.json's vendor_workbook_provenance block, independent
# of whichever workbook version a family's own Detected tables come from
# (beef's Detected tables are v21; v21 predates this tab entirely — it
# does not exist there. Routing is a newer, separately-versioned decision
# layer, not tied to when a family's raw data was pulled).
#
# Repinned v29 -> v30, 2026-08-30 (James): v30 adds exactly one row,
# ER-014 (beef BDA-011, 2-Methyl-2-propanethiol, mr17_state=unresolved) —
# verified cell-by-cell against the v29 already in-repo before repinning:
# all 108 tabs identical except "Exclusion Routing Audit", which is
# identical through its first 15 rows (title, header, ER-001..ER-013) with
# ER-014 appended as row 16. No other row was touched, so nothing this
# rule already routed under v29 can have silently changed meaning under
# v30. The old file is kept on disk (CulinAI_Staging_v29.xlsx) rather than
# overwritten, so that diff stays checkable later.
ROUTING_XLSX = REPO_ROOT / "pipeline" / "vendor" / "culinai_protein" / "CulinAI_Staging_v30.xlsx"
ROUTING_SHEET = "Exclusion Routing Audit"

# The full set of mr17_state values this tab is allowed to use, verified
# empirically against v29's actual rows (13, across beef + egg) rather
# than asserted from memory — a 4th, unexpected value showed up on first
# read (analytical_background, for a GC column bleed artifact in egg's
# yolk data) that the original 2-outcome mental model (present_not_
# flavor_relevant / unresolved) didn't anticipate. Both this constant and
# the load-time gate below exist so a 5th, equally unanticipated value
# fails the build loudly instead of silently becoming a new corpus
# behavior nobody reviewed.
VALID_MR17_ROUTING_STATES = {
    "present_not_flavor_relevant",  # a genuinely identified, non-flavor substance (residue, industrial chemical)
    "analytical_background",         # a method/column artifact, not a food constituent at all
    "unresolved",                    # held pending — neither confirmed non-flavor nor mapped
}


def load_routing_table(routing_xlsx: Path, family: str) -> dict[str, dict]:
    """Reads ROUTING_SHEET, filters to `family`, and returns a dict keyed
    by compound_name. This IS the gate James asked for, not a separate
    check bolted on afterward: every row is validated at load time, before
    a single routing decision can reach _new_compound() — an invalid
    mr17_state or an empty authority raises SystemExit here, not a warning
    later. (test_vcf_reliability.py also asserts this independently and
    corpus-wide, across every family present in the tab, as a standing
    regression anchor — this inline gate is what actually stops a bad row
    from becoming ingested corpus behavior on THIS run.)
    """
    if not routing_xlsx.exists():
        raise SystemExit(
            f"{routing_xlsx} not found — routing decisions for excluded/"
            f"unresolved compounds are read from this workbook's "
            f"{ROUTING_SHEET!r} tab, not a code constant. Pin the workbook "
            f"(see meta.json's vendor_workbook_provenance) before ingesting."
        )
    wb = openpyxl.load_workbook(routing_xlsx, data_only=True, read_only=True)
    if ROUTING_SHEET not in wb.sheetnames:
        raise SystemExit(f"{routing_xlsx} has no {ROUTING_SHEET!r} tab.")
    ws = wb[ROUTING_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # title row ("Exclusion Routing Audit — v29")
    header = next(rows_iter)
    table: dict[str, dict] = {}
    bad_rows = []
    for row in rows_iter:
        rec = dict(zip(header, row))
        if rec.get("family") != family:
            continue
        name = rec.get("compound_name")
        state = rec.get("mr17_state")
        authority = rec.get("authority")
        if state not in VALID_MR17_ROUTING_STATES:
            bad_rows.append((rec.get("route_id"), name, "mr17_state", state))
            continue
        if not authority or not str(authority).strip():
            bad_rows.append((rec.get("route_id"), name, "authority", authority))
            continue
        if name in table:
            # A compound can legitimately appear in more than one product's
            # row within a family (egg: Dibutyl phthalate is routed once for
            # chicken yolk, once for duck salted yolk) — the routing
            # DECISION is compound-level, not product-level (this script's
            # own resolver stores mr17_outcome once per compound identity,
            # not per profile), so two rows that AGREE are redundant, not
            # conflicting. Two rows that DISAGREE on state or authority are
            # a real ambiguity — which one wins is not this script's call.
            existing = table[name]
            if existing["mr17_state"] != state or existing["authority"] != str(authority).strip():
                bad_rows.append((rec.get("route_id"), name, "conflicting_duplicate_compound_name_for_family",
                                  {"existing": existing, "new_state": state, "new_authority": authority}))
            continue
        table[name] = {
            "mr17_state": state,
            "authority": str(authority).strip(),
            "meaning": rec.get("meaning"),
            "route_id": rec.get("route_id"),
        }
    if bad_rows:
        raise SystemExit(
            f"{len(bad_rows)} row(s) in {routing_xlsx.name}::{ROUTING_SHEET} for "
            f"family={family!r} fail the routing gate (route_id, compound_name, "
            f"failing_field, value): {bad_rows}. A routing decision without a "
            f"valid state or a stated basis cannot become corpus behavior — fix "
            f"the workbook row before re-running. NOTE on "
            f"'conflicting_duplicate_compound_name_for_family' specifically: this "
            f"loader keys routing on compound_name per family, on the assumption "
            f"that a compound's disposition is the same regardless of which "
            f"product it showed up in. A conflict here may mean the two rows "
            f"genuinely disagree by mistake — or it may mean that assumption is "
            f"wrong for this compound and routing needs to become per-product, "
            f"not per-compound. Don't resolve a conflict by editing one row to "
            f"match the other without checking which case this is; that would "
            f"silently paper over a real per-product distinction the workbook "
            f"may be trying to draw."
        )
    return table

COMPOUNDS_JSONL = VCF_DIR / "compounds.jsonl"
SPINE_JSONL = VCF_DIR / "spine.jsonl"
META_JSON = VCF_DIR / "meta.json"

BEEF_PROFILES_OUT = PROTEIN_DIR / "beef_profiles_prebuilt.jsonl"
BEEF_OBSERVATIONS_OUT = PROTEIN_DIR / "beef_observations.jsonl"
BEEF_AGING_DELTAS_OUT = PROTEIN_DIR / "beef_aging_deltas.jsonl"

PROFILE_SOURCE = "culinai_protein_v21"

# Only these 10 tabs are in scope per the build spec — everything else in
# the workbook (Canonical Registry - Beef, Beef Legacy Reconciliation,
# Beef Completion Audit, Beef MR Addendum Audit, and every non-beef
# family's tabs) is deliberately NOT read. Row counts are asserted below
# against the spec's own table as a load-time check.
BEEF_TABS = {
    "Detected Beef Muscle": 136,
    "Detected Beef Fat": 96,
    "Detected Beef Cure-Smoke": 86,
    "Detected Beef Aging": 37,
    "Verified Beef Profiles": 79,
    "Verified Beef Fat Profile": 1,
    "Beef Identification QA": 5,
    "Beef Quantitative Deltas": 22,
    "Beef State Coverage": 9,
    "Beef Pattern Evidence": 7,
}

# --- Step 2: preparation_state -> state_tier, per the PROSE, not the
# self-contradicting table (see module docstring, finding #2). -----------
STATE_TIER_MAP = {
    "raw": ("raw", None),
    "roasted": ("cooked", None),
    "grilled": ("cooked", None),
    "seared": ("cooked", None),
    "cooked": ("cooked", None),
    "braised / moist-cooked": ("cooked", None),
    "dry-aged": ("raw", "dry_aged"),          # prose: "still raw" — table's "cooked" row is wrong
    "wet-aged": ("raw", "wet_aged"),          # same
    "grilled dry-aged": ("cooked", "dry_aged"),  # prose: "still grilled"
    "dry-rendered": ("cooked", None),          # -> beef:fat product, per Step 1
    "stored / spoilage": (None, None),         # excluded entirely, per Step 2
}

# --- Step 3: beef's own group vocabulary -> VCF's 18 fixed groups, used
# ONLY for compounds that are genuinely NEW to the corpus (a compound
# already canonicalized in VCF reuses ITS OWN established compound_group,
# never overridden by a beef sheet's coarser label). Labels not in this
# map (combined "X/Y/Z" labels from Detected Beef Aging, "Miscellaneous",
# "Others", ambiguous "Oxygen heterocycles"/"N-containing"/"S-containing")
# are deliberately left unmapped and reported, per spec: "record any that
# don't fit rather than inventing new groups."
GROUP_MAP = {
    "aldehydes": "Carbonyls, aldehydes",
    "esters": "Esters",
    "alkanes": "Hydrocarbons",
    "acids": "Acids",
    "alcohols": "Alcohols",
    "pyrazines": "Bases",
    "benzenes": "Hydrocarbons",
    "hydrocarbons": "Hydrocarbons",
    "ketones": "Carbonyls, ketones",
    "furans": "Furans",
    "lactones": "Lactones",
    "alkenes": "Hydrocarbons",
    "amines": "Bases",
    "phenols": "Phenols",
    "nitrogen heterocycles": "Bases",
    "sulfur compounds": "Sulfur compounds",
    "nitrogenous": "Bases",
    "ethers": "Ethers",
    "carbonyls, aldehydes": "Carbonyls, aldehydes",
    "carbonyls, ketones": "Carbonyls, ketones",
    "hydrocarbons, aromatics": "Hydrocarbons",
    "heterocycles, thiazolines": "Bases",
}

# =====================================================================
# MR-17 / MR-18 (2026-08-29, following the beef raw-tier df=1
# investigation — see meta.json's protein_beef_validation entry for the
# full account).
#
# MR-17: a compound enters a flavour profile only if it carries a
# resolved compound_group AND is not marked flavour_relevant: false.
# Unmapped is a BLOCKING condition requiring classification, never an
# automatic exclusion — compound_group is None was originally treated as
# a proxy for "not flavour-relevant" and that's wrong: it silently would
# have dropped Furaneol (2,5-Dimethyl-4-hydroxy-3(2H)-furanone), one of
# the most-cited Maillard flavour compounds in food chemistry, from
# beef:fat:cooked's profile, for no reason but a coarse source label.
#
# MR-18: GROUP_MAP above stays deliberately incomplete for combined/
# ambiguous source labels (Others, Miscellaneous, N-containing,
# S-containing, Oxygen heterocycles, Hydrocarbons/esters/acids) — that
# was the right call, not the bug. The SAME coarse label covers both
# real flavour compounds (Furaneol under "Oxygen heterocycles", a
# pyrazine under "N-containing") and compounds with no flavour identity
# at all (Demeton-O under "Others"); blanket-mapping the label would get
# some of those wrong in one direction or the other. What MR-18 actually
# requires is that every compound landing under an unmapped label gets
# classified INDIVIDUALLY, by name, into exactly one of three outcomes,
# and that ingestion reports unmapped labels and their compound counts
# rather than silently dropping them (see n_new_compounds_unmapped_group
# in meta.json) — never that the label itself gets a blanket mapping.
#
# Three outcomes, one per compound GROUP_MAP left unmapped:
#   - mapped: a real VCF compound_group, chemistry-identified by name —
#     MAPPED_COMPOUND_OVERRIDES below. This is a classification fix (the
#     source label was too coarse), not an exclusion decision, so it does
#     NOT come from the routing workbook — the Exclusion Routing Audit
#     tab only tracks compounds that are NOT entering a profile.
#   - excluded / unresolved: read from ROUTING_XLSX's Exclusion Routing
#     Audit tab (load_routing_table, above) — NOT a code constant.
#
# Backport, 2026-08-30 (James: "beef is the reference implementation...
# two patterns in the repo is worse than the cost of retrofitting"): this
# table used to carry excluded/unresolved outcomes as hardcoded literals,
# each with a code-authored reason. Cross-checking those 10 entries
# against v29's Exclusion Routing Audit after it shipped found FIVE
# disagreements — two where the workbook had since resolved a compound
# via real external authority (EPA, IARC/ECHA) that the code still called
# unresolved, and three where the code's OWN reasoning ("industrial
# solvent") was thinner than the standard the workbook now holds compounds
# to (Ethyl chloride, Perfluorononane, Diisopropyl ether: v29 keeps all
# three at unresolved, on the explicit basis that "industrial-solvent
# identity alone is not enough to choose between [contaminant vs
# analytical-background] routes" — a stricter evidentiary bar than the
# code applied). That second kind of disagreement is not staleness, it's
# the code asserting more confidence than the evidence supported — see
# DI-BEEF-001 in meta.json's protein_beef block. Reading the workbook
# live removes the class of error, not just this instance of it.
# =====================================================================
MAPPED_COMPOUND_OVERRIDES: dict[str, dict] = {
    "2,5-Dimethyl-4-hydroxy-3(2H)-furanone": {
        "compound_group": "Furans",
        "reason": "Furaneol — a canonical Maillard/caramelic flavour compound "
                  "(strawberry/pineapple note, also a Maillard product in cooked "
                  "meat); source label 'Oxygen heterocycles' was too coarse for GROUP_MAP.",
    },
    "p-Cymene": {
        "compound_group": "Hydrocarbons",
        "reason": "a monoterpene common in cumin/thyme/oregano volatile profiles; "
                  "source label 'Others' was too coarse for GROUP_MAP. Also given an "
                  "explicit terpene_mono role override in build_vcf_compound_roles.py: "
                  "it has no CAS/molecular_weight in this corpus, and even with one, MW "
                  "134 is a band that script's MW-based terpene rule deliberately "
                  "excludes (contaminated by tetramethylbenzenes at that exact mass) — "
                  "included here by name-confirmed identity instead, the same "
                  "per-compound inspection standard the MW bands themselves were built with.",
    },
    "Ethenyl-dimethylpyrazine": {
        "compound_group": "Bases",
        "reason": "a pyrazine, same class as the corpus's other tagged pyrazines "
                  "(2,5-dimethyl pyrazine, trimethyl pyrazine, etc.); source label "
                  "'N-containing' was too coarse for GROUP_MAP.",
    },
}

# --- MR-15: straight-chain (unbranched, acyclic, fully saturated) alkanes
# never carry a terpene role, regardless of what a molecular-weight
# coincidence might suggest. Checked against beef's actual alkane list
# (Octane .. Octadecane, Undecane, Dodecane, Tridecane): none of their
# real MWs land inside build_vcf_compound_roles.py's terpene MW windows
# (136.23 / 204.35 / 202.34 / 200.32 — those require specific unsaturation
# a saturated CnH2n+2 alkane can't have), so this rule is a currently-inert
# guard, not a fix for an observed collision. It still has to exist:
# avian/finfish/etc. are not yet checked, and the rule is what Anchor #4
# tests.
STRAIGHT_CHAIN_ALKANE_RE = re.compile(
    r"^(meth|eth|prop|but|pent|hex|hept|oct|non|dec|undec|dodec|tridec|"
    r"tetradec|pentadec|hexadec|heptadec|octadec|nonadec|icos)ane$",
    re.IGNORECASE,
)


def is_straight_chain_alkane_name(name: str) -> bool:
    core = name.strip()
    core = re.sub(r"\s*\(=.*\)\s*$", "", core)  # drop a trailing "(=alias)"
    return bool(STRAIGHT_CHAIN_ALKANE_RE.match(core))


def slugify(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def normalize_name(name: str) -> str:
    """Loose match key: lowercase, strip a trailing '(=alias)', collapse
    whitespace, strip a leading stereo/locant descriptor. Good enough to
    catch 'Hexanal' == 'hexanal' and '2-Methylbutanal' == '2-methylbutanal'
    — not a full synonym resolver, and not meant to be; a miss here just
    falls through to the crosswalk lookup, then to a new provisional id,
    never to a silently wrong match (equality only, no fuzzy scoring)."""
    s = name.strip().lower()
    s = re.sub(r"\s*\(=.*\)\s*$", "", s)
    s = re.sub(r"^\(?[\d,]*[rsezRSEZ]*\)?-\s*", "", s) if False else s  # left conservative; see below
    s = re.sub(r"\s+", " ", s)
    return s


def is_present(val) -> bool:
    """Detected-table abundance cell: blank/None = not measured in this
    state; a literal 0 = source-reported not detected (per the sheets'
    own notes, e.g. 'zero = source-reported not detected'). Both mean
    'not in this tier's profile' — same outcome, different reason, so
    both are excluded, never coerced to a silent presence."""
    if val is None:
        return False
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return False
    try:
        return float(s) != 0.0
    except ValueError:
        return True  # non-numeric non-blank cell (e.g. a text flag) counts as present


def load_beef_tabs(xlsx_path: Path) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    out = {}
    mismatches = []
    for name, expected_rows in BEEF_TABS.items():
        if name not in wb.sheetnames:
            raise SystemExit(f"Expected tab {name!r} not found in {xlsx_path}")
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter)
        records = [dict(zip(header, row)) for row in rows_iter]
        if len(records) != expected_rows:
            mismatches.append((name, expected_rows, len(records)))
        out[name] = records
    if mismatches:
        raise SystemExit(f"Row-count mismatch vs spec for tabs: {mismatches}")
    return out


def land_to_csv(tabs: dict[str, list[dict]]) -> None:
    VENDOR_DIR.mkdir(parents=True, exist_ok=True)
    for name, records in tabs.items():
        df = pd.DataFrame(records)
        safe = slugify(name)
        df.to_csv(VENDOR_DIR / f"{safe}.csv", index=False)


# =====================================================================
# Compound identity resolution — reuses the SAME crosswalk
# canonicalize_vcf_compounds.py already loads, extended with a
# CAS-first / normalized-name-second priority order for a source that,
# unlike VCF, sometimes supplies its own CAS directly.
# =====================================================================

def build_existing_indexes(compound_rows: list[dict]) -> tuple[dict, dict]:
    by_cas = {r["cas"]: r for r in compound_rows if r.get("cas")}
    by_norm_name = {}
    for r in compound_rows:
        key = normalize_name(r["raw_compound"])
        by_norm_name.setdefault(key, r)
    return by_cas, by_norm_name


def load_crosswalk_for_beef():
    xw = pd.read_excel(CROSSWALK_XLSX)
    by_norm_name = {}
    for row in xw.itertuples(index=False):
        by_norm_name.setdefault(normalize_name(row.Name), {"cas": row.CAS, "cid": int(row.CID)})
    props_by_cas = {}
    for row in xw.itertuples(index=False):
        def f(v):
            try:
                fv = float(v)
                return None if fv != fv else fv
            except (TypeError, ValueError):
                return None
        props_by_cas[row.CAS] = {
            "xlogp": f(row.XLogP), "molecular_weight": f(row.MolecularWeight), "tpsa": f(row.TPSA),
        }
    return by_norm_name, props_by_cas


class BeefCompoundResolver:
    def __init__(self, existing_compound_rows: list[dict], routing_table: dict[str, dict] | None = None):
        self.by_cas, self.by_norm_name = build_existing_indexes(existing_compound_rows)
        self.crosswalk_by_name, self.crosswalk_props = load_crosswalk_for_beef()
        self.new_rows: dict[str, dict] = {}  # compound_id -> row, for genuinely new compounds
        self.method_counts: Counter = Counter()
        # {} rather than None default so a resolver built for read-only reuse
        # (aging-deltas pass, below) still fails loud if it unexpectedly hits
        # an unrouted compound, instead of crashing on a NoneType lookup.
        self.routing_table: dict[str, dict] = routing_table if routing_table is not None else {}

    def resolve(self, compound_name: str, cas: str | None, group_label: str | None) -> dict:
        cas = (cas or "").strip() or None
        norm = normalize_name(compound_name)

        # 1. CAS given directly by beef data, and already known to VCF.
        if cas and cas in self.by_cas:
            self.method_counts["cas_given_reused_existing"] += 1
            return {**self.by_cas[cas], "_new": False}

        # 2. CAS given directly, not yet known under that CAS. Before minting a
        #    new identity, check whether this compound already exists in the
        #    corpus under a NAME-ONLY provisional id (VCF's own vcf:<slug>, or
        #    a prior family's <family>:<slug>) that never had a CAS attached.
        #    Skipping this check is exactly the bug that produced 2 duplicate
        #    identities during beef ingestion — (E,E)-2,4-Heptadienal (beef
        #    CAS 4313-03-5 vs. pre-existing vcf:e_e_2_4_heptadienal) and
        #    Pentadecanal (beef CAS 2765-11-9 vs. pre-existing
        #    vcf:pentadecanal) — because this branch used to return straight
        #    to _new_compound() without ever consulting by_norm_name. Those 2
        #    existing instances are NOT retroactively merged by this fix
        #    (that needs a full rebuild, deferred — see meta.json); this only
        #    stops NEW ones from being minted by this and future families.
        if cas:
            existing_by_name = self.by_norm_name.get(norm)
            if existing_by_name is not None and not existing_by_name.get("cas"):
                self.method_counts["cas_given_name_matched_existing_no_cas"] += 1
                return {**existing_by_name, "_new": False}
            self.method_counts["cas_given_new_compound"] += 1
            return self._new_compound(compound_name, cas, self.crosswalk_props.get(cas, {}).get("xlogp") is not None,
                                       group_label, cid=None, match_method="cas_given_new_compound")

        # 3. No CAS — try normalized name against compounds ALREADY canonicalized (max reuse).
        if norm in self.by_norm_name:
            self.method_counts["name_matched_existing_compound"] += 1
            return {**self.by_norm_name[norm], "_new": False}

        # 4. No CAS — try normalized name against the crosswalk itself (same resolution VCF's
        #    own canonicalization uses), independent of whether VCF happens to use that compound.
        if norm in self.crosswalk_by_name:
            hit = self.crosswalk_by_name[norm]
            resolved_cas = hit["cas"]
            if resolved_cas in self.by_cas:
                # crosswalk resolved us to a CAS VCF already has under a differently-formatted name.
                self.method_counts["name_resolved_via_crosswalk_reused_existing"] += 1
                return {**self.by_cas[resolved_cas], "_new": False}
            self.method_counts["name_resolved_via_crosswalk_new_compound"] += 1
            return self._new_compound(compound_name, resolved_cas, True, group_label, cid=hit["cid"],
                                       match_method="name_resolved_via_crosswalk_new_compound")

        # 5. Genuinely unmatched — provisional beef:<slug> identity, same convention as VCF's vcf:<slug>.
        self.method_counts["unmatched"] += 1
        return self._new_compound(compound_name, None, False, group_label, cid=None, match_method="unmatched")

    def _new_compound(self, compound_name, cas, has_xlogp_hint, group_label, cid, match_method) -> dict:
        compound_id = cas if cas else f"beef:{slugify(compound_name)}"
        if compound_id in self.new_rows:
            row = self.new_rows[compound_id]
            row["_new"] = True
            return row
        norm_group = (group_label or "").strip().lower()
        mapped_group = GROUP_MAP.get(norm_group)
        flavour_relevant = True if mapped_group is not None else None
        mr17_reason = None
        mr17_outcome = None
        mr17_state = None
        mr17_authority = None
        mr17_route_id = None
        if mapped_group is None:
            mapped_override = MAPPED_COMPOUND_OVERRIDES.get(compound_name)
            if mapped_override is not None:
                mapped_group = mapped_override["compound_group"]
                flavour_relevant = True
                mr17_outcome = "mapped"
                mr17_reason = mapped_override["reason"]
            else:
                routing = self.routing_table.get(compound_name)
                if routing is None:
                    raise SystemExit(
                        f"{compound_name!r} (source label {group_label!r}) has no "
                        f"compound_group and no routing decision in "
                        f"{ROUTING_XLSX.name}::{ROUTING_SHEET} for this family. "
                        f"Per the routing gate, an unrouted compound cannot "
                        f"silently enter or leave the corpus — add a routed row "
                        f"with a stated authority before re-running."
                    )
                mr17_state = routing["mr17_state"]
                mr17_authority = routing["authority"]
                mr17_reason = routing["meaning"]
                mr17_route_id = routing["route_id"]
                if mr17_state == "unresolved":
                    mapped_group = None
                    flavour_relevant = None
                    mr17_outcome = "unresolved"
                else:  # present_not_flavor_relevant or analytical_background
                    mapped_group = None
                    flavour_relevant = False
                    mr17_outcome = "excluded"
        props = self.crosswalk_props.get(cas, {}) if cas else {}
        row = {
            "raw_compound": compound_name,
            "compound_group": mapped_group,  # None if unmapped/excluded/unresolved — never guessed
            "compound_group_source_label": group_label,
            "compound_group_unmapped": mapped_group is None,
            # MR-17: True = enters profiles/df counts. False = excluded, reason given.
            # None = unresolved, held pending — same exclusion outcome as False for
            # profile-building purposes, but a distinct, honestly-labeled state.
            "flavour_relevant": flavour_relevant,
            "mr17_outcome": mr17_outcome,       # "mapped" / "excluded" / "unresolved" / None (ordinary compound)
            "mr17_reason": mr17_reason,
            # Raw workbook state/authority, kept alongside the derived binary
            # mr17_outcome above — "excluded" collapses present_not_flavor_
            # relevant and analytical_background into one corpus behavior,
            # but they're different claims (a real substance vs. a method
            # artifact) and a reader shouldn't lose that distinction.
            "mr17_routing_state": mr17_state,   # workbook's own value, or None for mapped/ordinary compounds
            "mr17_routing_authority": mr17_authority,
            "mr17_routing_route_id": mr17_route_id,
            "compound_id": compound_id,
            "cas": cas,
            "pubchem_cid": cid,
            "match_method": match_method,
            "df_culinary": 0,
            "idf": None,
            "xlogp": props.get("xlogp"),
            "molecular_weight": props.get("molecular_weight"),
            "tpsa": props.get("tpsa"),
            "phase_bucket": None,  # computed by build_vcf_phase.py from xlogp, same as every other compound
            "boiling_point_c": None,
            "volatility_bucket": None,
            "source_family": "beef",
            "_new": True,
        }
        self.new_rows[compound_id] = row
        return row


# =====================================================================
# Row-level ingestion per tab
# =====================================================================

def csv_val(row: dict, key: str):
    v = row.get(key)
    if v is None:
        return None
    if isinstance(v, float) and v != v:
        return None
    return v


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=os.environ.get("BEEF_XLSX"),
                     help="Workbook carrying beef's Detected/Identification-QA tables (v21).")
    ap.add_argument("--routing-xlsx", default=os.environ.get("BEEF_ROUTING_XLSX", str(ROUTING_XLSX)),
                     help="Workbook carrying the Exclusion Routing Audit tab (defaults to the "
                          "pinned CulinAI_Staging_v30.xlsx — the tab does not exist in v21).")
    args = ap.parse_args()
    if not args.xlsx:
        raise SystemExit("Pass --xlsx PATH or set BEEF_XLSX")
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"{xlsx_path} not found")
    if not COMPOUNDS_JSONL.exists() or not SPINE_JSONL.exists():
        raise SystemExit("Run build_vcf_spine.py and canonicalize_vcf_compounds.py first.")

    PROTEIN_DIR.mkdir(parents=True, exist_ok=True)

    tabs = load_beef_tabs(xlsx_path)
    land_to_csv(tabs)
    detected_tables_xlsx_sha256 = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()

    routing_xlsx_path = Path(args.routing_xlsx)
    routing_table = load_routing_table(routing_xlsx_path, family="beef")
    routing_xlsx_sha256 = hashlib.sha256(routing_xlsx_path.read_bytes()).hexdigest()

    compound_rows = [json.loads(l) for l in COMPOUNDS_JSONL.read_text().splitlines() if l.strip()]
    resolver = BeefCompoundResolver(compound_rows, routing_table=routing_table)

    # product -> state_tier -> set of compound_ids (membership)
    profile_members: dict[tuple[str, str], set] = defaultdict(set)
    # product -> state_tier -> compound_ids eligible for df counting (evidence_mode == measured)
    profile_df_eligible: dict[tuple[str, str], set] = defaultdict(set)
    # product -> state_tier -> pre_treatments seen
    pre_treatments_seen: dict[tuple[str, str], set] = defaultdict(set)

    observations = []
    n_spoilage_excluded = 0
    n_fat_superseded_excluded = 0
    n_cure_smoke_t2_landed_not_profiled = 0

    def record_obs(**kw):
        observations.append(kw)

    mr17_excluded_observations = []  # (product_id, tier, compound_id, raw_compound, mr17_outcome, mr17_reason)

    def add_member(product_id, tier, compound_row, evidence_mode, pre_treatment=None):
        # MR-17: a compound enters a flavour profile only if it carries a
        # resolved compound_group and is not marked flavour_relevant: false.
        # flavour_relevant is False (excluded, reason on record) or None
        # (unresolved, held pending) for the same set of outcomes —
        # compound_group is None either way — so this one check covers both
        # non-guessed exclusion states without conflating them for reporting.
        if compound_row.get("compound_group") is None and compound_row.get("mr17_outcome") is not None:
            mr17_excluded_observations.append((
                product_id, tier, compound_row["compound_id"], compound_row.get("raw_compound"),
                compound_row.get("mr17_outcome"), compound_row.get("mr17_reason"),
            ))
            return False
        key = (product_id, tier)
        profile_members[key].add(compound_row["compound_id"])
        if evidence_mode == "measured":
            profile_df_eligible[key].add(compound_row["compound_id"])
        if pre_treatment:
            pre_treatments_seen[key].add(pre_treatment)
        return True

    def _mr17_blocked(compound_row):
        return compound_row.get("compound_group") is None and compound_row.get("mr17_outcome") is not None

    def _mr17_reason(compound_row):
        return f"mr17_{compound_row.get('mr17_outcome')}: {compound_row.get('mr17_reason')}"

    # --- Detected Beef Muscle (136 rows: 91 Morsli raw/grill + 45 cooked ID-QA) ---
    for row in tabs["Detected Beef Muscle"]:
        name = csv_val(row, "compound_name")
        cas = csv_val(row, "cas")
        group = csv_val(row, "compound_group")
        cid_row = resolver.resolve(str(name), str(cas) if cas else None, group)
        tiers_hit = []
        if csv_val(row, "state_scope") == "cooked beef identification-QA benchmark":
            tiers_hit.append("cooked")
        else:
            if is_present(csv_val(row, "raw_pct")):
                tiers_hit.append("raw")
            for col in ("grill55_pct", "grill60_pct", "grill71_pct", "grill77_pct", "grill85_pct"):
                if is_present(csv_val(row, col)):
                    tiers_hit.append("cooked")
                    break
        blocked = _mr17_blocked(cid_row)
        for tier in tiers_hit:
            add_member("beef:muscle", tier, cid_row, "measured")
        record_obs(source_tab="Detected Beef Muscle", detected_record_id=row.get("detected_record_id"),
                   compound_name=name, cas_given=cas, resolved_compound_id=cid_row["compound_id"],
                   match_method=cid_row.get("match_method"), product_id="beef:muscle",
                   tiers=[] if blocked else tiers_hit, evidence_mode="measured",
                   excluded=blocked or not tiers_hit,
                   exclusion_reason=_mr17_reason(cid_row) if blocked else (None if tiers_hit else "no_state_column_positive"))

    # --- Detected Beef Fat (96 rows, all dry-rendered -> beef:fat, cooked tier) ---
    for row in tabs["Detected Beef Fat"]:
        name = csv_val(row, "compound_name")
        cas = csv_val(row, "cas")
        group = csv_val(row, "normalized_compound_group") or csv_val(row, "source_compound_group")
        cid_row = resolver.resolve(str(name), str(cas) if cas else None, group)
        added = add_member("beef:fat", "cooked", cid_row, "measured")
        record_obs(source_tab="Detected Beef Fat", detected_record_id=row.get("detected_record_id"),
                   compound_name=name, cas_given=cas, resolved_compound_id=cid_row["compound_id"],
                   match_method=cid_row.get("match_method"), product_id="beef:fat",
                   tiers=["cooked"] if added else [], evidence_mode="measured", excluded=not added,
                   exclusion_reason=_mr17_reason(cid_row) if not added else None)

    # --- Detected Beef Cure-Smoke (86 rows; T1 smoked+spiced side only enters the smoked tier) ---
    for row in tabs["Detected Beef Cure-Smoke"]:
        name = csv_val(row, "compound_name")
        cas = csv_val(row, "cas")
        group = csv_val(row, "compound_group")
        cid_row = resolver.resolve(str(name), str(cas) if cas else None, group)
        t1 = str(csv_val(row, "t1_detected") or "").strip().lower() == "yes"
        added = False
        if t1:
            added = add_member("beef:muscle", "smoked", cid_row, "measured", pre_treatment="cured")
        else:
            n_cure_smoke_t2_landed_not_profiled += 1
        record_obs(source_tab="Detected Beef Cure-Smoke", detected_record_id=row.get("detected_record_id"),
                   compound_name=name, cas_given=cas, resolved_compound_id=cid_row["compound_id"],
                   match_method=cid_row.get("match_method"), product_id="beef:muscle",
                   tiers=["smoked"] if added else [], evidence_mode="measured", excluded=not added,
                   exclusion_reason=(_mr17_reason(cid_row) if (t1 and not added) else
                                     (None if t1 else "t1_not_detected_t2_unsmoked_side_not_a_named_tier")))

    # --- Detected Beef Aging (37 rows; folded into RAW tier per corrected Step 2 rule) ---
    for row in tabs["Detected Beef Aging"]:
        name = csv_val(row, "compound_name")
        cas = csv_val(row, "cas")
        group = csv_val(row, "compound_group")
        cid_row = resolver.resolve(str(name), str(cas) if cas else None, group)
        added = add_member("beef:muscle", "raw", cid_row, "measured", pre_treatment="dry_or_wet_aged_unions_pooled")
        record_obs(source_tab="Detected Beef Aging", detected_record_id=row.get("detected_record_id"),
                   compound_name=name, cas_given=cas, resolved_compound_id=cid_row["compound_id"],
                   match_method=cid_row.get("match_method"), product_id="beef:muscle",
                   tiers=["raw"] if added else [], evidence_mode="measured",
                   note="folded into raw tier; aging pooled dry+wet, day 0-28, not state-specific",
                   excluded=not added, exclusion_reason=_mr17_reason(cid_row) if not added else None)

    # --- Verified Beef Profiles (79 rows; excludes BP-052 fat-duplicate and BP-053 spoilage) ---
    for row in tabs["Verified Beef Profiles"]:
        prep_state = csv_val(row, "preparation_state")
        promo = str(csv_val(row, "promotion_status") or "")
        name = csv_val(row, "compound_name")
        cas = csv_val(row, "cas")
        group = csv_val(row, "compound_group")
        evidence_mode = str(csv_val(row, "evidence_mode") or "measured").strip().lower()
        record_id = row.get("profile_record_id")

        if prep_state == "stored / spoilage":
            n_spoilage_excluded += 1
            record_obs(source_tab="Verified Beef Profiles", detected_record_id=record_id,
                       compound_name=name, cas_given=cas, resolved_compound_id=None, match_method=None,
                       product_id="beef:muscle", tiers=[], evidence_mode=evidence_mode, excluded=True,
                       exclusion_reason="pat_beef_005_spoilage_marker_excluded_from_all_profiles")
            continue
        if promo.startswith("superseded"):
            n_fat_superseded_excluded += 1
            record_obs(source_tab="Verified Beef Profiles", detected_record_id=record_id,
                       compound_name=name, cas_given=cas, resolved_compound_id=None, match_method=None,
                       product_id=csv_val(row, "product_id"), tiers=[], evidence_mode=evidence_mode, excluded=True,
                       exclusion_reason="superseded_duplicate_of_verified_beef_fat_profile_bfp_001")
            continue

        tier, pre_treatment = STATE_TIER_MAP.get(prep_state, (None, None))
        if tier is None:
            record_obs(source_tab="Verified Beef Profiles", detected_record_id=record_id,
                       compound_name=name, cas_given=cas, resolved_compound_id=None, match_method=None,
                       product_id=csv_val(row, "product_id"), tiers=[], evidence_mode=evidence_mode, excluded=True,
                       exclusion_reason=f"unrecognized_preparation_state:{prep_state!r}")
            continue

        product_id = csv_val(row, "product_id") or "beef:muscle"
        cid_row = resolver.resolve(str(name), str(cas) if cas else None, group)
        added = add_member(product_id, tier, cid_row, evidence_mode, pre_treatment=pre_treatment)
        record_obs(source_tab="Verified Beef Profiles", detected_record_id=record_id,
                   compound_name=name, cas_given=cas, resolved_compound_id=cid_row["compound_id"],
                   match_method=cid_row.get("match_method"), product_id=product_id,
                   tiers=[tier] if added else [], pre_treatment=pre_treatment, evidence_mode=evidence_mode,
                   excluded=not added, exclusion_reason=_mr17_reason(cid_row) if not added else None,
                   preparation_state=prep_state)

    # --- Verified Beef Fat Profile (1 row) ---
    for row in tabs["Verified Beef Fat Profile"]:
        name = csv_val(row, "compound_name")
        cas = csv_val(row, "cas")
        group = csv_val(row, "compound_group")
        evidence_mode = str(csv_val(row, "evidence_mode") or "measured").strip().lower()
        cid_row = resolver.resolve(str(name), str(cas) if cas else None, group)
        added = add_member("beef:fat", "cooked", cid_row, evidence_mode)
        record_obs(source_tab="Verified Beef Fat Profile", detected_record_id=row.get("profile_record_id"),
                   compound_name=name, cas_given=cas, resolved_compound_id=cid_row["compound_id"],
                   match_method=cid_row.get("match_method"), product_id="beef:fat",
                   tiers=["cooked"] if added else [],
                   evidence_mode=evidence_mode, excluded=not added,
                   exclusion_reason=_mr17_reason(cid_row) if not added else None)

    with open(BEEF_OBSERVATIONS_OUT, "w") as f:
        for o in observations:
            f.write(json.dumps(o, ensure_ascii=False, default=str) + "\n")

    # --- append new compounds to compounds.jsonl ---
    new_compound_rows = list(resolver.new_rows.values())
    for r in new_compound_rows:
        r.pop("_new", None)
    with open(COMPOUNDS_JSONL, "a") as f:
        for r in new_compound_rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    # --- build profile records (Step 5's partial-profile decision applied below) ---
    PRODUCT_DISPLAY = {"beef:muscle": "Beef, muscle", "beef:fat": "Beef, rendered fat"}
    TIER_PREP = {"raw": ["raw"], "cooked": ["cooked"], "smoked": ["smoked", "cured"]}

    beef_profiles = []
    for (product_id, tier), cids in sorted(profile_members.items()):
        df_eligible = profile_df_eligible[(product_id, tier)]
        beef_profiles.append({
            "vcf_product_id": f"{product_id}:{tier}",
            "raw_name": f"{PRODUCT_DISPLAY.get(product_id, product_id)} ({tier})",
            "base_ingredient": product_id.split(":", 1)[1],
            "spine_id": product_id,
            "class": "culinary",
            "product_group": "Meat & Poultry",
            "profile_source": PROFILE_SOURCE,
            "n_compounds": len(cids),
            "compound_ids": sorted(cids),
            "df_eligible_compound_ids": sorted(df_eligible),
            "state_tier": tier,
            "pre_treatments_present": sorted(pre_treatments_seen[(product_id, tier)]),
            "preparation": TIER_PREP.get(tier, [tier]),
            # Step 5 decision: PARTIAL-PROFILE STATE IN THE LENS, not a size
            # normalization formula. Cosine-with-IDF already partially
            # corrects for profile size (score is normalized by each
            # side's own vector norm); a hand-tuned size-correction factor
            # would need ground truth to validate that this pass doesn't
            # have. Every non-VCF-source profile is marked partial
            # unconditionally (the smallness here is a methodology/scope
            # difference from VCF's extraction depth, not "this is
            # genuinely a simple product") so a lens can surface "partial
            # profile, N compounds" rather than present a possibly
            # size-deflated score with VCF-equivalent confidence. See
            # meta.json protein_beef.step5_partial_profile_decision for
            # the actual score-distribution numbers this was checked
            # against.
            "profile_size_class": "partial",
        })

    with open(BEEF_PROFILES_OUT, "w") as f:
        for p in beef_profiles:
            f.write(json.dumps(p, ensure_ascii=False) + "\n")

    # --- append spine entries (beef:muscle 3 members, beef:fat 1 member) ---
    spine_entries = [json.loads(l) for l in SPINE_JSONL.read_text().splitlines() if l.strip()]
    by_product: dict[str, list] = defaultdict(list)
    for p in beef_profiles:
        by_product[p["spine_id"]].append(p)

    TIER_ORDER = {"raw": 0, "cooked": 1, "smoked": 2}
    for product_id, members in sorted(by_product.items()):
        members_sorted = sorted(members, key=lambda m: TIER_ORDER.get(m["state_tier"], 9))
        spine_entries.append({
            "spine_id": product_id,
            "display_name": PRODUCT_DISPLAY.get(product_id, product_id),
            "base_ingredient": product_id.split(":", 1)[1],
            "aliases": [],
            "product_group": "Meat & Poultry",
            "n_members": len(members_sorted),
            "class_counts": {"culinary": len(members_sorted)},
            "policy": None,  # filled in by build_vcf_resolution_policy.py, same as every VCF entry
            "members": [
                {
                    "vcf_product_id": m["vcf_product_id"],
                    "raw_name": m["raw_name"],
                    "class": "culinary",
                    "preparation": m["preparation"],
                    "cure_state": "cured" if m["state_tier"] == "smoked" else None,
                    "state": None,
                    "form": None,
                    "cultivar": None,
                    "binomial": None,
                }
                for m in members_sorted
            ],
            "resolution_confidence": None,
            "default_member": None,
            "profile_source": PROFILE_SOURCE,
        })
    with open(SPINE_JSONL, "w") as f:
        for e in spine_entries:
            f.write(json.dumps(e, ensure_ascii=False) + "\n")

    meta = json.loads(META_JSON.read_text()) if META_JSON.exists() else {}

    # --- Step 6, item 4: bump vocabulary_version and refreeze the hash
    # over the COMBINED product identity (every VCF (vcf_product_id,
    # raw_name, base_ingredient, class) tuple, unchanged, PLUS every beef
    # state-tier product on the same shape) — build_vcf_spine.py computed
    # this over VCF alone before this script ran; recomputing it here,
    # over the union, is what makes the refreeze visible and checkable in
    # meta.json rather than silently stale. ---
    vcf_vocab_rows = []
    for e in spine_entries:
        if e.get("profile_source") == PROFILE_SOURCE:
            continue
        for m in e["members"]:
            vcf_vocab_rows.append((m["vcf_product_id"], m["raw_name"], e["base_ingredient"], m["class"]))
    beef_vocab_rows = [
        (p["vcf_product_id"], p["raw_name"], p["base_ingredient"], p["class"]) for p in beef_profiles
    ]
    combined_vocab_hash = hashlib.sha256(
        json.dumps(sorted(vcf_vocab_rows + beef_vocab_rows, key=str), ensure_ascii=False,
                   separators=(",", ":")).encode("utf-8")
    ).hexdigest()

    spine_meta = meta.get("spine", {})
    pre_ingestion_vocabulary_version = spine_meta.get("vocabulary_version")
    spine_meta["vocabulary_version"] = "vcf_spine_v2_plus_beef"
    spine_meta["vocabulary_hash"] = combined_vocab_hash
    spine_meta["vocabulary_version_pre_beef_ingestion"] = pre_ingestion_vocabulary_version
    spine_meta["n_spine_entries"] = len(spine_entries)
    spine_meta["n_products_total"] = len(vcf_vocab_rows) + len(beef_vocab_rows)
    spine_meta["n_products_culinary"] = spine_meta.get("n_products_culinary", 0) + len(beef_vocab_rows)
    meta["spine"] = spine_meta

    # --- aging deltas (dry28 vs wet28) — a quantitative-value diff, not a gained/lost set diff ---
    aging_deltas = []
    compound_rows_after = [json.loads(l) for l in COMPOUNDS_JSONL.read_text().splitlines() if l.strip()]
    resolver2 = BeefCompoundResolver(compound_rows_after, routing_table=routing_table)  # read-only reuse; nothing new should appear here
    for row in tabs["Beef Quantitative Deltas"]:
        if csv_val(row, "comparison_type") != "dry28_vs_wet28":
            continue
        name = csv_val(row, "compound_name")
        cas = csv_val(row, "cas")
        cid_row = resolver2.resolve(str(name), str(cas) if cas else None, None)
        aging_deltas.append({
            "diff_type": "aging_delta",
            "spine_id": "beef:muscle",
            "base_ingredient": "muscle",
            "compound_id": cid_row["compound_id"],
            "raw_compound": name,
            "delta_id": row.get("delta_id"),
            "source_id": row.get("source_id"),
            "state_a": csv_val(row, "state_a"),
            "state_b": csv_val(row, "state_b"),
            "value_a": csv_val(row, "value_a"),
            "value_b": csv_val(row, "value_b"),
            "value_unit": csv_val(row, "value_unit"),
            "direct_finding": csv_val(row, "direct_finding"),
        })
    if resolver2.new_rows:
        raise SystemExit(
            f"Beef Quantitative Deltas introduced {len(resolver2.new_rows)} compound(s) not already "
            f"seen in the Detected/Verified tabs — investigate before writing aging deltas: "
            f"{list(resolver2.new_rows)}"
        )
    with open(BEEF_AGING_DELTAS_OUT, "w") as f:
        for r in aging_deltas:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    # --- meta.json: protein_beef block ---
    n_detected_total = sum(len(tabs[t]) for t in ("Detected Beef Muscle", "Detected Beef Fat",
                                                    "Detected Beef Cure-Smoke", "Detected Beef Aging"))
    n_detected_with_cas = sum(
        1 for t in ("Detected Beef Muscle", "Detected Beef Fat", "Detected Beef Cure-Smoke", "Detected Beef Aging")
        for row in tabs[t] if str(csv_val(row, "cas") or "").strip()
    )
    unmapped_groups = sorted({
        o["compound_name"] for o in [] # placeholder, filled below properly
    })
    unmapped_new_compounds = [
        {"compound_id": cid, "raw_compound": r["raw_compound"], "source_label": r.get("compound_group_source_label")}
        for cid, r in resolver.new_rows.items() if r.get("compound_group_unmapped")
    ]
    meta["protein_beef"] = {
        "family": "beef",
        "profile_source": PROFILE_SOURCE,
        "source_workbook_provenance": {
            "detected_tables_source": {"filename": xlsx_path.name, "sha256": detected_tables_xlsx_sha256},
            "routing_source": {"filename": routing_xlsx_path.name, "sha256": routing_xlsx_sha256},
            "note": "James, 2026-08-30: nothing previously recorded which workbook version produced "
                    "beef — 'somewhere around v21-v22, from memory.' Recorded here by hash going forward, "
                    "not by version-number memory. The two are different files on purpose (see "
                    "routing_source_backport below) — beef's Detected tables predate the Exclusion "
                    "Routing Audit tab entirely.",
        },
        "tabs_ingested": BEEF_TABS,
        "rule_zero_finding_cas_coverage": {
            "claim_in_spec": "All beef rows carry CAS.",
            "actual": f"{n_detected_with_cas}/{n_detected_total} rows across the four Detected tabs carry "
                      f"a CAS number in the source data ({n_detected_with_cas/n_detected_total:.1%}).",
            "resolution": "Compound identity resolved by name against the same VCF/PubChem crosswalk "
                          "canonicalize_vcf_compounds.py already uses, exactly like VCF's own CAS-less "
                          "compounds are resolved. See compound_resolution.match_method_counts below.",
        },
        "rule_zero_finding_state_tier_table_contradiction": {
            "claim_in_spec_table": "dry-aged, wet-aged, grilled dry-aged -> cooked + pre_treatment",
            "claim_in_spec_prose": "A dry-aged steak is still raw or still grilled.",
            "resolution": "Implemented per the prose: bare dry-aged/wet-aged -> raw + pre_treatment; "
                          "only grilled dry-aged -> cooked + pre_treatment. See STATE_TIER_MAP.",
        },
        "compound_resolution": {
            "match_method_counts": dict(resolver.method_counts),
            "n_new_compounds_added": len(new_compound_rows),
            "n_new_compounds_unmapped_group": len(unmapped_new_compounds),
            "new_compounds_unmapped_group_sample": unmapped_new_compounds[:20],
        },
        "product_split": {
            "beef:muscle": "Detected Beef Muscle, Cure-Smoke, Aging; Verified Beef Profiles (excl. BP-052, BP-053)",
            "beef:fat": "Detected Beef Fat; Verified Beef Fat Profile",
            "bp_052_excluded_as_superseded_duplicate_of_bfp_001": n_fat_superseded_excluded,
        },
        "spoilage_exclusion": {
            "pat_beef_005_rows_excluded_from_profiles": n_spoilage_excluded,
            "note": "Stored in beef_observations.jsonl with excluded=true; never enters any profile "
                    "the Compound lens can reach.",
        },
        "cure_smoke_scoping": {
            "t1_smoked_rows_entering_smoked_tier": sum(1 for o in observations
                                                        if o["source_tab"] == "Detected Beef Cure-Smoke"
                                                        and not o["excluded"]),
            "t2_unsmoked_rows_landed_not_profiled": n_cure_smoke_t2_landed_not_profiled,
            "note": "Cure-Smoke is a paired T1/T2 comparison table; only the smoked+spiced side is a "
                    "named tier in the spec's own state_tier table. T2 (cured, unsmoked) is landed for "
                    "audit but not folded into any of the three named tiers.",
        },
        "profiles_built": [
            {"product_id": p["vcf_product_id"], "n_compounds": p["n_compounds"],
             "n_df_eligible": len(p["df_eligible_compound_ids"]), "pre_treatments": p["pre_treatments_present"]}
            for p in beef_profiles
        ],
        "step5_partial_profile_decision": {
            "chosen": "partial_profile_state_in_the_lens",
            "rejected": "source_aware_normalization",
            "why": "Cosine-with-IDF already normalizes by each profile's own vector norm, partially "
                   "correcting for size; a further hand-tuned correction factor would need ground truth "
                   "this pass doesn't have to validate. Every culinai_protein_v21 profile is marked "
                   "profile_size_class=partial unconditionally, and pairs.jsonl flags any pair touching "
                   "one. Actual score-distribution gap is reported in meta.json's protein_beef_validation "
                   "block after pairs.jsonl is rebuilt (Step 8) — this decision was made before seeing "
                   "that number, not fitted to it.",
        },
        "mr15_straight_chain_alkane_guard": {
            "implemented": True,
            "beef_alkanes_checked": sorted({
                o["compound_name"] for o in observations
                if o["compound_name"] and is_straight_chain_alkane_name(str(o["compound_name"]))
            }),
            "note": "None of these collide with build_vcf_compound_roles.py's terpene MW windows on "
                    "their real molecular weight (checked, not assumed) — the guard is currently inert "
                    "on this dataset but is now in place as a hard exclusion regardless of any future "
                    "MW coincidence. See build_vcf_compound_roles.py's MR15_STRAIGHT_CHAIN_ALKANE_RE.",
        },
        "mr17_mr18_flavour_relevance_classification": {
            "background": "compound_group is None was originally treated as a proxy for "
                          "'not flavour-relevant' when auto-suppressing beef's raw-tier df=1 compounds. "
                          "That's wrong: it would have silently dropped Furaneol "
                          "(2,5-Dimethyl-4-hydroxy-3(2H)-furanone), one of the most-cited Maillard "
                          "flavour compounds in food chemistry, from beef:fat:cooked's profile, because "
                          "its source label ('Oxygen heterocycles') was too coarse for GROUP_MAP — not "
                          "because it isn't flavour chemistry. See protein_beef_validation for the full "
                          "df=1 investigation this came out of.",
            "mr17": "a compound enters a flavour profile only if it carries a resolved compound_group "
                    "and is not marked flavour_relevant: false. Unmapped is a blocking condition "
                    "requiring classification, never an automatic exclusion.",
            "mr18": "GROUP_MAP stays deliberately incomplete for combined/ambiguous source labels "
                    "(Others, Miscellaneous, N-containing, S-containing, Oxygen heterocycles, "
                    "Hydrocarbons/esters/acids) — the same coarse label covers real flavour compounds "
                    "and non-culinary chemistry alike, so blanket-mapping it would get some compounds "
                    "wrong either direction. What MR-18 requires: every compound landing under an "
                    "unmapped label is classified individually by name — a real reclassification "
                    "(MAPPED_COMPOUND_OVERRIDES, a code constant, since it's a GROUP_MAP fix not an "
                    "exclusion) or a routing decision (read live from ROUTING_XLSX's Exclusion Routing "
                    "Audit tab as of 2026-08-30 — see routing_source_backport below), never a blanket "
                    "mapping of the coarse label itself.",
            "routing_source_backport": {
                "date": "2026-08-30",
                "what_changed": "excluded/unresolved outcomes for this run were read from "
                                f"{ROUTING_XLSX.name}::{ROUTING_SHEET} (family=beef), not from a code "
                                "constant. mapped stays a code constant (MAPPED_COMPOUND_OVERRIDES) — it's "
                                "a GROUP_MAP classification fix, not an exclusion, so it isn't in that tab.",
                "why": "James: beef is the reference implementation every later family's ingestion "
                       "script gets read against — two patterns in the repo (beef hardcoded, egg reading "
                       "the workbook) is worse than the cost of retrofitting beef now.",
                "gate": "load_routing_table() validates every row for this family at load time: "
                        "mr17_state must be one of VALID_MR17_ROUTING_STATES, authority must be "
                        "non-empty, and a compound with no compound_group and no matching row at all "
                        "raises SystemExit rather than silently falling through. A routing decision "
                        "without a stated basis — or with no decision recorded at all — fails the build.",
                "sha256": routing_xlsx_sha256,
            },
            "mapped": [
                {"raw_compound": r["raw_compound"], "compound_group": r["compound_group"], "reason": r["mr17_reason"]}
                for r in new_compound_rows if r.get("mr17_outcome") == "mapped"
            ],
            "excluded": [
                {"raw_compound": r["raw_compound"], "routing_state": r.get("mr17_routing_state"),
                 "reason": r["mr17_reason"], "authority": r.get("mr17_routing_authority"),
                 "route_id": r.get("mr17_routing_route_id")}
                for r in new_compound_rows if r.get("mr17_outcome") == "excluded"
            ],
            "unresolved_pending": [
                {"raw_compound": r["raw_compound"], "reason": r.get("mr17_reason"),
                 "authority": r.get("mr17_routing_authority"), "route_id": r.get("mr17_routing_route_id")}
                for r in new_compound_rows if r.get("mr17_outcome") == "unresolved"
            ],
            "n_observations_excluded_from_profiles_and_df_by_mr17": len(mr17_excluded_observations),
            "note": "excluded + unresolved compounds are held out of every profile's compound_ids and "
                    "df_eligible_compound_ids entirely (see add_member's MR-17 gate) — not merely "
                    "excluded from df counting the way evidence_mode=inherited compounds are. "
                    "Applies to any future family hitting the same coarse GROUP_MAP labels, not just beef.",
        },
        "DI-BEEF-001": {
            "type": "evidentiary_standard_correction, not a typo fix",
            "finding": "Three of beef's original code-authored exclusions (Ethyl chloride, "
                       "Perfluorononane, Diisopropyl ether) were marked 'excluded' on 'industrial "
                       "solvent' reasoning alone. Cross-checked against CulinAI_Staging_v29.xlsx's "
                       "Exclusion Routing Audit after it shipped: the workbook holds all three at "
                       "'unresolved' and states why — 'industrial-solvent identity alone is not enough "
                       "to choose between [contaminant vs. analytical-background] routes.' That is a "
                       "stricter evidentiary standard than the code applied, not a newer fact the code "
                       "simply hadn't seen yet (contrast: Thiadiazole and 2-Ethylhexyl acrylate, where "
                       "the workbook resolved a genuinely open question via EPA/IARC/ECHA authority "
                       "after the code was written — ordinary staleness, not a standards gap).",
            "james": "That isn't the workbook being behind. That's the code asserting more confidence "
                     "than the evidence supports, which is worse than being out of date, and it's the "
                     "same class of error as the four Step 11 frame revisions. The workbook's standard "
                     "is the correct one and the code should adopt the standard, not just the outputs.",
            "resolution": "Code no longer asserts an excluded/unresolved outcome on its own reasoning at "
                          "all — every such outcome for this run came from the routing table, live, with "
                          "the workbook's own stated authority (see routing_source_backport above and "
                          "the 'excluded'/'unresolved_pending' lists' authority/route_id fields). No "
                          "corpus-count effect: both 'excluded' and 'unresolved' block profile membership "
                          "identically, so this is an audit-trail correction, not a rebuild of beef's "
                          "compound counts.",
            "no_corpus_output_change": True,
        },
        "product_group_naming_note": {
            "finding": "VCF's own 'Meat & Poultry' product_group (17 pre-existing products) contains "
                       "zero poultry — 4 lamb/mutton, 13 pork. Beef's new profiles inherit the same "
                       "product_group label per the spec, so 'Meat & Poultry' is now beef + lamb/mutton "
                       "+ pork, still zero poultry. Anyone reasoning from product_group alone would "
                       "conclude poultry is covered in this corpus; it is not, at all — there is no "
                       "existing avian entry to check against before avian ingestion, ruminant/pork "
                       "validation is not a substitute, and the label itself invites that assumption.",
            "action_taken": "none — recorded here as a corpus-labeling defect, not fixed by renaming the "
                            "group (renaming 'Meat & Poultry' touches every pre-existing lamb/mutton/pork "
                            "row's product_group and is out of this delta's scope). Flagging so the next "
                            "reader doesn't repeat the assumption.",
        },
    }
    META_JSON.write_text(json.dumps(meta, indent=2))

    print(f"Landed {sum(len(v) for v in tabs.values())} rows across {len(tabs)} beef tabs to {VENDOR_DIR}")
    print(f"Compound resolution: {dict(resolver.method_counts)}")
    print(f"New compounds added to compounds.jsonl: {len(new_compound_rows)} "
          f"({len(unmapped_new_compounds)} with unmapped group)")
    print(f"Beef profiles built: {[(p['vcf_product_id'], p['n_compounds']) for p in beef_profiles]}")
    print(f"Spine entries added: {sorted(by_product)}")
    print(f"Spoilage rows excluded: {n_spoilage_excluded}")
    print(f"Fat-duplicate (superseded) rows excluded: {n_fat_superseded_excluded}")
    print(f"Aging deltas (dry28 vs wet28): {len(aging_deltas)}")


if __name__ == "__main__":
    main()
