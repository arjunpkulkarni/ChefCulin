"""
Spine near-duplicate clustering — candidate generation (James, 2026-08-29,
"Next Three Builds — Sequenced", Build 1).

Produces spine_cluster_proposals.xlsx for HUMAN REVIEW. Does NOT touch
spine.jsonl, does NOT merge any profiles, does NOT rebuild pairs.jsonl.
That is deliberate — the spec is explicit that these proposals are not
auto-applied.

Rule (as specified): propose a cluster when compound overlap >= 85% of the
smaller profile's own compound_ids, BOTH profiles have >= 20 compounds, AND
at least one of:
  - shared binomial genus (first token of the `binomial` field) — but see
    GENUS_MATCH_FALSE_FRIENDS: same genus is necessary but not sufficient
    either, the same lesson as containment below. Brassica oleracea's
    cultivar groups (cabbage, kale, kohlrabi, broccoli, cauliflower,
    brussels sprouts) share a genus AND a species without being
    culinarily substitutable — cluster_121 and cluster_83 were both
    rejected on exactly this basis (2026-08-30, James).
  - name containment after normalisation: every token of the shorter
    product's base_ingredient (or any one alias — see name_variants)
    appears, in order, among the longer product's tokens (whole-token
    subsequence, not a raw substring test — this is what keeps PINEAPPLE
    away from SWEETSOP, SUGAR APPLE, whose base_ingredient is "sugar
    apple"). Whole-token matching is necessary but not sufficient on its
    own: found 2026-08-30 (James, reviewing the 44-cluster output) that
    ARTICHOKE / JERUSALEM ARTICHOKE passed as a whole-token match despite
    being a thistle and a sunflower tuber that share nothing but a
    historical-misnomer common name — exactly the PINEAPPLE/SWEETSOP
    failure this rule exists to prevent, just one token later than the
    substring version would have let through. See
    CONTAINMENT_FALSE_FRIENDS for the fix: a small, named denylist of
    common-name tokens known to span unrelated species, in the same
    spirit as SPELLING_EQUIVALENCE below — not a general heuristic, since
    no shape reliably tells a coincidental shared common name apart from
    a real one.

Candidate pairs are read from pairs.jsonl in BOTH directions (a pair might
rank highly from A's side but not reach B's own top-20, or vice versa) —
top-20 per anchor is already generous for anything at 85%+ overlap, which
is almost always a top few match, but checking both directions costs
nothing and avoids a silent miss.

Genus extraction: the pre-parsed `binomial` field only covers 218/573 VCF
products (38%) — populated wherever the raw product name carried a
parenthetical binomial. It is NOT populated for every product a botanist
would consider genus-identifiable (e.g. no "blackberry" or "black
raspberry" row in this corpus carries a binomial at all, even though
"RASPBERRY (Rubus idaeus L.)" does for a sibling product) and it does NOT
appear on names using the bare "<Genus> SPECIES" placeholder convention
(e.g. "VACCINIUM SPECIES" itself has binomial=None even though the name
IS a genus placeholder) — that second gap is real and this script closes
it with a regex fallback; the first gap (no binomial recorded anywhere in
the corpus for a given common name) is NOT closed here, since doing so
would mean injecting outside botanical knowledge the corpus itself does
not contain, which is a bigger, separate decision than "propose spine
clusters from what's already in the data."
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

REPO_ROOT = Path(__file__).resolve().parents[2]
VCF_DIR = REPO_ROOT / "pipeline" / "artifacts" / "vcf"
SCRIPTS_DIR = Path(__file__).resolve().parent
PROFILES_JSONL = VCF_DIR / "profiles.jsonl"
PAIRS_JSONL = VCF_DIR / "pairs.jsonl"
PARSE_JSONL = VCF_DIR / "vcf_product_parse.jsonl"
OUT_XLSX = VCF_DIR / "spine_cluster_proposals.xlsx"

if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))
from build_vcf_resolution_policy import classify as resolution_policy_classify  # noqa: E402

MIN_OVERLAP = 0.85
MIN_PROFILE_SIZE = 20

SPECIES_PLACEHOLDER_RE = re.compile(r"\b([A-Za-z]+)\s+SPECIES\b", re.IGNORECASE)
TOKEN_RE = re.compile(r"[a-z0-9]+")

# James, 2026-08-29: "do it as a named spelling-equivalence table with a
# handful of entries, not a fuzzy matcher... explicit, reviewable, and it
# can't quietly start matching things you didn't intend." Each key maps to
# a single canonical token used ONLY for the containment comparison — it
# never touches display names, raw_name, or anything written to the
# proposals sheet.
SPELLING_EQUIVALENCE = {
    "whiskey": "whisky",
    "yoghurt": "yogurt",
    "aubergine": "eggplant",
}


def tokens(name: str) -> list[str]:
    return [SPELLING_EQUIVALENCE.get(t, t) for t in TOKEN_RE.findall((name or "").lower())]


def _subseq(short_tokens: list[str], long_tokens: list[str]) -> bool:
    """Every token of short_tokens appears in long_tokens, in order (whole-
    token subsequence, not a raw substring test)."""
    if not short_tokens:
        return False
    pos = 0
    for t in short_tokens:
        found = False
        while pos < len(long_tokens):
            if long_tokens[pos] == t:
                found = True
                pos += 1
                break
            pos += 1
        if not found:
            return False
    return True


def name_variants(r: dict) -> list[list[str]]:
    """All token sequences worth checking containment against for a
    product: its base_ingredient, plus each alias as its own sequence.

    Found 2026-08-30 (James, "the split"): containment previously checked
    base_ingredient only, which misses cases where the identifying word
    landed in `aliases` instead — not a taxonomy gap, a matching-logic one.
    Two concrete misses drove this:
      - "RASPBERRY, BLACKBERRY and BOYSENBERRY" has base_ingredient
        "raspberry" (2026-08-28, James: not a synonym pair, three distinct
        berries) with "blackberry and boysenberry" recorded as an alias —
        correctly parsed, but that means the literal word "blackberry"
        never reached base_tokens, so it couldn't contain-match
        "BLACKBERRY (fresh)"/"(heated)" even though the word is right
        there in the product's own aliases.
      - "SOYBEAN (Glycine max. L. merr.)" carries alias "glycine max. l.
        merr." — doesn't help THIS pair, but the same principle: aliases
        are part of a product's identity, not overflow, and matching
        logic that ignores them is incomplete on its own terms.
    Each alias is checked as its own token sequence (not merged into one
    bag) so subsequence order stays meaningful per phrase — "blackberry"
    is a subsequence of "blackberry and boysenberry" on its own; merging
    all aliases together would risk manufacturing subsequences across
    unrelated phrases that were never adjacent in any real name.
    """
    base = r.get("base_ingredient") or r["raw_name"]
    variants = [tokens(base)]
    for alias in r.get("aliases") or []:
        variants.append(tokens(alias))
    return [v for v in variants if v]


# James, 2026-08-30 (reviewing the 44-cluster proposals): cluster_237
# (ARTICHOKE / JERUSALEM ARTICHOKE) is the PINEAPPLE/SWEETSOP failure the
# whole-token subsequence rule was built to prevent, and it got through
# specifically BECAUSE it's a whole-token match rather than a raw
# substring — Cynara scolymus (a thistle) and Helianthus tuberosus (a
# sunflower tuber) share nothing but a historical-misnomer common name.
# Whole-token matching is necessary but not sufficient: a single shared
# token can still be a common-name collision between botanically unrelated
# plants rather than a real containment relationship. Named and excluded
# here rather than caught by a general heuristic (no shape reliably
# distinguishes "chinese cabbage contains cabbage-as-a-real-relative" from
# "jerusalem artichoke contains artichoke-as-a-coincidence" — that's
# exactly the outside botanical knowledge a rule can't manufacture) — same
# convention as DEMONYM_ADJECTIVES above. "chestnut" added proactively
# (James: "water chestnut against chestnut is the next one waiting" —
# water chestnut is an Eleocharis sedge tuber, unrelated to Castanea).
CONTAINMENT_FALSE_FRIENDS = {"artichoke", "chestnut"}


ABBREVIATED_GENUS_RE = re.compile(r"\b([A-Za-z])\.\s*[a-z]{3,}\b")


def has_partial_signal(tagged_genus: str, untagged_raw_name: str) -> bool:
    """James, 2026-08-30 (splitting the coverage-gap tab after triaging the
    70 asymmetric rows by hand): most of that batch shares NOTHING with
    the tagged side but a high compound-overlap score — Sweet Cherry/Beer,
    Tequila/Mulberry Spirit — pure aroma-chemistry coincidence, not a
    naming relationship a parser fix could ever resolve. A handful DO
    share something textual (Juniper Berry/Juniperus communis, Ethiopian
    Pepper "X. aethiopica"/Xylopia species) that just isn't formally
    parsed as a binomial yet. Splitting the tab on that distinction is
    what makes the actionable list short enough to actually work, per
    James — but this is a coarse, two-check proxy, not a taxonomy engine:
    (1) a token in the untagged name that's a prefix/suffix of the tagged
    genus (catches common-name/Latin stem overlap: "juniper" is a prefix
    of "juniperus"), or (2) an abbreviated-binomial letter ("X. species-
    epithet") matching the tagged genus's first letter. A "no_signal"
    call here means "nothing textual points at this," not "confirmed
    unrelated" — see the sheet's own header note.
    """
    tg = tagged_genus.lower()
    for t in tokens(untagged_raw_name):
        if len(t) >= 4 and (tg.startswith(t) or t.startswith(tg)):
            return True
    m = ABBREVIATED_GENUS_RE.search(untagged_raw_name or "")
    if m and m.group(1).lower() == tg[0]:
        return True
    return False


def containment_match(variants_a: list[list[str]], variants_b: list[list[str]]) -> bool:
    for ta in variants_a:
        for tb in variants_b:
            short, long_ = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
            if not _subseq(short, long_):
                continue
            if len(short) == 1 and short[0] in CONTAINMENT_FALSE_FRIENDS:
                # The entire match rests on one denylisted common-name
                # token — not evidence of a real relationship. Keep
                # checking other variant pairs rather than returning
                # False outright, in case a different (non-false-friend)
                # variant pair genuinely does contain-match.
                continue
            return True
    return False


def genus_of(raw_name: str, binomial: str | None) -> str | None:
    if binomial:
        first = binomial.strip().split()
        if first:
            return first[0].lower()
    m = SPECIES_PLACEHOLDER_RE.search(raw_name or "")
    if m:
        return m.group(1).lower()
    return None


def species_of(binomial: str | None) -> str | None:
    """First two words of a binomial ('genus species'), lowercased — used
    ONLY to check GENUS_MATCH_FALSE_FRIENDS below. genus_of() above stays
    genus-only, since genus is the real clustering signal; species-level
    identity only matters for this one carve-out."""
    if not binomial:
        return None
    parts = binomial.strip().split()
    if len(parts) < 2:
        return None
    return f"{parts[0].lower()} {parts[1].lower()}"


# James, 2026-08-30 (closing the door before it opens): cluster_121
# (Broccoli/Cauliflower) and cluster_83 (Brussels Sprouts/Cabbage) were
# both rejected for the identical reason — Brassica oleracea's cultivar
# groups (cabbage, kale, kohlrabi, broccoli, cauliflower, brussels
# sprouts, collard) are the same species but not culinarily
# substitutable, and the genus rule can't see that distinction; it only
# asks "same genus?", not "same genus, but are these actually
# interchangeable?" Three more pairs on the exact same door — Cabbage/
# Kohlrabi, Cabbage/Kale, Brussels Sprouts/Cabbage(cooked) — are sitting
# in the coverage-gap tab right now, one missing binomial away from
# generating their own cluster proposal the moment kale/kohlrabi/brussels
# sprouts ever get tagged. Denylisted at the SPECIES level (not the whole
# Brassica genus — Brassica campestris/B. rapa, e.g. Chinese cabbage vs
# turnip, is a separate, already-flagged conceptual gap, not this one) so
# a future binomial fix can't silently regenerate a cluster only to have
# it rejected by hand again. Same convention as CONTAINMENT_FALSE_FRIENDS
# below: a named, reviewed denylist, not a general "same genus isn't
# always substitutable" detector — no shape reliably tells that apart
# from a real generic/variant relationship (wine, whisky, fennel) without
# outside culinary knowledge a rule can't manufacture.
GENUS_MATCH_FALSE_FRIENDS = {"brassica oleracea"}


def main():
    profiles = [json.loads(l) for l in PROFILES_JSONL.read_text().splitlines() if l.strip()]
    profiles_by_id = {p["vcf_product_id"]: p for p in profiles}
    parse = {r["vcf_product_id"]: r for r in
             (json.loads(l) for l in PARSE_JSONL.read_text().splitlines() if l.strip())}
    pairs = [json.loads(l) for l in PAIRS_JSONL.read_text().splitlines() if l.strip()]

    genus = {}
    species = {}
    base_tokens = {}
    variants = {}
    for pid, r in parse.items():
        genus[pid] = genus_of(r["raw_name"], r.get("binomial"))
        species[pid] = species_of(r.get("binomial"))
        base_tokens[pid] = tokens(r.get("base_ingredient") or r["raw_name"])
        variants[pid] = name_variants(r)

    def n_compounds(pid):
        p = profiles_by_id.get(pid)
        return p["n_compounds"] if p else 0

    def compound_set(pid):
        p = profiles_by_id.get(pid)
        return set(p["compound_ids"]) if p else set()

    # binomial coverage — James asked for this number before anything else:
    # of VCF's 521 culinary products (beef's 4 have no binomial concept at
    # all, so they're outside this question), how many carry a binomial.
    culinary_vcf_ids = {pid for pid, r in parse.items() if r.get("class") == "culinary"}
    n_with_binomial = sum(1 for pid in culinary_vcf_ids if parse[pid].get("binomial"))
    binomial_coverage = {
        "n_culinary_vcf_products": len(culinary_vcf_ids),
        "n_with_binomial": n_with_binomial,
        "pct_with_binomial": round(100 * n_with_binomial / len(culinary_vcf_ids), 1),
    }

    # candidate pairs from pairs.jsonl, both directions, culinary only
    seen = set()
    candidates = []
    coverage_gap_seen = set()
    coverage_gap_rows = []
    for row in pairs:
        if row.get("suppressed_reason"):
            continue
        a, b = row["anchor_vcf_product_id"], row["match_vcf_product_id"]
        key = frozenset({a, b}) if not isinstance(a, list) else None
        if key is None or key in seen:
            continue
        na, nb = n_compounds(a), n_compounds(b)
        if na < MIN_PROFILE_SIZE or nb < MIN_PROFILE_SIZE:
            continue
        sa, sb = compound_set(a), compound_set(b)
        if not sa or not sb:
            continue
        shared = sa & sb
        # overlap relative to the SMALLER profile's own set, per spec
        smaller_id, smaller_set = (a, sa) if na <= nb else (b, sb)
        overlap = len(shared) / len(smaller_set)
        if overlap < MIN_OVERLAP:
            continue

        ga, gb = genus.get(a), genus.get(b)
        species_a, species_b = species.get(a), species.get(b)
        genus_match = (
            ga is not None and ga == gb
            and not (
                species_a is not None
                and species_a == species_b
                and species_a in GENUS_MATCH_FALSE_FRIENDS
            )
        )
        containment = containment_match(variants.get(a, []), variants.get(b, []))

        if not (genus_match or containment):
            # Met overlap+size but no genus/containment evidence either way.
            # Three cases, not two — found 2026-08-30 (James, "the split"
            # rebaseline) after the binomial-parser fix silently dropped 4
            # real pairs from BOTH this tab and the cluster list, with no
            # error and no row anywhere: Brassica campestris/Chinese
            # Cabbage, /Turnip, Juniper Berry/Juniperus communis, and Soy
            # Protein/Soybean. Each had exactly one side newly carrying a
            # binomial the other side still lacked — genus_match came back
            # False (can't match against None), which used to only mean
            # "both sides empty" but now also means "one side has a
            # confirmed genus and we simply don't know the other's" — a
            # fundamentally different, much weaker claim than "these two
            # are confirmed different taxa." Silently dropping that case is
            # the worst failure mode available: worse than a wrong cluster
            # or a coverage-gap row, because nothing points a reviewer at
            # it at all. It cost nothing THIS round only because all 4 had
            # already been resolved by hand outside this script's report.
            #   - both sides None            -> genuine coverage gap (as before)
            #   - exactly one side None       -> asymmetric evidence (NEW) —
            #     surfaced in the SAME tab with genus_call="asymmetric" so
            #     it can never again vanish without a trace
            #   - both sides non-None, differ -> a real rule decision on
            #     real evidence (both taxa confirmed and confirmed
            #     different) — correctly still excluded from this report
            if (ga is None or gb is None) and key not in coverage_gap_seen:
                coverage_gap_seen.add(key)
                is_asymmetric = (ga is None) != (gb is None)
                signal = None
                if is_asymmetric:
                    tagged_genus, untagged_pid = (ga, b) if ga is not None else (gb, a)
                    signal = (
                        "partial_signal"
                        if has_partial_signal(tagged_genus, parse[untagged_pid]["raw_name"])
                        else "no_signal"
                    )
                coverage_gap_rows.append({
                    "a": a, "b": b, "overlap": round(100 * overlap, 1),
                    "na": na, "nb": nb,
                    "genus_call": "asymmetric" if is_asymmetric else "n_a",
                    "signal": signal,  # only set for asymmetric rows
                })
            continue

        seen.add(key)
        basis = []
        if genus_match:
            basis.append(f"genus:{ga}")
        if containment:
            basis.append("containment")
        candidates.append({
            "a": a, "b": b, "overlap": overlap, "shared": len(shared),
            "na": na, "nb": nb, "basis": "+".join(basis),
        })

    # union-find into clusters (a chain of 85%+ pairs, e.g. multiple
    # VACCINIUM entries, is one proposed cluster, not N separate pairs)
    parent = {}

    def find(x):
        parent.setdefault(x, x)
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(x, y):
        rx, ry = find(x), find(y)
        if rx != ry:
            parent[rx] = ry

    for c in candidates:
        union(c["a"], c["b"])

    clusters = defaultdict(list)
    for c in candidates:
        clusters[find(c["a"])].append(c)

    # Resolution-policy preview: what build_vcf_resolution_policy.py's OWN
    # classify() would produce if a cluster's members were merged under one
    # spine_id. Reused directly from that script, not reimplemented, so this
    # preview can't drift from what actually runs later. This is a PREVIEW,
    # not a decision — James asked to set spine_id and policy in the same
    # review pass rather than discover the policy consequence later.
    cluster_policy_preview = {}
    for root, members in clusters.items():
        product_ids = sorted({pid for c in members for pid in (c["a"], c["b"])})
        fake_members = [
            {"form": parse[pid].get("form"), "cure_state": parse[pid].get("cure_state"),
             "binomial": parse[pid].get("binomial"), "preparation": parse[pid].get("preparation") or [],
             "state": parse[pid].get("state"), "cultivar": parse[pid].get("cultivar")}
            for pid in product_ids
        ]
        policy, confidence = resolution_policy_classify(fake_members)
        cluster_policy_preview[root] = (policy, confidence)

    rows_out = []
    for root, members in sorted(clusters.items(), key=lambda kv: -len(kv[1])):
        product_ids = sorted({pid for c in members for pid in (c["a"], c["b"])})
        policy, confidence = cluster_policy_preview[root]
        for pid in product_ids:
            others = [c for c in members if pid in (c["a"], c["b"])]
            rows_out.append({
                "cluster_id": f"cluster_{root}",
                "vcf_product_id": pid,
                "raw_name": parse[pid]["raw_name"],
                "current_spine_id": next((p["spine_id"] for p in profiles if p["vcf_product_id"] == pid), None),
                "n_compounds": n_compounds(pid),
                "cluster_members": ", ".join(parse[m]["raw_name"] for m in product_ids if m != pid),
                "max_overlap_pct": round(100 * max(c["overlap"] for c in others), 1),
                "match_basis": ", ".join(sorted({c["basis"] for c in others})),
                "policy_if_merged": f"{policy} ({confidence})",
                "approve": "",
                "policy_override": "",
            })

    # --- write xlsx ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cluster Proposals"
    headers = ["cluster_id", "vcf_product_id", "raw_name", "current_spine_id", "n_compounds",
               "cluster_members", "max_overlap_pct", "match_basis", "policy_if_merged",
               "approve (y/n)", "policy_override (blank = accept policy_if_merged)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows_out:
        ws.append([r["cluster_id"], r["vcf_product_id"], r["raw_name"], r["current_spine_id"],
                   r["n_compounds"], r["cluster_members"], r["max_overlap_pct"], r["match_basis"],
                   r["policy_if_merged"], "", ""])
    for col, width in zip("ABCDEFGHIJK", [14, 10, 45, 16, 12, 55, 14, 22, 20, 12, 38]):
        ws.column_dimensions[col].width = width

    # --- coverage-gap pairs, split 2026-08-30 (James, after the 70-row
    # asymmetric triage): the single "Coverage Gap - Manual Call" sheet
    # mixed a short, genuinely actionable list with a long tail of pure
    # compound-overlap coincidences (Sweet Cherry/Beer, Tequila/Mulberry
    # Spirit — 40 of the original 70 asymmetric rows had no textual
    # relationship at all), which trains a reader to skim past everything.
    # Split by has_partial_signal(): the short "needs eyes" sheet stays
    # visible and first; everything with no textual signal — every n_a
    # row (neither side has a binomial at all) plus asymmetric rows that
    # failed the signal check — moves to a second sheet, hidden by
    # default, informational only, no review expected.
    needs_review_rows = [
        r for r in coverage_gap_rows
        if r["genus_call"] == "asymmetric" and r["signal"] == "partial_signal"
    ]
    no_signal_rows = [
        r for r in coverage_gap_rows
        if r["genus_call"] == "n_a" or r["signal"] == "no_signal"
    ]

    ws2 = wb.create_sheet("Coverage Gap - Needs Review")  # Excel sheet-name limit is 31 chars
    ws2.append(["vcf_product_id_a", "raw_name_a", "vcf_product_id_b", "raw_name_b",
                "overlap_pct", "n_a", "n_b",
                "note (one side has a confirmed genus, the other doesn't, but shares a "
                "name fragment or abbreviated-genus letter with it — worth a look before "
                "assuming no relationship; a coarse proxy, not a taxonomy engine — see "
                "has_partial_signal() docstring)"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in sorted(needs_review_rows, key=lambda r: -r["overlap"]):
        ws2.append([r["a"], parse[r["a"]]["raw_name"], r["b"], parse[r["b"]]["raw_name"],
                    r["overlap"], r["na"], r["nb"], "asymmetric, partial_signal"])
    for col, width in zip("ABCDEFGH", [16, 40, 16, 40, 12, 8, 8, 40]):
        ws2.column_dimensions[col].width = width

    ws3 = wb.create_sheet("Coverage Gap - No Signal (Info)")
    ws3.append(["vcf_product_id_a", "raw_name_a", "vcf_product_id_b", "raw_name_b",
                "overlap_pct", "n_a", "n_b",
                "genus_call (n_a=neither side has a binomial; asymmetric=one side does but "
                "shares nothing textual with it — no_signal means 'nothing points at this,' "
                "not 'confirmed unrelated')"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for r in sorted(no_signal_rows, key=lambda r: -r["overlap"]):
        ws3.append([r["a"], parse[r["a"]]["raw_name"], r["b"], parse[r["b"]]["raw_name"],
                    r["overlap"], r["na"], r["nb"],
                    r["genus_call"] if r["genus_call"] == "n_a" else "asymmetric, no_signal"])
    for col, width in zip("ABCDEFGH", [16, 40, 16, 40, 12, 8, 8, 30]):
        ws3.column_dimensions[col].width = width
    ws3.sheet_state = "hidden"  # collapsed by default — informational, no review expected

    wb.save(OUT_XLSX)

    n_asymmetric = sum(1 for r in coverage_gap_rows if r["genus_call"] == "asymmetric")
    print(f"Binomial coverage: {n_with_binomial}/{len(culinary_vcf_ids)} VCF culinary products "
          f"({binomial_coverage['pct_with_binomial']}%) carry a binomial.")
    print(f"Wrote {len(rows_out)} candidate rows across {len(clusters)} proposed clusters to {OUT_XLSX}")
    for root, members in sorted(clusters.items(), key=lambda kv: -len(kv[1])):
        names = sorted({parse[pid]["raw_name"] for c in members for pid in (c["a"], c["b"])})
        policy, confidence = cluster_policy_preview[root]
        print(f"  {names}  (basis: {sorted({c['basis'] for c in members})}, policy_if_merged={policy}/{confidence})")
    print(f"Coverage-gap pairs (overlap+size gate met, no genus/containment evidence to cluster on): "
          f"{len(coverage_gap_rows)} total ({n_asymmetric} asymmetric, "
          f"{len(coverage_gap_rows) - n_asymmetric} n_a).")
    print(f"  'Coverage Gap - Needs Review': {len(needs_review_rows)} rows — one side has a "
          f"confirmed genus and shares a name fragment with the other; actually worth a look.")
    print(f"  'Coverage Gap - No Signal (Info)': {len(no_signal_rows)} rows — hidden sheet, "
          f"informational only, no review expected (pure compound-overlap coincidence or "
          f"neither side genus-confirmable at all).")

    return {
        "rows_out": rows_out, "clusters": clusters, "genus": genus, "base_tokens": base_tokens,
        "n_compounds": n_compounds, "compound_set": compound_set, "parse": parse,
        "profiles_by_id": profiles_by_id, "binomial_coverage": binomial_coverage,
        "coverage_gap_rows": coverage_gap_rows, "cluster_policy_preview": cluster_policy_preview,
    }


if __name__ == "__main__":
    main()
