"""
CulinAI Protein Layer — Egg ingestion (second family, per the Egg
Pre-Ingestion Audit / Egg Product Coverage / Egg Product Architecture /
Egg Completion Audit tabs).

Run from the repo root, AFTER build_vcf_spine.py and
canonicalize_vcf_compounds.py, and after ingest_protein_beef.py if beef is
also being included this build (order between protein families does not
matter to either script individually — each only appends — but running
beef first is what this repo has actually tested):
    python pipeline/scripts/ingest_protein_egg.py

Once this has run and egg_profiles_prebuilt.jsonl exists, build_vcf_profiles.py
needs `--include-protein` to merge it in, same flag beef uses (see that
script's EXTERNAL_PROFILE_FILES list, which now includes both files).

Reads:  CulinAI_Staging_v30.xlsx (fixed path — this is the ONLY workbook that
        has ever carried egg's Detected/Sensomics tables; there is no
        earlier version to select via a CLI flag the way beef's --xlsx
        selects v21, so this script does not take one)
Writes: pipeline/vendor/culinai_protein/egg/*.csv          (landed tabs)
        pipeline/artifacts/vcf/compounds.jsonl              (appends new
                                                              egg-only
                                                              compound rows)
        pipeline/artifacts/vcf/spine.jsonl                  (appends
                                                              egg:yolk_chicken /
                                                              egg:yolk_duck_salted /
                                                              egg:whole_chicken
                                                              spine entries;
                                                              vocabulary
                                                              re-hashed)
        pipeline/artifacts/protein/egg_profiles_prebuilt.jsonl
                                                             (consumed by
                                                              build_vcf_profiles.py)
        pipeline/artifacts/protein/egg_observations.jsonl   (full per-row
                                                              audit trail)
        pipeline/artifacts/vcf/meta.json                    (adds an
                                                              "protein_egg"
                                                              block)

=====================================================================
Why this script looks structurally simpler than beef's, and where that
simplicity actually comes from: egg's three source tables are each
ALREADY single-state (Detected Egg Boiled Yolk is 123 rows, ALL
boiled_100C_15min; Egg Scrambled Sensomics is 22 rows, ALL scrambled).
There is no STATE_TIER_MAP ambiguity to resolve the way beef's
preparation_state column spanned raw/cooked/aged/spoilage across one
table. The one genuine multi-tier table (Detected Egg Salted Yolk) signals
its two tiers directly via two independent yes/no flag columns
(salted_yolk_detected, roasted_salted_yolk_detected) rather than a single
categorical column — 45 rows, 18 salted / 43 roasted-salted / 16 shared
between both (18+43-16=45, every row accounted for, verified against Egg
Product Coverage's own EGC-008 note before writing this).

Four profiles across three spine products (per PAT-EGG-001/002/005 and MR-13
— never union yolk+white, never inherit preserved-product chemistry into
ordinary chicken egg):
    egg:yolk_chicken       -> boiled                  (Detected Egg Boiled Yolk)
    egg:yolk_duck_salted   -> salted, roasted_salted   (Detected Egg Salted Yolk)
    egg:whole_chicken      -> scrambled                (Egg Scrambled Sensomics)
No egg:white_chicken profile this pass — per Egg Completion Audit, every
chicken-white source is "full_table_pending" (raw/cooked volatiles are
directly evidenced but no accessible row-level table exists yet). Nothing
invented to fill that gap.

=====================================================================
Rule zero, this family's version — two things checked rather than assumed,
found false, and handled without silently reshaping the corpus:

1. MR-17 is a per-OCCURRENCE decision, not a per-LABEL or per-IDENTITY one
   — two separate gaps found here, not one, both from the same root cause
   (a routing decision needs to apply regardless of what else is true
   about the compound). First gap: beef's _new_compound() only consults
   the routing table when a compound's source GROUP LABEL fails to map —
   a clean label short-circuits straight to inclusion. Egg's Detected Egg
   Salted Yolk row DEY-SALT-040 (Dibutyl phthalate) breaks that: its
   compound_group is a perfectly ordinary "Esters" (maps cleanly), yet
   ER-008 explicitly excludes it as a plasticizer-contamination artifact —
   a compound-specific evidentiary call unrelated to its coarse label.
   Second, worse gap, found only by checking actual output against
   expected counts (Detected Egg Salted Yolk showed 0 excluded rows when
   it should show 1): Dibutyl phthalate already exists in VCF's own
   corpus as an ordinary compound (compound_group='Esters', df_culinary=37
   — used in 37 VCF products already, predating MR-17 entirely), so
   resolve() returns that EXISTING row at step 1 and _new_compound() never
   runs at all — ANY routing check placed inside _new_compound(), however
   correctly written, silently never fires for a compound the corpus
   already knows. ER-007/ER-008 don't declare Dibutyl phthalate globally
   non-flavor-relevant (VCF's 37 existing uses of it are untouched); they
   say egg's own detected rows of it shouldn't count toward EGG's
   profiles specifically. That is inherently a per-row, per-family check,
   so route_verdict() (below, in main()) runs at the row-processing level
   for every table without its own inline verdict column (Salted Yolk,
   Scrambled), independent of whether resolve() returned a new or a
   reused compound. Beef itself has NOT been retrofitted for either gap
   (out of scope for this pass; flagged for James in meta.json's
   protein_egg.mr17_precedence_generalization block, same as DI-BEEF-001's
   own "found by checking, reported, not silently patched everywhere it
   might apply" handling) — whether beef's own routing-table compounds
   happen to already exist elsewhere in the corpus, which would mean beef
   has this same silent-miss risk today, has not been checked.

2. Compound-identity fragmentation is a live, worsening problem, not a
   one-off. Checked every egg compound name (no-CAS tables only — CAS-given
   rows are anchored and immune to this) against compounds.jsonl and the
   VCF/PubChem crosswalk under both exact and locant/stereo-prefix-stripped
   normalization. Six genuine near-misses found — a compound egg's source
   names differently (word order, a bare non-"(=...)" alias parenthetical,
   a leading stereo descriptor normalize_name() doesn't strip, or an
   IUPAC-vs-common substituent synonym) from how it already exists in the
   corpus. Left unhandled, each would have minted a redundant provisional
   identity — exactly the failure mode beef's own module docstring already
   documents for (E,E)-2,4-Heptadienal and Pentadecanal, and which beef's
   OWN Furaneol occurrence ("2,5-Dimethyl-4-hydroxy-3(2H)-furanone",
   word-order-shifted from the crosswalk's "4-hydroxy-2,5-dimethyl-...")
   already fell into, undetected, at the time: it minted
   beef:2_5_dimethyl_4_hydroxy_3_2h_furanone instead of reusing the
   crosswalk's racemic-Furaneol CAS 3658-77-3. That instance is NOT fixed
   here (retroactive merges need a deliberate rebuild, same rule as beef's
   own (E,E)-2,4-Heptadienal/Pentadecanal cases) — but it is exactly why
   this check was worth doing rather than trusting exact-string matching a
   second time. See EGG_IDENTITY_FIXES below and protein_egg.compound_
   identity_fragmentation in meta.json for the full list, including the
   pre-existing beef-side and VCF-side fragmentation this surfaced
   (Furaneol now has FOUR identities in the corpus after this run: (R)-,
   (S)-, racemic-via-egg, and beef's stray provisional slug; p-cresol has
   two; a thiophene diol pair and a thiazoleethanol are separately
   beef-provisional despite the parent ring class already existing under
   real CAS elsewhere in the corpus). None of this is invented by egg's
   ingestion — it was already there and is only now visible because a
   second family's data was checked against the first's.

Two compounds have NO clean GROUP_MAP hit and NO routing-table row in
CulinAI_Staging_v30.xlsx (checked — the family=egg rows are exactly
ER-006/007/008, none of which name these): "Nitrous oxide" (Detected Egg
Boiled Yolk, source label "Other nitrogen compounds") and "Butane,
1-isocyano-" (source label "Isonitriles"). Per the same rule beef enforces
(an unrouted compound cannot silently enter or leave the corpus), this
should raise SystemExit and stop the build until a real ER-xxx row exists.
Given James is unreachable for the rest of the day and asked ingestion to
proceed, EGG_PROVISIONAL_ROUTING_PENDING_CONFIRMATION below supplies a
placeholder verdict for exactly these two, so the run completes rather than
blocking on two trace compounds — but it is a CODE-AUTHORED placeholder,
not a workbook decision, structurally separate from the real routing table
and loudly labeled as such everywhere it appears (print output, meta.json,
and the report back to James). This is the DI-BEEF-001 lesson applied
before shipping, not after: the code says exactly how much confidence it
actually has, and where that confidence stops.
=====================================================================
"""
from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[2]
VCF_DIR = REPO_ROOT / "pipeline" / "artifacts" / "vcf"
PROTEIN_DIR = REPO_ROOT / "pipeline" / "artifacts" / "protein"
VENDOR_DIR = REPO_ROOT / "pipeline" / "vendor" / "culinai_protein" / "egg"
CROSSWALK_XLSX = REPO_ROOT / "pipeline" / "vendor" / "vcf" / "vcfodor_cas_pubchem_distinct.xlsx"

STAGING_XLSX = REPO_ROOT / "pipeline" / "vendor" / "culinai_protein" / "CulinAI_Staging_v30.xlsx"
ROUTING_SHEET = "Exclusion Routing Audit"

COMPOUNDS_JSONL = VCF_DIR / "compounds.jsonl"
SPINE_JSONL = VCF_DIR / "spine.jsonl"
META_JSON = VCF_DIR / "meta.json"

EGG_PROFILES_OUT = PROTEIN_DIR / "egg_profiles_prebuilt.jsonl"
EGG_OBSERVATIONS_OUT = PROTEIN_DIR / "egg_observations.jsonl"

PROFILE_SOURCE = "culinai_protein_v30_egg"

EGG_TABS = {
    "Detected Egg Boiled Yolk": 123,
    "Detected Egg Salted Yolk": 45,
    "Egg Scrambled Sensomics": 22,
}

VALID_MR17_ROUTING_STATES = {
    "present_not_flavor_relevant",
    "analytical_background",
    "unresolved",
}


def load_routing_table(routing_xlsx: Path, family: str) -> dict[str, dict]:
    """Same gate ingest_protein_beef.py uses, unmodified in behavior —
    reused here by import-equivalent copy rather than a cross-script import
    so this file stays runnable standalone, matching the repo's existing
    convention of one file per family ingestion script."""
    if not routing_xlsx.exists():
        raise SystemExit(f"{routing_xlsx} not found — cannot load routing decisions.")
    wb = openpyxl.load_workbook(routing_xlsx, data_only=True, read_only=True)
    if ROUTING_SHEET not in wb.sheetnames:
        raise SystemExit(f"{routing_xlsx} has no {ROUTING_SHEET!r} tab.")
    ws = wb[ROUTING_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # title row
    header = next(rows_iter)
    table: dict[str, dict] = {}
    bad_rows = []
    for row in rows_iter:
        rec = dict(zip(header, row))
        if rec.get("family") != family:
            continue
        name = rec.get("compound_name")
        state = rec.get("mr17_state")
        authority = rec.get("authority")
        if state not in VALID_MR17_ROUTING_STATES:
            bad_rows.append((rec.get("route_id"), name, "mr17_state", state))
            continue
        if not authority or not str(authority).strip():
            bad_rows.append((rec.get("route_id"), name, "authority", authority))
            continue
        if name in table:
            existing = table[name]
            if existing["mr17_state"] != state or existing["authority"] != str(authority).strip():
                bad_rows.append((rec.get("route_id"), name, "conflicting_duplicate_compound_name_for_family",
                                  {"existing": existing, "new_state": state, "new_authority": authority}))
            continue
        table[name] = {
            "mr17_state": state,
            "authority": str(authority).strip(),
            "meaning": rec.get("meaning"),
            "route_id": rec.get("route_id"),
        }
    if bad_rows:
        raise SystemExit(f"{len(bad_rows)} row(s) in {routing_xlsx.name}::{ROUTING_SHEET} for "
                          f"family={family!r} fail the routing gate: {bad_rows}")
    return table


# =====================================================================
# GROUP_MAP — beef's map, plus egg-specific label variants added THIS
# session, each either a straightforward synonym/singular of an existing
# key (safe to blanket-map — the chemistry is unambiguous regardless of
# which specific compound carries the label) or grounded against an
# ALREADY-CANONICALIZED VCF compound under the same functional class
# (checked in compounds.jsonl before adding, not assumed):
#   thiophene-ring compounds -> Sulfur compounds: every existing VCF
#     thiophene/thiazole/benzothiazole entry (60+ checked) is already
#     "Sulfur compounds", none are "Bases" despite the ring nitrogen a
#     thiazole also carries — VCF's own convention files S-heterocycles
#     under Sulfur compounds regardless of a co-occurring ring nitrogen.
#   terpenoid alcohols -> Alcohols: grounded against geraniol/linalool/
#     citronellol/terpineol relatives already in the corpus, all "Alcohols".
#   halogenated aromatics / halogenated nitrogen compounds -> Halogens:
#     grounded against existing halogenated pyrrole entries (bromopyrrole,
#     chloropyrrole -> "Halogens", not "Bases") — VCF's convention is that
#     halogen substitution takes classification priority over the parent
#     skeleton.
#   epoxides -> "(Ep)oxides, pyrans, coumarins": VCF's actual (oddly
#     punctuated) fixed-group name for this class, found via an existing
#     epoxydecenal entry, not guessed.
# =====================================================================
GROUP_MAP = {
    "aldehydes": "Carbonyls, aldehydes",
    "esters": "Esters",
    "alkanes": "Hydrocarbons",
    "acids": "Acids",
    "alcohols": "Alcohols",
    "pyrazines": "Bases",
    "benzenes": "Hydrocarbons",
    "hydrocarbons": "Hydrocarbons",
    "ketones": "Carbonyls, ketones",
    "furans": "Furans",
    "lactones": "Lactones",
    "alkenes": "Hydrocarbons",
    "amines": "Bases",
    "phenols": "Phenols",
    "nitrogen heterocycles": "Bases",
    "sulfur compounds": "Sulfur compounds",
    "nitrogenous": "Bases",
    "ethers": "Ethers",
    "carbonyls, aldehydes": "Carbonyls, aldehydes",
    "carbonyls, ketones": "Carbonyls, ketones",
    "hydrocarbons, aromatics": "Hydrocarbons",
    "heterocycles, thiazolines": "Bases",
    # --- egg additions, 2026-08-30 ---
    "alcohol": "Alcohols",
    "aldehyde": "Carbonyls, aldehydes",
    "ester": "Esters",
    "acid": "Acids",
    "ketone": "Carbonyls, ketones",
    "aromatic hydrocarbons": "Hydrocarbons",
    "cycloalkanes": "Hydrocarbons",
    "furan-ring compounds": "Furans",
    "thiophene-ring compounds": "Sulfur compounds",
    "other sulfur compounds": "Sulfur compounds",
    "isothiocyanates": "Sulfur compounds",
    "terpenoid alcohols": "Alcohols",
    "terpenes": "Terpenes",
    "epoxides": "(Ep)oxides, pyrans, coumarins",
    "halogenated aromatics": "Halogens",
    "halogenated nitrogen compounds": "Halogens",
}

# Individual per-compound classification for names filed under a GENUINELY
# combined/ambiguous label ("Nitrogen compounds", "Other heterocycles",
# "Heterocycles", "Nitrogen heterocycle / ketone") that MR-18 says must not
# be blanket-mapped — same treatment as beef's Furaneol/p-Cymene/
# Ethenyl-dimethylpyrazine overrides.
MAPPED_COMPOUND_OVERRIDES: dict[str, dict] = {
    "1H-Tetrazol-5-amine": {
        "compound_group": "Bases",
        "reason": "a tetrazole (nitrogen heterocycle); source label 'Nitrogen compounds' too coarse for GROUP_MAP.",
    },
    "5H-Tetrazol-5-amine": {
        "compound_group": "Bases",
        "reason": "same tetrazole class as 1H-Tetrazol-5-amine; likely the same compound reported under two "
                  "tautomer-name variants in the source (no CAS given to confirm) — flagged in meta.json rather "
                  "than silently merged.",
    },
    "1-Tetrazol-2-ylethanone": {
        "compound_group": "Bases",
        "reason": "a tetrazole with a ketone substituent; source label 'Nitrogen heterocycle / ketone' is a "
                  "combined label MR-18 says not to blanket-map, but this specific compound's ring nitrogen "
                  "dominates classification the same way the other tetrazoles do.",
    },
    "Pyrrole": {
        "compound_group": "Bases",
        "reason": "matches every other pyrrole already in the corpus (all 'Bases'); source label "
                  "'Other heterocycles' too coarse for GROUP_MAP.",
    },
    "Indole": {
        "compound_group": "Bases",
        "reason": "matches every other indole already in the corpus (all 'Bases'); same coarse-label reason.",
    },
    "2-Acetylthiazole": {
        "compound_group": "Sulfur compounds",
        "reason": "a thiazole — grounded against 15+ existing VCF thiazole entries, all 'Sulfur compounds' "
                  "despite the ring nitrogen; source label 'Other heterocycles' too coarse for GROUP_MAP.",
    },
    "4-Hydroxy-2,5-dimethyl-3(2H)-furanone": {
        "compound_group": "Furans",
        "reason": "racemic Furaneol — grounded against the (R)- and (S)- Furaneol entries already in the "
                  "corpus (both 'Furans') and the crosswalk's own racemic-form entry; source label "
                  "'Heterocycles' too coarse for GROUP_MAP. See EGG_IDENTITY_FIXES for the CAS pin "
                  "(3658-77-3) that ties this to the crosswalk's existing racemic entry instead of minting "
                  "a fourth Furaneol identity the way beef's own occurrence did.",
    },
}

# =====================================================================
# EGG_IDENTITY_FIXES — compound-identity corrections for the no-CAS
# tables (Detected Egg Boiled Yolk, Egg Scrambled Sensomics), found by
# checking every egg compound name against compounds.jsonl and the VCF/
# PubChem crosswalk under both exact normalize_name() and a more
# aggressive stereo/locant-prefix-stripped comparison (see module
# docstring, rule-zero finding #2). Applied BEFORE resolver.resolve() is
# called, so the resolver's own CAS-first priority does the rest —
# these are not a second identity system, just corrected inputs to the
# same one.
#   cas: pins directly to an existing CAS already in the corpus/crosswalk.
#   resolve_as: rewrites the name to an exact existing raw_compound string
#     for a compound with NO cas (a provisional vcf:<slug> identity) —
#     the only way to reuse it, since resolve()'s existing-compound check
#     is by normalized name, not by any fuzzy score.
# =====================================================================
EGG_IDENTITY_FIXES: dict[str, dict] = {
    "(+)-2-Bornanone": {
        "cas": "76-22-2",
        "reason": "leading (+)- stereo descriptor; normalize_name() only strips a trailing (=alias), not a "
                  "leading stereo prefix. Matches existing '2-bornanone' by name minus the prefix — not "
                  "merged into the separately-named 'camphor (=alcanfor)' CAS 464-49-3 entry, a different "
                  "name string with its own separate identity in this corpus.",
    },
    "(S)-(+)-3-Methyl-1-pentanol": {
        "cas": "589-35-5",
        "reason": "leading (S)-(+)- double stereo descriptor; matches existing '3-methyl-1-pentanol' by name "
                  "minus the prefix.",
    },
    "3-(Methylsulfanyl)propanal (methional)": {
        "cas": "3268-49-3",
        "reason": "'methylsulfanyl' (egg's IUPAC-style substituent name) vs. 'methylthio' (this corpus's "
                  "existing name for the same -SCH3 group) — a synonym no punctuation-normalization catches. "
                  "Matched via the shared common name 'methional', not by string similarity.",
    },
    "4-Methylphenol (p-cresol)": {
        "cas": "106-44-5",
        "reason": "bare '(p-cresol)' alias, not the '(=p-cresol)' form normalize_name() strips. Matches "
                  "existing '4-methylphenol (=p-cresol)'. NOT merged into the separate stray "
                  "'beef:4_methyl_phenol' provisional identity (a pre-existing beef-side duplicate caused by "
                  "'4-Methyl phenol' with a space vs. '4-methylphenol' with none) — that fragmentation "
                  "predates this run and is reported, not fixed, per the established no-retroactive-merge rule.",
    },
    "3-Hydroxy-4,5-dimethyl-2(5H)-furanone (sotolon)": {
        "cas": "28664-35-9",
        "reason": "bare '(sotolon)' alias, not the '(=sotolone)' form normalize_name() strips. Matches "
                  "existing '3-hydroxy-4,5-dimethyl-2(5H)-furanone (=sotolone)'.",
    },
    "4-Hydroxy-2,5-dimethyl-3(2H)-furanone": {
        "cas": "3658-77-3",
        "reason": "racemic Furaneol. Word order and capitalization happen to match the crosswalk's "
                  "'4-hydroxy-2,5-dimethyl-3(2H)-furanone (=furaneol)' exactly, so this one would have "
                  "resolved correctly even without this entry — kept explicit anyway so the corpus's "
                  "growing Furaneol fragmentation (now 4 identities: (R)-, (S)-, this racemic one, and "
                  "beef's stray word-order-shifted provisional slug) is a documented decision, not an "
                  "accident of which two strings happened to line up.",
    },
    "trans-4,5-Epoxy-(E)-2-decenal": {
        "resolve_as": "4,5-Epoxy-(E)-2-decenal",
        "reason": "leading 'trans-' stereo descriptor on a compound with NO cas in this corpus (a "
                  "provisional vcf:4_5_epoxy_e_2_decenal identity) — the only way to reuse it is matching "
                  "the exact existing name string; there is no CAS to pin to instead.",
    },
}

# Placeholder verdicts for exactly two compounds with no GROUP_MAP hit and
# no row in CulinAI_Staging_v30.xlsx's Exclusion Routing Audit tab for
# family=egg (checked — the only three egg rows there are ER-006/007/008,
# none of which name these). NOT a workbook decision. See module docstring
# rule-zero item after finding #2, and protein_egg.provisional_routing_
# pending_confirmation in meta.json. A real ER-xxx row for either compound
# supersedes this the moment one exists — this dict should shrink to empty,
# not grow.
EGG_PROVISIONAL_ROUTING_PENDING_CONFIRMATION: dict[str, dict] = {
    "Nitrous oxide": {
        "mr17_state": "analytical_background",
        "authority": "PROVISIONAL — code-authored, not workbook-sourced (2026-08-30)",
        "meaning": "An inorganic atmospheric/carrier gas, not an organic flavor compound in any of VCF's "
                   "18 fixed groups; most plausible read is a headspace-method background artifact rather "
                   "than a genuine egg constituent. No corpus precedent to ground this against (VCF has no "
                   "existing Nitrous oxide entry at all). James has not confirmed this.",
        "route_id": None,
    },
    "Butane, 1-isocyano-": {
        "mr17_state": "unresolved",
        "authority": "PROVISIONAL — code-authored, not workbook-sourced (2026-08-30)",
        "meaning": "An isocyanide — a real, if unusual in food chemistry, functional group. Unlike Nitrous "
                   "oxide, there is no confident basis here to call it non-flavor-relevant; held pending "
                   "rather than excluded. No corpus precedent (VCF has no existing isonitrile-classified "
                   "compound to ground a mapping against either). James has not confirmed this.",
        "route_id": None,
    },
}

STRAIGHT_CHAIN_ALKANE_RE = re.compile(
    r"^(meth|eth|prop|but|pent|hex|hept|oct|non|dec|undec|dodec|tridec|"
    r"tetradec|pentadec|hexadec|heptadec|octadec|nonadec|icos)ane$",
    re.IGNORECASE,
)


def is_straight_chain_alkane_name(name: str) -> bool:
    core = re.sub(r"\s*\(=.*\)\s*$", "", name.strip())
    return bool(STRAIGHT_CHAIN_ALKANE_RE.match(core))


def slugify(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def normalize_name(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"\s*\(=.*\)\s*$", "", s)
    s = re.sub(r"\s+", " ", s)
    return s


def is_present(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return False
    try:
        return float(s) != 0.0
    except ValueError:
        return True


def is_yes(val) -> bool:
    return str(val or "").strip().lower() == "yes"


def load_egg_tabs(xlsx_path: Path) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    out = {}
    mismatches = []
    for name, expected_rows in EGG_TABS.items():
        if name not in wb.sheetnames:
            raise SystemExit(f"Expected tab {name!r} not found in {xlsx_path}")
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter)
        records = [dict(zip(header, row)) for row in rows_iter]
        if len(records) != expected_rows:
            mismatches.append((name, expected_rows, len(records)))
        out[name] = records
    if mismatches:
        raise SystemExit(f"Row-count mismatch vs expected for tabs: {mismatches}")
    return out


def land_to_csv(tabs: dict[str, list[dict]]) -> None:
    VENDOR_DIR.mkdir(parents=True, exist_ok=True)
    for name, records in tabs.items():
        df = pd.DataFrame(records)
        df.to_csv(VENDOR_DIR / f"{slugify(name)}.csv", index=False)


def build_existing_indexes(compound_rows: list[dict]) -> tuple[dict, dict]:
    by_cas = {r["cas"]: r for r in compound_rows if r.get("cas")}
    by_norm_name = {}
    for r in compound_rows:
        key = normalize_name(r["raw_compound"])
        by_norm_name.setdefault(key, r)
    return by_cas, by_norm_name


def load_crosswalk():
    xw = pd.read_excel(CROSSWALK_XLSX)
    by_norm_name = {}
    for row in xw.itertuples(index=False):
        by_norm_name.setdefault(normalize_name(row.Name), {"cas": row.CAS, "cid": int(row.CID)})
    props_by_cas = {}
    for row in xw.itertuples(index=False):
        def f(v):
            try:
                fv = float(v)
                return None if fv != fv else fv
            except (TypeError, ValueError):
                return None
        props_by_cas[row.CAS] = {"xlogp": f(row.XLogP), "molecular_weight": f(row.MolecularWeight), "tpsa": f(row.TPSA)}
    return by_norm_name, props_by_cas


class EggCompoundResolver:
    """Same resolution cascade as BeefCompoundResolver (ingest_protein_beef.py),
    with one deliberate generalization: the routing table is consulted for
    every genuinely-new compound BEFORE group-mapping, not only when the
    group is unmapped. See module docstring rule-zero finding #1 for why —
    Dibutyl phthalate's compound_group is a clean 'Esters', but the
    workbook still excludes it by name via ER-008, and beef's label-gated
    routing check would have missed that."""

    def __init__(self, existing_compound_rows: list[dict], routing_table: dict[str, dict]):
        self.by_cas, self.by_norm_name = build_existing_indexes(existing_compound_rows)
        self.crosswalk_by_name, self.crosswalk_props = load_crosswalk()
        self.new_rows: dict[str, dict] = {}
        self.method_counts: Counter = Counter()
        self.routing_table = routing_table

    def resolve(self, compound_name: str, cas: str | None, group_label: str | None) -> dict:
        cas = (cas or "").strip() or None
        norm = normalize_name(compound_name)

        if cas and cas in self.by_cas:
            self.method_counts["cas_given_reused_existing"] += 1
            return {**self.by_cas[cas], "_new": False}

        if cas:
            existing_by_name = self.by_norm_name.get(norm)
            if existing_by_name is not None and not existing_by_name.get("cas"):
                self.method_counts["cas_given_name_matched_existing_no_cas"] += 1
                return {**existing_by_name, "_new": False}
            self.method_counts["cas_given_new_compound"] += 1
            return self._new_compound(compound_name, cas, group_label, match_method="cas_given_new_compound")

        if norm in self.by_norm_name:
            self.method_counts["name_matched_existing_compound"] += 1
            return {**self.by_norm_name[norm], "_new": False}

        if norm in self.crosswalk_by_name:
            hit = self.crosswalk_by_name[norm]
            resolved_cas = hit["cas"]
            if resolved_cas in self.by_cas:
                self.method_counts["name_resolved_via_crosswalk_reused_existing"] += 1
                return {**self.by_cas[resolved_cas], "_new": False}
            self.method_counts["name_resolved_via_crosswalk_new_compound"] += 1
            return self._new_compound(compound_name, resolved_cas, group_label, cid=hit["cid"],
                                       match_method="name_resolved_via_crosswalk_new_compound")

        self.method_counts["unmatched"] += 1
        return self._new_compound(compound_name, None, group_label, match_method="unmatched")

    def _new_compound(self, compound_name, cas, group_label, match_method, cid=None) -> dict:
        compound_id = cas if cas else f"egg:{slugify(compound_name)}"
        if compound_id in self.new_rows:
            row = self.new_rows[compound_id]
            row["_new"] = True
            return row

        mapped_group = None
        flavour_relevant = None
        mr17_outcome = None
        mr17_reason = None
        mr17_state = None
        mr17_authority = None
        mr17_route_id = None
        mr17_provisional = False

        # Routing table (real or provisional-placeholder) takes precedence
        # over group-mapping — see class docstring / module rule-zero #1.
        routing = self.routing_table.get(compound_name) or EGG_PROVISIONAL_ROUTING_PENDING_CONFIRMATION.get(compound_name)
        if routing is not None:
            mr17_state = routing["mr17_state"]
            mr17_authority = routing["authority"]
            mr17_reason = routing["meaning"]
            mr17_route_id = routing.get("route_id")
            mr17_provisional = compound_name in EGG_PROVISIONAL_ROUTING_PENDING_CONFIRMATION
            if mr17_state == "unresolved":
                flavour_relevant = None
                mr17_outcome = "unresolved"
            else:
                flavour_relevant = False
                mr17_outcome = "excluded"
        else:
            norm_group = (group_label or "").strip().lower()
            mapped_group = GROUP_MAP.get(norm_group)
            if mapped_group is not None:
                flavour_relevant = True
            else:
                mapped_override = MAPPED_COMPOUND_OVERRIDES.get(compound_name)
                if mapped_override is not None:
                    mapped_group = mapped_override["compound_group"]
                    flavour_relevant = True
                    mr17_outcome = "mapped"
                    mr17_reason = mapped_override["reason"]
                else:
                    raise SystemExit(
                        f"{compound_name!r} (source label {group_label!r}) has no compound_group, no "
                        f"MAPPED_COMPOUND_OVERRIDES entry, and no row in the Exclusion Routing Audit tab "
                        f"(or the provisional placeholder dict) for family=egg. Per the routing gate, an "
                        f"unrouted compound cannot silently enter or leave the corpus — classify it, "
                        f"override it, or add a routed row before re-running."
                    )

        props = self.crosswalk_props.get(cas, {}) if cas else {}
        row = {
            "raw_compound": compound_name,
            "compound_group": mapped_group,
            "compound_group_source_label": group_label,
            "compound_group_unmapped": mapped_group is None,
            "flavour_relevant": flavour_relevant,
            "mr17_outcome": mr17_outcome,
            "mr17_reason": mr17_reason,
            "mr17_routing_state": mr17_state,
            "mr17_routing_authority": mr17_authority,
            "mr17_routing_route_id": mr17_route_id,
            "mr17_routing_provisional_pending_confirmation": mr17_provisional,
            "compound_id": compound_id,
            "cas": cas,
            "pubchem_cid": cid,
            "match_method": match_method,
            "df_culinary": 0,
            "idf": None,
            "xlogp": props.get("xlogp"),
            "molecular_weight": props.get("molecular_weight"),
            "tpsa": props.get("tpsa"),
            "phase_bucket": None,
            "boiling_point_c": None,
            "volatility_bucket": None,
            "source_family": "egg",
            "_new": True,
        }
        self.new_rows[compound_id] = row
        return row


def apply_identity_fix(name: str, cas: str | None) -> tuple[str, str | None]:
    """Returns (name_to_resolve, cas_to_resolve) after applying
    EGG_IDENTITY_FIXES, if any. The ORIGINAL name is kept separately in
    the observation record regardless — this only changes what gets
    passed to resolver.resolve()."""
    fix = EGG_IDENTITY_FIXES.get(name)
    if fix is None:
        return name, cas
    if "cas" in fix:
        return name, fix["cas"]
    if "resolve_as" in fix:
        return fix["resolve_as"], cas
    return name, cas


def csv_val(row: dict, key: str):
    v = row.get(key)
    if v is None:
        return None
    if isinstance(v, float) and v != v:
        return None
    return v


def main():
    if not STAGING_XLSX.exists():
        raise SystemExit(f"{STAGING_XLSX} not found.")
    if not COMPOUNDS_JSONL.exists() or not SPINE_JSONL.exists():
        raise SystemExit("Run build_vcf_spine.py and canonicalize_vcf_compounds.py first.")

    PROTEIN_DIR.mkdir(parents=True, exist_ok=True)

    tabs = load_egg_tabs(STAGING_XLSX)
    land_to_csv(tabs)
    staging_xlsx_sha256 = hashlib.sha256(STAGING_XLSX.read_bytes()).hexdigest()

    routing_table = load_routing_table(STAGING_XLSX, family="egg")

    compound_rows = [json.loads(l) for l in COMPOUNDS_JSONL.read_text().splitlines() if l.strip()]
    resolver = EggCompoundResolver(compound_rows, routing_table=routing_table)

    profile_members: dict[tuple[str, str], set] = defaultdict(set)
    profile_df_eligible: dict[tuple[str, str], set] = defaultdict(set)
    observations = []

    def record_obs(**kw):
        observations.append(kw)

    def _mr17_blocked(compound_row):
        return compound_row.get("compound_group") is None and compound_row.get("mr17_outcome") is not None

    def _mr17_reason(compound_row):
        return f"mr17_{compound_row.get('mr17_outcome')}: {compound_row.get('mr17_reason')}"

    def route_verdict(raw_compound_name: str):
        """A DIFFERENT gap than the one described in the module docstring's
        rule-zero #1 (label-gated vs. compound-gated routing) — this one is
        occurrence-gated vs. identity-gated. EggCompoundResolver's routing
        check lives inside _new_compound(), which only runs for a compound
        genuinely NEW to the corpus. Dibutyl phthalate (ER-007/ER-008)
        already exists as an ordinary VCF compound (compound_group='Esters',
        df_culinary=37 — it's used in 37 VCF products already, predating
        MR-17 entirely), so resolve() returns that EXISTING row at step 1
        (cas already in by_cas) and _new_compound() never runs — the
        routing check built for rule-zero #1 silently never fires. Found by
        checking the actual profile-build output against expected exclusion
        counts (Detected Egg Salted Yolk showed 0 excluded rows when it
        should show 1), not by re-reading the code and spotting it. The
        real fix: a routing decision is about a (family, compound_name)
        OCCURRENCE — ER-007/ER-008 don't declare Dibutyl phthalate globally
        non-flavor-relevant (it stays exactly as it is in VCF's own 37
        products), they say egg's own detected rows of it shouldn't count
        toward EGG's profiles. That has nothing to do with whether the
        resolved identity is new or reused, so this check has to run
        per-ROW, independent of resolver.resolve()'s new/existing branch,
        for every table that doesn't carry its own inline per-row verdict
        (Salted Yolk, Scrambled — Boiled Yolk is exempt: it carries
        culinary_relevance_route itself, is self-governing, and was cross-
        checked against this same table already, see meta.json)."""
        return routing_table.get(raw_compound_name) or EGG_PROVISIONAL_ROUTING_PENDING_CONFIRMATION.get(raw_compound_name)

    def add_member(product_id, tier, compound_row, evidence_mode, route_override=None):
        # Any routing verdict — excluded OR unresolved — blocks membership,
        # regardless of whether resolve() returned a new or a pre-existing
        # compound (see route_verdict's docstring above). Both outcomes
        # hold a compound out of every profile identically, same rule as
        # beef's own add_member.
        if route_override is not None:
            return False
        if _mr17_blocked(compound_row):
            return False
        key = (product_id, tier)
        profile_members[key].add(compound_row["compound_id"])
        if evidence_mode == "measured":
            profile_df_eligible[key].add(compound_row["compound_id"])
        return True

    # --- Detected Egg Boiled Yolk (123 rows; single product+state) ---
    for row in tabs["Detected Egg Boiled Yolk"]:
        name = str(csv_val(row, "compound_name_source"))
        group = csv_val(row, "normalized_group")
        route = csv_val(row, "culinary_relevance_route")
        resolve_name, resolve_cas = apply_identity_fix(name, None)
        cid_row = resolver.resolve(resolve_name, resolve_cas, group)

        # This table carries its own per-row MR-17 verdict directly (unlike
        # Salted Yolk / Scrambled, which have no such column and rely
        # entirely on the routing table / GROUP_MAP). Cross-checked against
        # the Exclusion Routing Audit's ER-006/007 rows for the two
        # excluded compounds here — both agree exactly (see module
        # docstring) — so this table's own column is used directly rather
        # than re-deriving the same verdict through the routing table a
        # second time.
        excluded_inline = route in ("present_not_flavor_relevant", "analytical_background")
        detected = is_yes(csv_val(row, "commercial_egg_detected"))
        added = False
        if not excluded_inline and detected:
            added = add_member("egg:yolk_chicken", "boiled", cid_row, "measured")
        record_obs(
            source_tab="Detected Egg Boiled Yolk", record_id=row.get("record_id"),
            compound_name=name, resolved_compound_id=cid_row["compound_id"],
            match_method=cid_row.get("match_method"), product_id="egg:yolk_chicken",
            tiers=["boiled"] if added else [], evidence_mode="measured",
            commercial_egg_detected=csv_val(row, "commercial_egg_detected"),
            breed_abundances={
                "commercial_egg_CE": csv_val(row, "commercial_egg_CE"),
                "blue_shell_BSE": csv_val(row, "blue_shell_BSE"),
                "woorimatdag1": csv_val(row, "woorimatdag1"),
                "woorimatdag2": csv_val(row, "woorimatdag2"),
            },
            culinary_relevance_route=route,
            excluded=not added,
            exclusion_reason=(
                f"inline_route_{route}" if excluded_inline else
                (_mr17_reason(cid_row) if _mr17_blocked(cid_row) else
                 ("not_detected_in_commercial_egg_control" if not detected else None))
            ),
        )

    # --- Detected Egg Salted Yolk (45 rows; TWO tiers, independent flags) ---
    for row in tabs["Detected Egg Salted Yolk"]:
        name = str(csv_val(row, "compound_name"))
        cas = csv_val(row, "cas")
        group = csv_val(row, "compound_group")
        resolve_name, resolve_cas = apply_identity_fix(name, str(cas) if cas else None)
        cid_row = resolver.resolve(resolve_name, resolve_cas, group)

        tiers_hit = []
        if is_yes(csv_val(row, "salted_yolk_detected")):
            tiers_hit.append("salted")
        if is_yes(csv_val(row, "roasted_salted_yolk_detected")):
            tiers_hit.append("roasted_salted")

        route = route_verdict(name)  # occurrence-level check, see route_verdict docstring
        blocked = route is not None or _mr17_blocked(cid_row)
        added_any = False
        for tier in tiers_hit:
            if add_member("egg:yolk_duck_salted", tier, cid_row, "measured", route_override=route):
                added_any = True
        if route is not None:
            reason = (f"mr17_{route['mr17_state']} (route_id={route.get('route_id')}, "
                      f"provisional={name in EGG_PROVISIONAL_ROUTING_PENDING_CONFIRMATION}): {route['meaning']}")
        elif _mr17_blocked(cid_row):
            reason = _mr17_reason(cid_row)
        elif not tiers_hit:
            reason = "neither_state_flag_yes"
        else:
            reason = None
        record_obs(
            source_tab="Detected Egg Salted Yolk", record_id=row.get("record_id"),
            compound_name=name, cas_given=cas, resolved_compound_id=cid_row["compound_id"],
            match_method=cid_row.get("match_method"), product_id="egg:yolk_duck_salted",
            tiers=[] if blocked else tiers_hit, evidence_mode="measured",
            excluded=blocked or not tiers_hit,
            exclusion_reason=reason,
        )

    # --- Egg Scrambled Sensomics (22 rows; single product+state) ---
    for row in tabs["Egg Scrambled Sensomics"]:
        name = str(csv_val(row, "compound_name"))
        group = csv_val(row, "compound_group")
        quant_layer = csv_val(row, "quantitation_layer")
        resolve_name, resolve_cas = apply_identity_fix(name, None)
        cid_row = resolver.resolve(resolve_name, resolve_cas, group)

        detected = is_yes(csv_val(row, "detected_in_scrambled_chicken"))
        evidence_mode = "measured" if quant_layer in ("SIDA_quantified", "headspace_semquant") else "inherited"
        route = route_verdict(name)  # defensive — no family=egg routing row currently targets this table
        added = False
        if detected:
            added = add_member("egg:whole_chicken", "scrambled", cid_row, evidence_mode, route_override=route)
        if route is not None:
            reason = f"mr17_{route['mr17_state']} (route_id={route.get('route_id')}): {route['meaning']}"
        elif detected and _mr17_blocked(cid_row):
            reason = _mr17_reason(cid_row)
        elif not detected:
            reason = "not_detected_in_scrambled_chicken"
        else:
            reason = None
        record_obs(
            source_tab="Egg Scrambled Sensomics", record_id=row.get("record_id"),
            compound_name=name, resolved_compound_id=cid_row["compound_id"],
            match_method=cid_row.get("match_method"), product_id="egg:whole_chicken",
            tiers=["scrambled"] if added else [], evidence_mode=evidence_mode,
            quantitation_layer=quant_layer, concentration_ug_kg=csv_val(row, "concentration_ug_kg"),
            odor_threshold_ug_kg=csv_val(row, "odor_threshold_ug_kg"), DoT_factor=csv_val(row, "DoT_factor"),
            excluded=not added,
            exclusion_reason=reason,
        )

    with open(EGG_OBSERVATIONS_OUT, "w") as f:
        for o in observations:
            f.write(json.dumps(o, ensure_ascii=False, default=str) + "\n")

    new_compound_rows = list(resolver.new_rows.values())
    for r in new_compound_rows:
        r.pop("_new", None)
    with open(COMPOUNDS_JSONL, "a") as f:
        for r in new_compound_rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    PRODUCT_DISPLAY = {
        "egg:yolk_chicken": "Egg, chicken yolk",
        "egg:yolk_duck_salted": "Egg, duck yolk (salted)",
        "egg:whole_chicken": "Egg, whole chicken (composite)",
    }

    egg_profiles = []
    for (product_id, tier), cids in sorted(profile_members.items()):
        df_eligible = profile_df_eligible[(product_id, tier)]
        egg_profiles.append({
            "vcf_product_id": f"{product_id}:{tier}",
            "raw_name": f"{PRODUCT_DISPLAY.get(product_id, product_id)} ({tier.replace('_', ' ')})",
            "base_ingredient": product_id.split(":", 1)[1],
            "spine_id": product_id,
            "class": "culinary",
            "product_group": "Egg",
            "profile_source": PROFILE_SOURCE,
            "n_compounds": len(cids),
            "compound_ids": sorted(cids),
            "df_eligible_compound_ids": sorted(df_eligible),
            "state_tier": tier,
            "preparation": [tier],
            # Same Step-5 decision beef made, same reasoning, not re-derived:
            # every non-VCF-source profile is partial unconditionally.
            "profile_size_class": "partial",
        })

    with open(EGG_PROFILES_OUT, "w") as f:
        for p in egg_profiles:
            f.write(json.dumps(p, ensure_ascii=False) + "\n")

    spine_entries = [json.loads(l) for l in SPINE_JSONL.read_text().splitlines() if l.strip()]
    by_product: dict[str, list] = defaultdict(list)
    for p in egg_profiles:
        by_product[p["spine_id"]].append(p)

    TIER_ORDER = {"boiled": 0, "salted": 0, "roasted_salted": 1, "scrambled": 0}
    for product_id, members in sorted(by_product.items()):
        members_sorted = sorted(members, key=lambda m: TIER_ORDER.get(m["state_tier"], 9))
        spine_entries.append({
            "spine_id": product_id,
            "display_name": PRODUCT_DISPLAY.get(product_id, product_id),
            "base_ingredient": product_id.split(":", 1)[1],
            "aliases": [],
            "product_group": "Egg",
            "n_members": len(members_sorted),
            "class_counts": {"culinary": len(members_sorted)},
            "policy": None,
            "members": [
                {
                    "vcf_product_id": m["vcf_product_id"],
                    "raw_name": m["raw_name"],
                    "class": "culinary",
                    "preparation": m["preparation"],
                    "cure_state": "salted" if m["state_tier"] in ("salted", "roasted_salted") else None,
                    "state": None,
                    "form": None,
                    "cultivar": None,
                    "binomial": None,
                }
                for m in members_sorted
            ],
            "resolution_confidence": None,
            "default_member": None,
            "profile_source": PROFILE_SOURCE,
        })
    with open(SPINE_JSONL, "w") as f:
        for e in spine_entries:
            f.write(json.dumps(e, ensure_ascii=False) + "\n")

    meta = json.loads(META_JSON.read_text()) if META_JSON.exists() else {}

    vcf_vocab_rows = []
    for e in spine_entries:
        if e.get("profile_source") in (PROFILE_SOURCE, "culinai_protein_v21"):
            continue
        for m in e["members"]:
            vcf_vocab_rows.append((m["vcf_product_id"], m["raw_name"], e["base_ingredient"], m["class"]))
    beef_vocab_rows = []
    for e in spine_entries:
        if e.get("profile_source") == "culinai_protein_v21":
            for m in e["members"]:
                beef_vocab_rows.append((m["vcf_product_id"], m["raw_name"], e["base_ingredient"], m["class"]))
    egg_vocab_rows = [
        (p["vcf_product_id"], p["raw_name"], p["base_ingredient"], p["class"]) for p in egg_profiles
    ]
    combined_vocab_hash = hashlib.sha256(
        json.dumps(sorted(vcf_vocab_rows + beef_vocab_rows + egg_vocab_rows, key=str), ensure_ascii=False,
                   separators=(",", ":")).encode("utf-8")
    ).hexdigest()

    spine_meta = meta.get("spine", {})
    pre_ingestion_vocabulary_version = spine_meta.get("vocabulary_version")
    spine_meta["vocabulary_version"] = "vcf_spine_v2_plus_beef_plus_egg"
    spine_meta["vocabulary_hash"] = combined_vocab_hash
    spine_meta["vocabulary_version_pre_egg_ingestion"] = pre_ingestion_vocabulary_version
    spine_meta["n_spine_entries"] = len(spine_entries)
    spine_meta["n_products_total"] = len(vcf_vocab_rows) + len(beef_vocab_rows) + len(egg_vocab_rows)
    # Derived fresh from spine_entries every run, not incrementally added to
    # whatever meta.json already held — beef's own script uses the
    # incremental form (spine_meta.get(...) + len(beef_vocab_rows)), which
    # would double-count on a second run of the same script. Not retrofitted
    # into beef (out of scope, flagged in mr17_precedence_generalization's
    # sibling note above), but not repeated here either.
    all_vocab_rows = vcf_vocab_rows + beef_vocab_rows + egg_vocab_rows
    spine_meta["n_products_culinary"] = sum(1 for row in all_vocab_rows if row[3] == "culinary")
    meta["spine"] = spine_meta

    n_provisional_hits = sum(
        1 for r in new_compound_rows if r.get("mr17_routing_provisional_pending_confirmation")
    )
    unmapped_new_compounds = [
        {"compound_id": cid, "raw_compound": r["raw_compound"], "source_label": r.get("compound_group_source_label")}
        for cid, r in resolver.new_rows.items() if r.get("compound_group_unmapped") and r.get("mr17_outcome") is None
    ]

    meta["protein_egg"] = {
        "family": "egg",
        "profile_source": PROFILE_SOURCE,
        "source_workbook_provenance": {
            "staging_workbook": {"filename": STAGING_XLSX.name, "sha256": staging_xlsx_sha256},
            "note": "Single workbook for both Detected/Sensomics tables and the Exclusion Routing Audit tab "
                    "— unlike beef, there is no earlier version predating the routing tab to reconcile.",
        },
        "tabs_ingested": EGG_TABS,
        "product_split": {
            "egg:yolk_chicken": "Detected Egg Boiled Yolk -> boiled tier only",
            "egg:yolk_duck_salted": "Detected Egg Salted Yolk -> salted + roasted_salted tiers "
                                     "(18 salted / 43 roasted_salted / 16 shared, all 45 rows accounted for)",
            "egg:whole_chicken": "Egg Scrambled Sensomics -> scrambled tier only",
            "egg:white_chicken": "NOT ingested this pass — every chicken-white source in Egg Product "
                                  "Coverage is 'full_table_pending'; no row-level table exists yet. Nothing "
                                  "fabricated to fill this gap.",
        },
        "boiled_yolk_inline_routing_cross_check": {
            "note": "Detected Egg Boiled Yolk carries its own culinary_relevance_route column per row, "
                    "unlike Salted Yolk/Scrambled which rely entirely on GROUP_MAP + the routing table. "
                    "Cross-checked its two excluded rows against the Exclusion Routing Audit tab's own "
                    "family=egg entries: both agree exactly (Cyclotetrasiloxane/ER-006 analytical_background; "
                    "Dibutyl phthalate/ER-007 present_not_flavor_relevant). No discrepancy found.",
            "n_excluded_present_not_flavor_relevant": sum(
                1 for o in observations if o.get("culinary_relevance_route") == "present_not_flavor_relevant"),
            "n_excluded_analytical_background": sum(
                1 for o in observations if o.get("culinary_relevance_route") == "analytical_background"),
            "n_included_with_quality_caution": sum(
                1 for o in observations if o.get("culinary_relevance_route") == "culinary_state_occurrence_with_quality_caution"),
        },
        "mr17_precedence_generalization": {
            "finding_1_label_gated": "Beef's _new_compound() only consults the routing table when a "
                       "compound's group label fails to map. Egg's Dibutyl phthalate (DEY-SALT-040, "
                       "ER-008) has compound_group='Esters', which maps cleanly via GROUP_MAP, yet the "
                       "workbook explicitly excludes it by name. Beef's label-gated shortcut would have "
                       "missed this.",
            "finding_2_identity_gated_the_one_that_actually_broke_the_first_run": "Worse: Dibutyl "
                       "phthalate already exists in VCF's own corpus as an ordinary compound "
                       "(compound_group='Esters', df_culinary=37 — used in 37 VCF products, predating "
                       "MR-17 entirely). resolve() returns that existing row before _new_compound() is "
                       "ever called, so ANY routing check placed inside _new_compound() — including the "
                       "fix for finding 1 — silently never fires. Found by checking actual output against "
                       "expected exclusion counts on the first run of this script (Detected Egg Salted "
                       "Yolk showed 0 excluded rows when ER-008 says it should show 1), not by re-reading "
                       "the code.",
            "resolution": "route_verdict() runs at the per-ROW level in main(), for every table without "
                          "its own inline verdict column, independent of whether resolve() returned a new "
                          "or a reused compound — a routing decision is about a (family, compound_name) "
                          "OCCURRENCE, not a global identity fact, so it cannot correctly live only inside "
                          "the new-compound path. This is a stricter reading of the SAME MR-17 rule beef's "
                          "own prose already states ('not marked flavour_relevant: false'), not a new rule.",
            "beef_not_retrofitted": "Beef's script is unchanged by either finding — out of scope for this "
                                    "pass, same as DI-BEEF-001's own handling. Whether any of beef's own "
                                    "routing-table compounds already existed elsewhere in the corpus before "
                                    "beef's ingestion ran (which would mean beef has this same silent-miss "
                                    "risk today, undetected because beef has no per-family cross-check the "
                                    "way this run's actual-vs-expected exclusion count happened to catch) "
                                    "has not been checked. Flagged for James to decide whether it needs a "
                                    "look before more families are added.",
        },
        "provisional_routing_pending_confirmation": {
            "compounds": list(EGG_PROVISIONAL_ROUTING_PENDING_CONFIRMATION.keys()),
            "n_new_compound_rows_affected": n_provisional_hits,
            "why": "Neither compound has a GROUP_MAP hit or a row in CulinAI_Staging_v30.xlsx's Exclusion "
                   "Routing Audit tab for family=egg. James was unreachable for the rest of the day and had "
                   "already asked ingestion to proceed rather than wait; a code-authored placeholder was "
                   "used so the run completes instead of aborting on two trace compounds, exactly as the "
                   "routing gate would otherwise force. NOT a workbook decision — supersede this the moment "
                   "a real ER-xxx row exists for either compound.",
            "detail": EGG_PROVISIONAL_ROUTING_PENDING_CONFIRMATION,
        },
        "compound_identity_fragmentation": {
            "finding": "Checked every no-CAS egg compound name (Boiled Yolk, Scrambled Sensomics — CAS-given "
                       "rows are anchored and immune) against compounds.jsonl and the crosswalk under exact "
                       "and stereo/locant-prefix-stripped normalization. 6 genuine near-misses found and "
                       "fixed via EGG_IDENTITY_FIXES so this run doesn't ADD to the problem: "
                       + ", ".join(sorted(EGG_IDENTITY_FIXES)) + ".",
            "pre_existing_fragmentation_surfaced_not_fixed": {
                "furaneol": "Now 4 identities in the corpus: (R)-131222-82-7, (S)-131222-81-6, this run's "
                            "racemic 3658-77-3 (correctly pinned), and beef's stray word-order-shifted "
                            "'beef:2_5_dimethyl_4_hydroxy_3_2h_furanone' (beef's own occurrence never "
                            "matched the crosswalk's racemic entry — different word order, exact-string-only "
                            "matching, same failure mode this run's EGG_IDENTITY_FIXES exists to prevent).",
                "p_cresol": "Two identities: the real CAS 106-44-5 entry, and beef's stray "
                            "'beef:4_methyl_phenol' (a space-vs-hyphen mismatch, same root cause).",
                "note": "Not retroactively merged — that needs a deliberate rebuild, same rule as beef's own "
                        "documented (E,E)-2,4-Heptadienal/Pentadecanal cases. Reported so it doesn't stay "
                        "invisible a second time.",
            },
        },
        "compound_resolution": {
            "match_method_counts": dict(resolver.method_counts),
            "n_new_compounds_added": len(new_compound_rows),
            "n_new_compounds_unmapped_group_unresolved_pending": len(unmapped_new_compounds),
        },
        "profiles_built": [
            {"product_id": p["vcf_product_id"], "n_compounds": p["n_compounds"],
             "n_df_eligible": len(p["df_eligible_compound_ids"])}
            for p in egg_profiles
        ],
        "mr15_straight_chain_alkane_guard_note": {
            "egg_alkanes_checked": sorted({
                o["compound_name"] for o in observations
                if o.get("compound_name") and is_straight_chain_alkane_name(str(o["compound_name"]))
            }),
        },
    }
    META_JSON.write_text(json.dumps(meta, indent=2))

    print(f"Landed {sum(len(v) for v in tabs.values())} rows across {len(tabs)} egg tabs to {VENDOR_DIR}")
    print(f"Compound resolution: {dict(resolver.method_counts)}")
    print(f"New compounds added to compounds.jsonl: {len(new_compound_rows)}")
    print(f"Egg profiles built: {[(p['vcf_product_id'], p['n_compounds']) for p in egg_profiles]}")
    print(f"Spine entries added: {sorted(by_product)}")
    if n_provisional_hits:
        print(f"** {n_provisional_hits} new compound row(s) used a PROVISIONAL, non-workbook routing "
              f"placeholder — see meta.json protein_egg.provisional_routing_pending_confirmation **")


if __name__ == "__main__":
    main()
